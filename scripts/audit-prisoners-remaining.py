#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl

from tidy_orchestrator.artifacts import (
    canonical_json_bytes,
    domain_digest,
    sha256_digest,
)
from tidy_orchestrator.prisoners_release import semantic_cells

ROOT = Path(__file__).resolve().parents[1]
FIX = ROOT / "fixtures" / "product-prototype"
DIRECT = ROOT / ".product-prototype" / "prisoners-remaining-phase1" / "direct"
TARGET = FIX / "prisoners-remaining-acceptance-audit-v1.json"
SCHEMA = "tidy.prisoners-remaining-acceptance-audit/v1"
MARKERS = {
    "n.a.": "not_applicable",
    "na": "not_available",
    "n.p.": "suppressed",
    "np": "suppressed",
    "..": "not_available",
}


def build() -> dict[str, Any]:
    plan = json.loads(
        (FIX / "prisoners-remaining-semantic-map-plan-v1.json").read_text()
    )
    members = []
    totals = Counter()
    workbook_cache: dict[str, Any] = {}
    semantic_cache: dict[str, dict] = {}
    for family in plan["families"]:
        family_id = family["familyId"]
        cohort = json.loads((FIX / f"prisoners-{family_id}.json").read_text())
        contract = json.loads(
            (FIX / "acceptance" / f"prisoners-{family_id}-v1.json").read_text()
        )
        for entry in cohort["workbooks"]:
            year = str(entry["year"])
            direct = json.loads((DIRECT / family_id / f"{year}.json").read_text())
            rows = direct["execution"]["tables"][0]["rows"]
            headers = [
                item["name"] for item in direct["recipe"]["tables"][0]["headers"]
            ]
            addresses = [row["_source"]["address"] for row in rows]
            keys = [tuple(str(row.get(header)) for header in headers) for row in rows]
            if len(addresses) != len(set(addresses)) or len(keys) != len(set(keys)):
                raise RuntimeError(f"non-unique source/key {family_id} {year}")
            path = entry["path"]
            if path not in workbook_cache:
                workbook_cache[path] = openpyxl.load_workbook(
                    FIX / path, data_only=False, read_only=False
                )
                semantic_cache[path] = semantic_cells(FIX / path)
            sheet = workbook_cache[path][entry["sheet"]]
            formulas = {
                cell.coordinate
                for cell in sheet._cells.values()
                if isinstance(cell.value, str) and cell.value.startswith("=")
            }
            markers = Counter(
                value
                for row in rows
                if isinstance((value := row["published value"]), str)
            )
            unknown_markers = set(markers) - set(MARKERS)
            if unknown_markers:
                raise RuntimeError(f"unknown markers {unknown_markers!r}")
            numeric = [
                row["published value"]
                for row in rows
                if isinstance(row["published value"], int | float)
                and not isinstance(row["published value"], bool)
            ]
            member = {
                "familyId": family_id,
                "year": entry["year"],
                "sheet": entry["sheet"],
                "executionPath": path,
                "executionDigest": entry["contentDigest"],
                "geometry": {
                    "maxRow": sheet.max_row,
                    "maxColumn": sheet.max_column,
                    "mergeCount": len(sheet.merged_cells.ranges),
                },
                "semanticCellCount": sum(
                    1
                    for sheet_name, _address in semantic_cache[path]
                    if sheet_name == entry["sheet"]
                ),
                "formulaCount": len(formulas),
                "selectedFormulaCount": len(formulas & set(addresses)),
                "observationCount": len(rows),
                "observedCount": len(numeric),
                "markerCounts": dict(sorted(markers.items())),
                "zeroCount": sum(value == 0 for value in numeric),
                "negativeCount": sum(value < 0 for value in numeric),
                "sourceAddressUnique": True,
                "canonicalKeyUnique": True,
                "executionWarningCount": len(direct["execution"].get("warnings", [])),
                "recipeDigest": contract["expectedRecipeDigestsByYear"][year],
                "replayMapDigest": entry["replayResponse"]["contentDigest"],
            }
            members.append(member)
            totals.update(
                members=1,
                observations=len(rows),
                observed=len(numeric),
                markers=sum(markers.values()),
                zeros=member["zeroCount"],
                negatives=member["negativeCount"],
                formulas=len(formulas),
                selectedFormulas=member["selectedFormulaCount"],
                warnings=member["executionWarningCount"],
            )
            for marker, count in markers.items():
                totals[f"marker:{marker}"] += count
    for workbook in workbook_cache.values():
        workbook.close()
    parity = []
    for source_name, bounded_name in [
        (
            "prisoners-australia-2025-national-source.xlsx",
            "prisoners-australia-2025-national-remaining-bounded.xlsx",
        ),
        (
            "prisoners-australia-2024-federal-source.xlsx",
            "prisoners-australia-2024-federal-remaining-bounded.xlsx",
        ),
        (
            "prisoners-australia-2025-federal-source.xlsx",
            "prisoners-australia-2025-federal-remaining-bounded.xlsx",
        ),
    ]:
        source = FIX / "workbooks" / source_name
        bounded = FIX / "workbooks" / bounded_name
        source_cells = semantic_cells(source)
        bounded_cells = semantic_cells(bounded)
        parity.append(
            {
                "sourcePath": f"workbooks/{source_name}",
                "sourceDigest": sha256_digest(source.read_bytes()),
                "boundedPath": f"workbooks/{bounded_name}",
                "boundedDigest": sha256_digest(bounded.read_bytes()),
                "sourceSemanticCellCount": len(source_cells),
                "boundedSemanticCellCount": len(bounded_cells),
                "coordinateValueFormulaParity": source_cells == bounded_cells,
            }
        )
    audit: dict[str, Any] = {
        "schemaVersion": SCHEMA,
        "recordedAt": "2026-08-25T12:00:00+00:00",
        "acceptanceAuthority": False,
        "trainingEligibility": False,
        "familyCount": len(plan["families"]),
        "memberCount": totals["members"],
        "observationCount": totals["observations"],
        "observedCount": totals["observed"],
        "markerCount": totals["markers"],
        "markerCounts": {
            marker: totals[f"marker:{marker}"] for marker in sorted(MARKERS)
        },
        "zeroCount": totals["zeros"],
        "negativeCount": totals["negatives"],
        "formulaCount": totals["formulas"],
        "selectedFormulaCount": totals["selectedFormulas"],
        "executionWarningCount": totals["warnings"],
        "safeTotalPolicy": "no-equations-no-inference",
        "boundedWorkbookParity": parity,
        "members": members,
    }
    audit["auditDigest"] = domain_digest(SCHEMA, audit)
    return audit


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    audit = build()
    rendered = canonical_json_bytes(audit) + b"\n"
    if args.check:
        if TARGET.read_bytes() != rendered:
            raise SystemExit("remaining Prisoners audit drift")
    else:
        TARGET.write_bytes(rendered)
    print(
        json.dumps(
            {
                "families": audit["familyCount"],
                "members": audit["memberCount"],
                "observations": audit["observationCount"],
                "zeros": audit["zeroCount"],
                "markers": audit["markerCount"],
                "digest": audit["auditDigest"],
            },
            separators=(",", ":"),
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
