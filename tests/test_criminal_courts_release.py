from __future__ import annotations

import copy
import csv
import json
import re
import shutil
import subprocess
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, range_boundaries

import tidy_orchestrator.product_prototype as product_prototype_module
from tidy_orchestrator.artifacts import domain_digest, sha256_digest
from tidy_orchestrator.criminal_courts_release import (
    CriminalCourtsReleaseError,
    build_family_membership,
    build_source_inventory,
    verify_criminal_courts_release,
)
from tidy_orchestrator.large_batch import LargeBatchSpec, verify_large_batch_evidence

PROJECT = Path(__file__).parents[1]
FIXTURES = PROJECT / "fixtures" / "product-prototype"
DOWNLOADS = FIXTURES / "criminal-courts-release-downloads-v1.json"
CROSSWALK = FIXTURES / "criminal-courts-release-family-crosswalk-v1.json"
INVENTORY = FIXTURES / "criminal-courts-release-source-inventory-v1.json"
MEMBERSHIP = FIXTURES / "criminal-courts-release-family-membership-v1.json"
SENTENCE_FINE_FAMILIES = (
    "criminal-courts-guilty-outcome-principal-sentence-by-length-and-court-level",
    "criminal-courts-custody-by-offence-sentence-length-all-courts-anzsoc-2011",
    "criminal-courts-custody-by-offence-sentence-length-all-courts-anzsoc-2023",
    "criminal-courts-custody-by-offence-sentence-length-higher-courts-anzsoc-2011",
    "criminal-courts-custody-by-offence-sentence-length-higher-courts-anzsoc-2023",
    "criminal-courts-custody-by-offence-sentence-length-magistrates-courts-anzsoc-2011",
    "criminal-courts-custody-by-offence-sentence-length-magistrates-courts-anzsoc-2023",
    "criminal-courts-custody-by-offence-sentence-length-childrens-courts-anzsoc-2011",
    "criminal-courts-custody-by-offence-sentence-length-childrens-courts-anzsoc-2023",
    "criminal-courts-community-service-work-by-offence-length-court-jurisdiction-anzsoc-2011",
    "criminal-courts-community-service-work-by-offence-length-court-jurisdiction-anzsoc-2023",
    "criminal-courts-fines-by-offence-amount-court-jurisdiction-anzsoc-2011",
    "criminal-courts-fines-by-offence-amount-court-jurisdiction-anzsoc-2023",
    "criminal-courts-custody-by-offence-indigenous-status-length-jurisdiction-anzsoc-2011",
)
INDIGENOUS_STATUS_FAMILIES = (
    "criminal-courts-custody-by-offence-indigenous-status-length-jurisdiction-anzsoc-2023",
    "criminal-courts-main-defendants-finalised-excluding-traffic-offences-summary-characteristics-by-indigenous-status-selected-sta-b7c24101f2",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-traffic-offences-court-level-by-indigenous-status-selected-s-95ed08966c",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-traffic-offences-summary-characteristics-by-indigenous-statu-eb8b64429d",
    "criminal-courts-main-defendants-with-a-guilty-outcome-excluding-traffic-offences-principal-sentence-and-age-by-indigenous-stat-904a1a9407",
    "criminal-courts-main-rate-of-defendants-finalised-excluding-transfers-and-traffic-offences-crude-and-age-standardised-by-selec-0ec1bf61ab",
)
YOUTH_FAMILIES = (
    "criminal-courts-youth-summary-characteristics-australia",
    "criminal-courts-youth-summary-characteristics-by-jurisdiction",
    "criminal-courts-youth-finalised-sex-age-by-principal-offence",
    "criminal-courts-youth-guilty-outcome-offence-by-principal-sentence",
    "criminal-courts-youth-guilty-outcome-sex-age-by-principal-sentence",
    "criminal-courts-youth-indigenous-status-by-jurisdiction",
    "criminal-courts-youth-indigenous-status-age-by-jurisdiction",
    "criminal-courts-youth-guilty-outcome-sex-age-by-principal-offence",
    "criminal-courts-youth-guilty-outcome-sentence-length-by-jurisdiction",
)
PRELIMINARY_ANZSOC_2023_FAMILIES = (
    "criminal-courts-preliminary-anzsoc-2023-defendants-finalised-excluding-transfers-and-traffic-offences-preliminary-anzsoc-2023--9011820c8c",
    "criminal-courts-preliminary-anzsoc-2023-defendants-finalised-excluding-transfers-preliminary-anzsoc-2023-principal-offence-by--73da8de2bb",
    "criminal-courts-preliminary-anzsoc-2023-defendants-finalised-excluding-transfers-preliminary-anzsoc-2023-principal-offence-sta-44f35a79b0",
    "criminal-courts-preliminary-anzsoc-2023-defendants-finalised-excluding-transfers-sex-and-age-by-preliminary-anzsoc-2023-princi-e9e65af772",
    "criminal-courts-preliminary-anzsoc-2023-defendants-finalised-preliminary-anzsoc-2023-principal-offence-by-method-of-finalisati-1ae417f119",
    "criminal-courts-preliminary-anzsoc-2023-defendants-with-a-guilty-outcome-sex-and-preliminary-anzsoc-2023-principal-offence-by--b7e20f2c51",
)
NEW_SOUTH_WALES_FAMILIES = (
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-new-s-4d8d4a9753",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-b264f70ae9",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--08d5fe90c2",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--3dc8642f72",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-06556eaa3c",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-ff440749dc",
    "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-new-south-wales-be08f0a014",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-new-south-wales-275c7f7671",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-new-south-wales-and-99870bae7c",
)
VICTORIA_FAMILIES = (
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-victo-05e49f77ef",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-9118c3e02d",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--c83b6750b7",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--c86993e818",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-39801494e2",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-173f3bac6d",
    "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-victoria-3594164858",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-victoria-and-4a552df2a2",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-victoria-b85a0d076d",
)
QUEENSLAND_FAMILIES = (
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-queen-3ce9f383e3",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-7ae5115676",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--f8da01fb54",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--0841d7f5ad",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-f847eba164",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-57f4983e16",
    "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-queensland-bef784f7d8",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-queensland-7022b11692",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-queensland-and-2925b8b521",
)
AUSTRALIAN_CAPITAL_TERRITORY_FAMILIES = (
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-austr-4200e414a2",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-79856bc0b9",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--0b4eef6926",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--7c61d6c40e",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-6c72c5eba2",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-139aaf92e0",
    "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-australian-capital-territory-5705af16d6",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-australian-capital-territory-1f84e9447d",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-australian-capital-territory-and-b377949ac0",
)
TASMANIA_FAMILIES = (
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-tasma-1e1718730a",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-f2593de546",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--cdc489d600",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--34dbc091f5",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-08dd480268",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-0a5e590f31",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-0cd32616d1",
    "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-tasmania-1d97a0925b",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-tasmania-and-4a82019ceb",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-tasmania-d897254a26",
)
NORTHERN_TERRITORY_FAMILIES = (
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-north-9e1eae4d24",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-2dc791882d",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--dff73e6680",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--d4b9910476",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-fb8ea665a2",
    "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-7b3957ee89",
    "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-northern-territory-2e3b1c96f5",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-northern-territory-27c9d31040",
    "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-northern-territory-and-cd3b98cdfb",
)
GUILTY_OUTCOME_FAMILIES = (
    "criminal-courts-guilty-outcome-summary-by-jurisdiction",
    "criminal-courts-guilty-outcome-sex-age-by-sentence-and-court-level",
    "criminal-courts-guilty-outcome-sex-age-by-sentence-all-courts",
    "criminal-courts-guilty-outcome-sex-age-by-sentence-higher-courts",
    "criminal-courts-guilty-outcome-sex-age-by-sentence-magistrates-courts",
    "criminal-courts-guilty-outcome-sex-age-by-sentence-childrens-courts",
    "criminal-courts-guilty-outcome-offence-by-sentence",
    "criminal-courts-guilty-outcome-offence-by-duration",
)


def _load(path: Path) -> dict[str, object]:
    return json.loads(path.read_text())


def _copy_release_project(target: Path) -> None:
    fixture_root = target / "fixtures" / "product-prototype"
    (fixture_root / "workbooks").mkdir(parents=True)
    downloads = _load(DOWNLOADS)
    for declaration in downloads["downloads"]:
        relative = declaration["path"]
        shutil.copyfile(FIXTURES / relative, fixture_root / relative)
    for path in (DOWNLOADS, CROSSWALK, INVENTORY, MEMBERSHIP):
        shutil.copyfile(path, fixture_root / path.name)
    for path in (
        FIXTURES / "data-asset-status-v1.json",
        FIXTURES / "batch-workbook-normalization-v1.json",
    ):
        shutil.copyfile(path, fixture_root / path.name)
    status = _load(FIXTURES / "data-asset-status-v1.json")
    for declaration in status["cohorts"]:
        cohort = PROJECT / declaration["cohortPath"]
        if cohort.is_file():
            destination = target / declaration["cohortPath"]
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(cohort, destination)


def test_release_verifier_proves_complete_four_release_custody() -> None:
    report = verify_criminal_courts_release(PROJECT)

    assert report == {
        "verified": True,
        "releaseCount": 4,
        "downloadCount": 69,
        "reviewedExclusionDownloadCount": 4,
        "substantiveCubeCount": 65,
        "numberedDataSheetCount": 430,
        "familyCount": 198,
        "registeredMemberCount": 374,
        "pendingSemanticContractCount": 56,
        "providerCalls": 0,
        "inventoryDigest": report["inventoryDigest"],
        "membershipDigest": report["membershipDigest"],
    }
    inventory = _load(INVENTORY)
    assert inventory["numberedDataSheetCountsByRelease"] == {
        "2021-22": 94,
        "2022-23": 102,
        "2023-24": 116,
        "2024-25": 118,
    }
    downloads = inventory["downloads"]
    assert sum(item["kind"] == "guide" for item in downloads) == 4
    assert sum(item["kind"] == "cube" for item in downloads) == 65
    assert any(
        sheet["name"].startswith("FDV Table ")
        and sheet["classification"] == "numbered-data"
        for item in downloads
        for sheet in item["sheets"]
    )
    assert any(
        sheet["name"].startswith("ANZSOC 2023 Table ")
        and sheet["classification"] == "numbered-data"
        for item in downloads
        for sheet in item["sheets"]
    )
    membership = _load(MEMBERSHIP)
    assert Counter(
        member["cubeId"]
        for family in membership["families"]
        for member in family["members"]
        if not member["registered"]
    ) == {
        "family-and-domestic-violence-offences-experimental-data": 39,
        "family-and-domestic-violence-offences-australia-experimental-data": 9,
        (
            "family-and-domestic-violence-offences-states-and-territories-"
            "experimental-data"
        ): 8,
    }


def test_generated_inventory_and_membership_are_byte_reproducible() -> None:
    inventory = build_source_inventory(PROJECT)
    membership = build_family_membership(PROJECT, inventory)

    assert json.loads(INVENTORY.read_text()) == inventory
    assert json.loads(MEMBERSHIP.read_text()) == membership
    assert sum(len(family["members"]) for family in membership["families"]) == 430
    assert (
        sum(
            member["registered"]
            for family in membership["families"]
            for member in family["members"]
        )
        == 374
    )
    wa_mixed = next(
        family
        for family in membership["families"]
        if family["familyId"].endswith("western-australia-and-be8fa3884d")
    )
    assert wa_mixed["semanticTitle"] == (
        "Defendants finalised, summary characteristics by court level — "
        "mixed concorded history — Western Australia"
    )
    nt_mixed = next(
        family
        for family in membership["families"]
        if family["familyId"].endswith("northern-territory-and-cd3b98cdfb")
    )
    assert nt_mixed["semanticTitle"] == (
        "Defendants finalised, summary characteristics by court level — "
        "mixed concorded history — Northern Territory"
    )
    act_mixed = next(
        family
        for family in membership["families"]
        if family["familyId"].endswith("australian-capital-territory-and-b377949ac0")
    )
    assert act_mixed["semanticTitle"] == (
        "Defendants finalised, summary characteristics by court level — "
        "mixed concorded history — Australian Capital Territory"
    )


def test_generator_check_and_cli_are_cwd_independent(tmp_path: Path) -> None:
    generated = subprocess.run(
        [
            str(PROJECT / "scripts" / "generate-criminal-courts-release-inventory.py"),
            "--check",
        ],
        cwd=tmp_path,
        check=True,
        capture_output=True,
        text=True,
    )
    assert generated.stdout == ""
    assert generated.stderr == ""

    cli = subprocess.run(
        [str(PROJECT / "scripts" / "tidy-criminal-courts-release"), "verify"],
        cwd=tmp_path,
        check=True,
        capture_output=True,
        text=True,
    )
    assert json.loads(cli.stdout)["registeredMemberCount"] == 374
    assert json.loads(cli.stdout)["providerCalls"] == 0

    registered_cli = subprocess.run(
        [
            "uv",
            "run",
            "--project",
            str(PROJECT),
            "tidy-criminal-courts-release",
            "verify",
        ],
        cwd=tmp_path,
        check=True,
        capture_output=True,
        text=True,
    )
    assert json.loads(registered_cli.stdout)["numberedDataSheetCount"] == 430


def test_guilty_outcome_cluster_preserves_semantics_and_provenance() -> None:
    membership = _load(MEMBERSHIP)
    families = {family["familyId"]: family for family in membership["families"]}
    assert (
        sum(
            len(families[family_id]["members"]) for family_id in GUILTY_OUTCOME_FAMILIES
        )
        == 19
    )
    assert all(
        all(member["registered"] for member in families[family_id]["members"])
        for family_id in GUILTY_OUTCOME_FAMILIES
    )

    rows: list[dict[str, object]] = []
    value_status_counts: dict[str, int] = {}
    for family_id in GUILTY_OUTCOME_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["acceptedWorkbookCount"] == len(run["workbooks"])
        assert all(
            item["decision"] == "prototype_auto_accepted" for item in run["workbooks"]
        )
        for status, count in manifest["valueStatusCounts"].items():
            value_status_counts[status] = value_status_counts.get(status, 0) + count
        by_asset: dict[tuple[str, str], set[str]] = {}
        for row in family_rows:
            key = (row["source_workbook_digest"], row["source_sheet"])
            assert row["source_cell"] not in by_asset.setdefault(key, set())
            by_asset[key].add(row["source_cell"])
            assert row["publication_vintage_date"]
            assert row["raw_value"] is not None
        rows.extend(family_rows)

    assert len(rows) == 25602
    assert value_status_counts == {
        "not_applicable": 3,
        "observed": 25493,
        "suppressed": 106,
    }
    assert {row["unit_id"] for row in rows} == {"person", "week", "year"}
    assert {row["measure_id"] for row in rows if row["unit_id"] == "week"} == {
        "defendant-mean-duration",
        "defendant-median-duration",
        "sentence-mean-duration",
        "sentence-median-duration",
    }
    assert all(
        row["raw_value"] in {"..", "na", "n.a.", "np", "n.p."}
        for row in rows
        if row["value_status"] != "observed"
    )

    offence_rows = [row for row in rows if "principal_offence_id" in row]
    assert all(
        row["principal_offence_id"].startswith(("ANZSOC_2011_", "OFFENCE_"))
        for row in offence_rows
        if row["publication_vintage_date"] < "2025-01-01"
    )
    assert all(
        row["principal_offence_id"].startswith(("ANZSOC_2023_", "OFFENCE_"))
        for row in offence_rows
        if row["publication_vintage_date"] >= "2025-01-01"
    )
    split_rows = [
        row
        for row in rows
        if row["source_sheet"] in {"Table 10", "Table 11", "Table 12", "Table 13"}
        and row["publication_vintage_date"] == "2025-06-30"
        and "court_level_id" in row
    ]
    assert {row["court_level_id"] for row in split_rows} == {
        "ALL_COURTS",
        "CHILDRENS_COURTS",
        "HIGHER_COURTS",
        "MAGISTRATES_COURTS",
    }
    assert all(
        str(row["raw_court_level"]).startswith("This table contains")
        for row in split_rows
    )

    for family_id in (
        "criminal-courts-guilty-outcome-offence-by-sentence",
        "criminal-courts-guilty-outcome-offence-by-duration",
    ):
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        assert {rule["dimension"] for rule in contract["allowedExecutionWarnings"]} == {
            "court_level",
            "observation_period",
        }


def test_sentence_and_fine_cluster_preserves_classification_and_measures() -> None:
    membership = _load(MEMBERSHIP)
    families = {family["familyId"]: family for family in membership["families"]}
    assert (
        sum(len(families[family_id]["members"]) for family_id in SENTENCE_FINE_FAMILIES)
        == 31
    )
    assert all(
        all(member["registered"] for member in families[family_id]["members"])
        for family_id in SENTENCE_FINE_FAMILIES
    )
    rows: list[dict[str, object]] = []
    measure_counts: dict[str, int] = {}
    value_status_counts: dict[str, int] = {}
    source_cells: set[tuple[str, str, str]] = set()
    for family_id in SENTENCE_FINE_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["acceptedWorkbookCount"] == len(run["workbooks"])
        assert all(
            item["decision"] == "prototype_auto_accepted" for item in run["workbooks"]
        )
        for measure, count in manifest["measureCounts"].items():
            measure_counts[measure] = measure_counts.get(measure, 0) + count
        for status, count in manifest["valueStatusCounts"].items():
            value_status_counts[status] = value_status_counts.get(status, 0) + count
        for row in family_rows:
            source_key = (
                row["source_workbook_digest"],
                row["source_sheet"],
                row["source_cell"],
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["publication_vintage_date"]
            assert row["raw_value"] is not None
            if family_id.endswith("anzsoc-2011") and "principal_offence_id" in row:
                assert row["principal_offence_id"].startswith(
                    ("ANZSOC_2011_", "OFFENCE_TOTAL")
                )
            if family_id.endswith("anzsoc-2023") and "principal_offence_id" in row:
                assert row["principal_offence_id"].startswith(
                    ("ANZSOC_2023_", "OFFENCE_TOTAL")
                )
            if row["measure_id"] == "defendant-proportion":
                assert row["principal_offence_id"] == "OFFENCE_TOTAL"
            if "fine_amount_id" in row:
                assert row["fine_amount_id"].startswith("FINE_")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        if family_id == SENTENCE_FINE_FAMILIES[0]:
            assert contract["allowedExecutionWarnings"] == []
        else:
            assert len(contract["allowedExecutionWarnings"]) == 1
            warning_rule = contract["allowedExecutionWarnings"][0]
            assert warning_rule["dimension"] == "jurisdiction"
            assert warning_rule["requireCanonicalOutputEquivalence"] is True
            assert warning_rule["expectedHeaderSourcesByYear"]
        rows.extend(family_rows)

    assert len(rows) == 43284
    assert measure_counts == {
        "defendant-count": 25844,
        "defendant-proportion": 2622,
        "fine-mean-amount": 2103,
        "fine-median-amount": 2103,
        "sentence-mean-duration": 5306,
        "sentence-median-duration": 5306,
    }
    assert value_status_counts == {
        "not_applicable": 1962,
        "observed": 40646,
        "suppressed": 676,
    }
    assert {(row["measure_id"], row["unit_id"]) for row in rows} == {
        ("defendant-count", "person"),
        ("defendant-proportion", "percent"),
        ("fine-mean-amount", "dollar"),
        ("fine-median-amount", "dollar"),
        ("sentence-mean-duration", "hour"),
        ("sentence-mean-duration", "month"),
        ("sentence-median-duration", "hour"),
        ("sentence-median-duration", "month"),
    }


def test_indigenous_status_cluster_preserves_periods_rates_and_null_markers() -> None:
    membership = _load(MEMBERSHIP)
    families = {family["familyId"]: family for family in membership["families"]}
    assert (
        sum(
            len(families[family_id]["members"])
            for family_id in INDIGENOUS_STATUS_FAMILIES
        )
        == 17
    )
    assert all(
        all(member["registered"] for member in families[family_id]["members"])
        for family_id in INDIGENOUS_STATUS_FAMILIES
    )

    rows: list[dict[str, object]] = []
    measure_counts: dict[str, int] = {}
    value_status_counts: dict[str, int] = {}
    source_cells: set[tuple[str, str, str]] = set()
    warning_dimensions: dict[str, set[str]] = {}
    for family_id in INDIGENOUS_STATUS_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        uses_observation_period = (
            contract.get("referenceDateDimension") == "observation_period"
        )
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["acceptedWorkbookCount"] == len(run["workbooks"])
        assert all(
            item["decision"] == "prototype_auto_accepted" for item in run["workbooks"]
        )
        for measure, count in manifest["measureCounts"].items():
            measure_counts[measure] = measure_counts.get(measure, 0) + count
        for status, count in manifest["valueStatusCounts"].items():
            value_status_counts[status] = value_status_counts.get(status, 0) + count
        for row in family_rows:
            source_key = (
                row["source_workbook_digest"],
                row["source_sheet"],
                row["source_cell"],
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["publication_vintage_date"]
            assert row["raw_value"] is not None
            if uses_observation_period:
                assert "observation_period_id" in row
                assert row["reference_date"] == row["observation_period_id"]
        if uses_observation_period:
            assert any(
                row["publication_vintage_date"] != row["reference_date"]
                for row in family_rows
            )
        warning_dimensions[family_id] = {
            rule["dimension"] for rule in contract["allowedExecutionWarnings"]
        }
        assert all(
            rule["requireCanonicalOutputEquivalence"] is True
            and rule["expectedHeaderSourcesByYear"]
            for rule in contract["allowedExecutionWarnings"]
        )
        rows.extend(family_rows)

    assert len(rows) == 24643
    assert measure_counts == {
        "age-standardised-defendant-rate": 382,
        "crude-defendant-rate": 393,
        "defendant-count": 21018,
        "defendant-proportion": 144,
        "indigenous-to-non-indigenous-rate-ratio": 382,
        "mean-defendant-age": 856,
        "median-defendant-age": 856,
        "sentence-mean-duration": 306,
        "sentence-median-duration": 306,
    }
    assert value_status_counts == {
        "not_applicable": 318,
        "not_available": 794,
        "observed": 23479,
        "suppressed": 52,
    }
    assert {
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    } == {
        ("..", "not_applicable"),
        ("na", "not_available"),
        ("np", "suppressed"),
    }
    assert {(row["measure_id"], row["unit_id"]) for row in rows} == {
        ("age-standardised-defendant-rate", "per-100000-persons-aged-10-plus"),
        ("crude-defendant-rate", "per-100000-persons-aged-10-plus"),
        ("defendant-count", "person"),
        ("defendant-proportion", "percent"),
        ("indigenous-to-non-indigenous-rate-ratio", "ratio"),
        ("mean-defendant-age", "year"),
        ("median-defendant-age", "year"),
        ("sentence-mean-duration", "month"),
        ("sentence-median-duration", "month"),
    }
    offence_rows = [row for row in rows if "principal_offence_id" in row]
    assert all(
        row["principal_offence_id"].startswith(("ANZSOC_2023_", "OFFENCE_TOTAL"))
        for row in offence_rows
    )
    assert {
        row["characteristic_group_id"]
        for row in rows
        if "characteristic_group_id" in row
        and str(row["characteristic_group_id"]).startswith("GROUP_PRINCIPAL_OFFENCE")
    } == {
        "GROUP_PRINCIPAL_OFFENCE_ANZSOC_2011",
        "GROUP_PRINCIPAL_OFFENCE_ANZSOC_2023",
    }
    for family_id, dimensions in warning_dimensions.items():
        assert dimensions == (
            {"observation_period"}
            if family_id
            in {
                INDIGENOUS_STATUS_FAMILIES[1],
                INDIGENOUS_STATUS_FAMILIES[3],
                INDIGENOUS_STATUS_FAMILIES[4],
            }
            else {"jurisdiction"}
        )


def test_youth_cluster_preserves_classification_periods_measures_and_markers() -> None:
    membership = _load(MEMBERSHIP)
    youth_members = [
        member
        for family in membership["families"]
        for member in family["members"]
        if member["cubeId"] == "youth-defendants-australia"
    ]
    assert len(youth_members) == 24
    assert all(member["registered"] for member in youth_members)

    rows: list[dict[str, object]] = []
    measure_counts: dict[str, int] = {}
    value_status_counts: dict[str, int] = {}
    source_cells: set[tuple[str, str, str]] = set()
    warning_dimensions: dict[str, set[str]] = {}
    for family_id in YOUTH_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["acceptedWorkbookCount"] == len(run["workbooks"])
        assert all(
            item["decision"] == "prototype_auto_accepted" for item in run["workbooks"]
        )
        for measure, count in manifest["measureCounts"].items():
            measure_counts[measure] = measure_counts.get(measure, 0) + count
        for status, count in manifest["valueStatusCounts"].items():
            value_status_counts[status] = value_status_counts.get(status, 0) + count
        uses_observation_period = (
            contract.get("referenceDateDimension") == "observation_period"
        )
        for row in family_rows:
            source_key = (
                row["source_workbook_digest"],
                row["source_sheet"],
                row["source_cell"],
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["publication_vintage_date"]
            assert row["raw_value"] is not None
            if uses_observation_period:
                assert row["reference_date"] == row["observation_period_id"]
            if "principal_offence_id" in row:
                expected_prefix = (
                    "ANZSOC_2023_"
                    if row["publication_vintage_date"] == "2025-06-30"
                    else "ANZSOC_2011_"
                )
                assert row["principal_offence_id"].startswith(
                    (expected_prefix, "OFFENCE_TOTAL")
                )
        if uses_observation_period:
            assert any(
                row["publication_vintage_date"] != row["reference_date"]
                for row in family_rows
            )
        warning_dimensions[family_id] = {
            rule["dimension"] for rule in contract["allowedExecutionWarnings"]
        }
        assert all(
            rule["requireCanonicalOutputEquivalence"] is True
            and rule["expectedHeaderSourcesByYear"]
            for rule in contract["allowedExecutionWarnings"]
        )
        rows.extend(family_rows)

    assert len(rows) == 10302
    assert measure_counts == {
        "defendant-count": 9762,
        "mean-case-duration": 69,
        "mean-defendant-age": 69,
        "median-case-duration": 69,
        "median-defendant-age": 69,
        "sentence-mean-duration": 132,
        "sentence-median-duration": 132,
    }
    assert value_status_counts == {
        "not_applicable": 2,
        "not_available": 447,
        "observed": 9847,
        "suppressed": 6,
    }
    assert {
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    } == {
        ("..", "not_applicable"),
        ("na", "not_available"),
        ("np", "suppressed"),
    }
    assert {(row["measure_id"], row["unit_id"]) for row in rows} == {
        ("defendant-count", "person"),
        ("mean-case-duration", "week"),
        ("mean-defendant-age", "year"),
        ("median-case-duration", "week"),
        ("median-defendant-age", "year"),
        ("sentence-mean-duration", "month"),
        ("sentence-median-duration", "month"),
    }
    assert {
        row["characteristic_group_id"]
        for row in rows
        if str(row.get("characteristic_group_id", "")).startswith(
            "GROUP_PRINCIPAL_OFFENCE"
        )
    } == {
        "GROUP_PRINCIPAL_OFFENCE_ANZSOC_2011",
        "GROUP_PRINCIPAL_OFFENCE_ANZSOC_2023",
        "GROUP_PRINCIPAL_OFFENCE_ANZSOC_2023_WITH_CONCORDED_ANZSOC_2011_SERIES",
    }
    assert "NON_INDIGENOUS_AND_NOT_STATED" in {
        row.get("indigenous_status_id") for row in rows
    }
    assert {"CHAR_10_13_YEARS", "CHAR_14_17_YEARS"} <= {
        row.get("characteristic_category_id") for row in rows
    }
    sentence_length_csv = FIXTURES / (
        "criminal-courts-youth-guilty-outcome-sentence-length-by-jurisdiction-"
        "evidence/canonical-observations.csv"
    )
    with sentence_length_csv.open(newline="") as handle:
        assert "Table 80 " in {row["source_sheet"] for row in csv.DictReader(handle)}
    expected_warning_dimensions = {
        YOUTH_FAMILIES[2]: {"sex"},
        YOUTH_FAMILIES[4]: {"sex"},
        YOUTH_FAMILIES[5]: {"jurisdiction"},
        YOUTH_FAMILIES[6]: {"jurisdiction"},
        YOUTH_FAMILIES[7]: {"sex"},
    }
    assert warning_dimensions == {
        family_id: expected_warning_dimensions.get(family_id, set())
        for family_id in YOUTH_FAMILIES
    }


def test_preliminary_cluster_preserves_dual_classification_custody() -> None:
    membership = _load(MEMBERSHIP)
    preliminary_members = [
        member
        for family in membership["families"]
        for member in family["members"]
        if member["cubeId"] == "preliminary-anzsoc-2023-principal-offence"
    ]
    assert len(preliminary_members) == 6
    assert all(member["registered"] for member in preliminary_members)
    assert {member["classificationContext"] for member in preliminary_members} == {
        "preliminary-anzsoc-2023"
    }

    source = load_workbook(
        FIXTURES / "workbooks/criminal-courts-2023-24-cube-15-source.xlsx",
        data_only=False,
        read_only=True,
    )
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str]] = set()
    warning_dimensions: dict[str, set[str]] = {}
    for family_id in PRELIMINARY_ANZSOC_2023_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        cohort = _load(FIXTURES / f"{family_id}.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        replay = cohort["workbooks"][0]["replayResponse"]

        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert run["acceptedWorkbookCount"] == 1
        assert run["workbooks"][0]["decision"] == "prototype_auto_accepted"
        assert replay["acceptanceAuthority"] is False
        assert contract["trainingEligibility"] is False
        assert contract["totalValidation"] == "not_applicable"
        assert contract["totalEquations"] == []
        warning_dimensions[family_id] = {
            rule["dimension"] for rule in contract["allowedExecutionWarnings"]
        }

        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", row["source_cell"])
            assert match is not None
            source_key = (row["source_sheet"], row["source_cell"])
            assert source_key not in source_cells
            source_cells.add(source_key)
            cell = source[row["source_sheet"]].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"] == row["value"]
            assert row["publication_vintage_date"] == "2024-06-30"
            assert row["reference_date"] == "2024-06-30"
            assert row["measure_id"] == "defendant-count"
            assert row["unit_id"] == "person"
            assert row["value_status"] == "observed"
            assert str(row["principal_offence_id"]).startswith("PRELIM_ANZSOC_2023_")
        rows.extend(family_rows)

    assert len(rows) == 2371
    assert sum(row["value"] == 0 for row in rows) == 204
    assert {row["source_sheet"] for row in rows} == {
        f"ANZSOC 2023 Table {table}" for table in range(1, 7)
    }
    concordance_rows = [
        row for row in rows if row["source_sheet"] == "ANZSOC 2023 Table 6"
    ]
    assert len(concordance_rows) == 306
    assert all(
        str(row["principal_offence_anzsoc_2011_id"]).startswith("ANZSOC_2011_")
        for row in concordance_rows
    )
    assert all(
        "principal_offence_anzsoc_2011_id" not in row
        for row in rows
        if row["source_sheet"] != "ANZSOC 2023 Table 6"
    )
    method_family = PRELIMINARY_ANZSOC_2023_FAMILIES[4]
    assert warning_dimensions == {
        family_id: ({"method_of_finalisation"} if family_id == method_family else set())
        for family_id in PRELIMINARY_ANZSOC_2023_FAMILIES
    }
    method_contract = _load(FIXTURES / "acceptance" / f"{method_family}-v1.json")
    warning_rule = method_contract["allowedExecutionWarnings"][0]
    assert warning_rule["requireCanonicalOutputEquivalence"] is True
    assert {
        header_source
        for sources in warning_rule["expectedHeaderSourcesByYear"]["2023"].values()
        for header_source in sources
    } == {"R5C2", "R24C2", "R43C2", "R62C2", "R81C2", "R100C2"}


def test_new_south_wales_cluster_preserves_source_and_classification() -> None:
    membership = _load(MEMBERSHIP)
    members = [
        member
        for family in membership["families"]
        for member in family["members"]
        if member["cubeId"] == "defendants-finalised-new-south-wales"
    ]
    assert len(members) == 22
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 5,
        "2022-23": 5,
        "2023-24": 6,
        "2024-25": 6,
    }
    assert all(member["registered"] for member in members)

    source_books = {
        release: load_workbook(
            FIXTURES
            / next(
                member["sourcePath"]
                for member in members
                if member["releaseId"] == release
            ),
            data_only=False,
            read_only=False,
        )
        for release in {member["releaseId"] for member in members}
    }
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str, str]] = set()
    warning_dimensions: set[str] = set()
    warning_count = 0
    for family_id in NEW_SOUTH_WALES_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert contract["totalValidation"] == "not_applicable"
        assert contract["totalEquations"] == []
        assert (
            manifest["warningCountsByYear"] == contract["expectedWarningCountsByYear"]
        )
        observed_warning_counts = {
            str(workbook["year"]): workbook["executionWarningCount"]
            for workbook in run["workbooks"]
        }
        assert observed_warning_counts == contract["expectedWarningCountsByYear"]
        warning_count += sum(observed_warning_counts.values())
        warning_dimensions.update(
            rule["dimension"] for rule in contract["allowedExecutionWarnings"]
        )
        assert all(
            rule["requireCanonicalOutputEquivalence"] is True
            and set(rule["expectedHeaderSourcesByYear"])
            == {str(year) for year in manifest["manualReplayYears"]}
            for rule in contract["allowedExecutionWarnings"]
        )

        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", str(row["source_cell"]))
            assert match is not None
            release_start = int(str(row["publication_vintage_date"])[:4]) - 1
            release = f"{release_start}-{str(release_start + 1)[-2:]}"
            cell = source_books[release][str(row["source_sheet"])].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"]
            source_key = (
                str(row["source_workbook_digest"]),
                str(row["source_sheet"]),
                str(row["source_cell"]),
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["reference_date"] == row["observation_period_id"]
            assert row["jurisdiction_id"] == "NSW"
        rows.extend(family_rows)

    assert len(rows) == 17_691
    assert len(source_cells) == len(rows)
    assert Counter(str(row["measure_id"]) for row in rows) == {
        "defendant-count": 16_827,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter(str(row["value_status"]) for row in rows) == {
        "observed": 17_578,
        "not_applicable": 61,
        "not_available": 52,
    }
    assert sum(row["value"] == 0 for row in rows) == 926
    assert {str(row["classification_context_id"]) for row in rows} == {
        "ANZSOC_2011",
        "ANZSOC_2023",
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
    }
    old_offences = {
        str(row["principal_offence_id"])
        for row in rows
        if row.get("classification_context_id") == "ANZSOC_2011"
        and "principal_offence_id" in row
    }
    new_offences = {
        str(row["principal_offence_id"])
        for row in rows
        if row.get("classification_context_id") == "ANZSOC_2023"
        and "principal_offence_id" in row
    }
    assert old_offences & new_offences
    assert warning_dimensions == {"court_level", "observation_period"}
    assert warning_count == 10_807
    assert any(row["publication_vintage_date"] != row["reference_date"] for row in rows)
    assert "Table 19 " in {str(row["source_sheet"]) for row in rows}
    with (
        FIXTURES / f"{NEW_SOUTH_WALES_FAMILIES[5]}-evidence/canonical-observations.csv"
    ).open(newline="") as handle:
        assert "Table 19 " in {row["source_sheet"] for row in csv.DictReader(handle)}


def test_victoria_cluster_preserves_source_and_classification() -> None:
    membership = _load(MEMBERSHIP)
    members = [
        member
        for family in membership["families"]
        for member in family["members"]
        if member["cubeId"] == "defendants-finalised-victoria"
    ]
    assert len(members) == 22
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 5,
        "2022-23": 5,
        "2023-24": 6,
        "2024-25": 6,
    }
    assert all(member["registered"] for member in members)

    source_books = {
        release: load_workbook(
            FIXTURES
            / next(
                member["sourcePath"]
                for member in members
                if member["releaseId"] == release
            ),
            data_only=False,
            read_only=False,
        )
        for release in {member["releaseId"] for member in members}
    }
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str, str]] = set()
    warning_dimensions: set[str] = set()
    warning_header_sources: dict[str, set[str]] = {}
    warning_count = 0
    for family_id in VICTORIA_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert contract["totalValidation"] == "not_applicable"
        assert contract["totalEquations"] == []
        assert (
            manifest["warningCountsByYear"] == contract["expectedWarningCountsByYear"]
        )
        observed_warning_counts = {
            str(workbook["year"]): workbook["executionWarningCount"]
            for workbook in run["workbooks"]
        }
        assert observed_warning_counts == contract["expectedWarningCountsByYear"]
        warning_count += sum(observed_warning_counts.values())
        warning_dimensions.update(
            rule["dimension"] for rule in contract["allowedExecutionWarnings"]
        )
        assert contract["trainingEligibility"] is False
        cohort = _load(FIXTURES / f"{family_id}.json")
        assert all(
            workbook["replayResponse"]["acceptanceAuthority"] is False
            for workbook in cohort["workbooks"]
        )
        for rule in contract["allowedExecutionWarnings"]:
            assert rule["requireCanonicalOutputEquivalence"] is True
            assert set(rule["expectedHeaderSourcesByYear"]) == {
                str(year) for year in manifest["manualReplayYears"]
            }
            sources = warning_header_sources.setdefault(rule["dimension"], set())
            sources.update(
                source
                for aliases in rule["expectedHeaderSourcesByYear"].values()
                for bound_sources in aliases.values()
                for source in bound_sources
            )

        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", str(row["source_cell"]))
            assert match is not None
            release_start = int(str(row["publication_vintage_date"])[:4]) - 1
            release = f"{release_start}-{str(release_start + 1)[-2:]}"
            cell = source_books[release][str(row["source_sheet"])].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"]
            source_key = (
                str(row["source_workbook_digest"]),
                str(row["source_sheet"]),
                str(row["source_cell"]),
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["reference_date"] == row["observation_period_id"]
            assert row["jurisdiction_id"] == "VIC"
        rows.extend(family_rows)

    assert len(rows) == 17_524
    assert len(source_cells) == len(rows)
    assert Counter(str(row["measure_id"]) for row in rows) == {
        "defendant-count": 16_660,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter(str(row["value_status"]) for row in rows) == {
        "observed": 17_411,
        "not_applicable": 61,
        "not_available": 52,
    }
    assert {
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    } == {("..", "not_applicable"), ("na", "not_available")}
    assert sum(row["value"] == 0 for row in rows) == 1_419
    assert {str(row["classification_context_id"]) for row in rows} == {
        "ANZSOC_2011",
        "ANZSOC_2023",
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
    }
    old_offences = {
        str(row["principal_offence_id"])
        for row in rows
        if row.get("classification_context_id") == "ANZSOC_2011"
        and "principal_offence_id" in row
    }
    new_offences = {
        str(row["principal_offence_id"])
        for row in rows
        if row.get("classification_context_id") == "ANZSOC_2023"
        and "principal_offence_id" in row
    }
    assert old_offences & new_offences
    assert warning_dimensions == {"court_level", "observation_period"}
    assert warning_header_sources == {
        "court_level": {
            "R5C2",
            "R6C2",
            "R55C2",
            "R61C2",
            "R62C2",
            "R104C2",
            "R117C2",
            "R118C2",
            "R153C2",
            "R173C2",
            "R174C2",
        },
        "observation_period": {
            "R5C2",
            "R6C2",
            "R28C2",
            "R29C2",
            "R30C2",
            "R31C2",
        },
    }
    assert warning_count == 10_720
    assert any(row["publication_vintage_date"] != row["reference_date"] for row in rows)
    assert "Table 32" in {str(row["source_sheet"]) for row in rows}
    with (
        FIXTURES / f"{VICTORIA_FAMILIES[3]}-evidence/canonical-observations.csv"
    ).open(newline="") as handle:
        assert "Table 32" in {row["source_sheet"] for row in csv.DictReader(handle)}


def test_queensland_cluster_preserves_source_classification_and_warnings() -> None:
    membership = _load(MEMBERSHIP)
    members = [
        member
        for family in membership["families"]
        for member in family["members"]
        if member["cubeId"] == "defendants-finalised-queensland"
    ]
    assert len(members) == 22
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 5,
        "2022-23": 5,
        "2023-24": 6,
        "2024-25": 6,
    }
    assert all(member["registered"] for member in members)

    source_books = {
        release: load_workbook(
            FIXTURES
            / next(
                member["sourcePath"]
                for member in members
                if member["releaseId"] == release
            ),
            data_only=False,
            read_only=False,
        )
        for release in {member["releaseId"] for member in members}
    }
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str, str]] = set()
    warning_dimensions: set[str] = set()
    warning_header_sources: dict[str, set[str]] = {}
    warning_count = 0
    for family_id in QUEENSLAND_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        cohort = _load(FIXTURES / f"{family_id}.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert manifest["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert contract["totalValidation"] == "not_applicable"
        assert contract["totalEquations"] == []
        assert contract["trainingEligibility"] is False
        assert all(
            workbook["replayResponse"]["acceptanceAuthority"] is False
            for workbook in cohort["workbooks"]
        )
        observed_warning_counts = {
            str(workbook["year"]): workbook["executionWarningCount"]
            for workbook in run["workbooks"]
        }
        assert observed_warning_counts == contract["expectedWarningCountsByYear"]
        assert observed_warning_counts == manifest["warningCountsByYear"]
        warning_count += sum(observed_warning_counts.values())
        warning_dimensions.update(
            rule["dimension"] for rule in contract["allowedExecutionWarnings"]
        )
        for rule in contract["allowedExecutionWarnings"]:
            assert rule["requireCanonicalOutputEquivalence"] is True
            assert set(rule["expectedHeaderSourcesByYear"]) == {
                str(year) for year in manifest["manualReplayYears"]
            }
            warning_header_sources.setdefault(rule["dimension"], set()).update(
                source
                for aliases in rule["expectedHeaderSourcesByYear"].values()
                for bound_sources in aliases.values()
                for source in bound_sources
            )

        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", str(row["source_cell"]))
            assert match is not None
            release_start = int(str(row["publication_vintage_date"])[:4]) - 1
            release = f"{release_start}-{str(release_start + 1)[-2:]}"
            cell = source_books[release][str(row["source_sheet"])].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"]
            source_key = (
                str(row["source_workbook_digest"]),
                str(row["source_sheet"]),
                str(row["source_cell"]),
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["reference_date"] == row["observation_period_id"]
            assert row["jurisdiction_id"] == "QLD"
        rows.extend(family_rows)

    assert len(rows) == 17_415
    assert len(source_cells) == len(rows)
    assert Counter(str(row["measure_id"]) for row in rows) == {
        "defendant-count": 16_551,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter(str(row["value_status"]) for row in rows) == {
        "observed": 17_162,
        "not_available": 136,
        "not_applicable": 117,
    }
    assert {
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    } == {("..", "not_applicable"), ("na", "not_available")}
    assert sum(row["value"] == 0 for row in rows) == 986
    assert {str(row["classification_context_id"]) for row in rows} == {
        "ANZSOC_2011",
        "ANZSOC_2023",
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
    }
    assert warning_dimensions == {"court_level", "observation_period"}
    assert warning_header_sources == {
        "court_level": {
            "R5C2",
            "R6C2",
            "R55C2",
            "R61C2",
            "R62C2",
            "R104C2",
            "R117C2",
            "R118C2",
            "R153C2",
            "R173C2",
            "R174C2",
        },
        "observation_period": {
            "R5C2",
            "R6C2",
            "R28C2",
            "R29C2",
            "R30C2",
            "R31C2",
        },
    }
    assert warning_count == 10_640
    assert any(row["publication_vintage_date"] != row["reference_date"] for row in rows)


def test_australian_capital_territory_aliases_are_exhaustively_source_bound() -> None:
    def normalize(value: str) -> str:
        return " ".join(value.strip().split())

    aliases: dict[tuple[str, str], str] = {}
    represented: set[tuple[str, str]] = set()
    aggregate_alias_count = 0
    for family_id in AUSTRALIAN_CAPITAL_TERRITORY_FAMILIES:
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        for dimension, dimension_aliases in contract["aliases"].items():
            aggregate_alias_count += len(dimension_aliases)
            for raw, target in dimension_aliases.items():
                identity = (dimension, normalize(raw))
                previous = aliases.setdefault(identity, target)
                assert previous == target, f"normalized alias collision: {identity}"
        rows = json.loads(
            (FIXTURES / f"{family_id}-evidence/canonical-observations.json").read_text()
        )
        for row in rows:
            for dimension in contract["requiredDimensions"]:
                raw_field = f"raw_{dimension}"
                identity = (dimension, normalize(str(row[raw_field])))
                represented.add(identity)
                target_field = product_prototype_module._DIMENSION_FIELDS[dimension]
                assert row[target_field] == aliases[identity]

    assert aggregate_alias_count == 630
    workbook_strings: set[str] = set()
    for release in ("2021-22", "2022-23", "2023-24", "2024-25"):
        for kind in ("source", "normalized"):
            path = (
                FIXTURES
                / "workbooks"
                / f"criminal-courts-{release}-cube-11-{kind}.xlsx"
            )
            workbook = load_workbook(path, data_only=False, read_only=True)
            try:
                workbook_strings.update(
                    normalize(cell.value)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str)
                )
            finally:
                workbook.close()
    assert {raw for _dimension, raw in aliases} <= workbook_strings
    assert set(aliases) - represented == set()


def test_australian_capital_territory_cluster_preserves_exact_semantics() -> None:
    membership = _load(MEMBERSHIP)
    families = {
        family["familyId"]: family
        for family in membership["families"]
        if family["familyId"] in AUSTRALIAN_CAPITAL_TERRITORY_FAMILIES
    }
    assert set(families) == set(AUSTRALIAN_CAPITAL_TERRITORY_FAMILIES)
    members = [member for family in families.values() for member in family["members"]]
    assert len(members) == 22
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 5,
        "2022-23": 5,
        "2023-24": 6,
        "2024-25": 6,
    }
    assert all(member["registered"] for member in members)
    assert {member["cubeId"] for member in members} == {
        "defendants-finalised-australian-capital-territory"
    }
    assert all(
        "Australian Capital Territory" in member["publishedTitle"] for member in members
    )
    assert families[AUSTRALIAN_CAPITAL_TERRITORY_FAMILIES[-1]]["semanticTitle"] == (
        "Defendants finalised, summary characteristics by court level — "
        "mixed concorded history — Australian Capital Territory"
    )

    workbooks = {
        release: load_workbook(
            FIXTURES
            / "workbooks"
            / f"criminal-courts-{release}-cube-11-normalized.xlsx",
            data_only=False,
            read_only=False,
        )
        for release in ("2021-22", "2022-23", "2023-24", "2024-25")
    }
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str, str]] = set()
    warning_count = 0
    for family_id in AUSTRALIAN_CAPITAL_TERRITORY_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        cohort = _load(FIXTURES / f"{family_id}.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
        assert contract["trainingEligibility"] is False
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert manifest["providerCalls"] == run["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["historicalReplayIsAcceptanceAuthority"] is False
        assert run["trainingEligibility"] is False
        assert all(
            workbook["replayResponse"]["acceptanceAuthority"] is False
            for workbook in cohort["workbooks"]
        )
        observed_warnings = {
            str(workbook["year"]): workbook["executionWarningCount"]
            for workbook in run["workbooks"]
        }
        assert observed_warnings == contract["expectedWarningCountsByYear"]
        assert observed_warnings == manifest["warningCountsByYear"]
        warning_count += sum(observed_warnings.values())
        decisions = {
            (workbook["workbookDigest"], workbook["sheet"]): workbook["decisionId"]
            for workbook in run["workbooks"]
        }
        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", str(row["source_cell"]))
            assert match is not None
            release_start = int(str(row["publication_vintage_date"])[:4]) - 1
            release = f"{release_start}-{str(release_start + 1)[-2:]}"
            cell = workbooks[release][str(row["source_sheet"])].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"]
            assert cell.data_type != "f"
            key = (
                str(row["source_workbook_digest"]),
                str(row["source_sheet"]),
                str(row["source_cell"]),
            )
            assert key not in source_cells
            source_cells.add(key)
            assert row["acceptance_decision_digest"] == decisions[key[:2]]
            assert row["reference_date"] == row["observation_period_id"]
            assert row["jurisdiction_id"] == "ACT"
        rows.extend(family_rows)

    assert len(rows) == len(source_cells) == 16_057
    assert Counter(str(row["measure_id"]) for row in rows) == {
        "defendant-count": 15_193,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter(str(row["value_status"]) for row in rows) == {
        "observed": 15_920,
        "not_available": 88,
        "not_applicable": 49,
    }
    assert {
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    } == {("..", "not_applicable"), ("na", "not_available")}
    assert sum(row["value"] == 0 for row in rows) == 2_844
    assert Counter(str(row["classification_context_id"]) for row in rows) == {
        "ANZSOC_2011": 11_283,
        "ANZSOC_2023": 2_104,
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023": 2_670,
    }
    assert warning_count == 9_969
    method_ids = {
        "CHAR_GUILTY_EX_PARTE",
        "CHAR_TRANSFER_TO_OTHER_COURT_LEVELS",
        "CHAR_WITHDRAWN_BY_PROSECUTION",
        "CHAR_TOTAL_FINALISED",
    }
    assert all(
        row["characteristic_group_id"] == "GROUP_METHOD_OF_FINALISATION"
        for row in rows
        if row.get("characteristic_category_id") in method_ids
    )
    assert not any(
        row.get("characteristic_group_id") == "GROUP_GUILTY_EX_PARTE" for row in rows
    )


def test_tasmania_aliases_are_exhaustively_source_bound() -> None:
    def normalize(value: str) -> str:
        return " ".join(value.strip().split())

    aliases: dict[tuple[str, str], str] = {}
    represented: set[tuple[str, str]] = set()
    aggregate_alias_count = 0
    for family_id in TASMANIA_FAMILIES:
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        for dimension, dimension_aliases in contract["aliases"].items():
            aggregate_alias_count += len(dimension_aliases)
            for raw, target in dimension_aliases.items():
                identity = (dimension, normalize(raw))
                previous = aliases.setdefault(identity, target)
                assert previous == target, f"normalized alias collision: {identity}"
        rows = json.loads(
            (FIXTURES / f"{family_id}-evidence/canonical-observations.json").read_text()
        )
        for row in rows:
            for dimension in contract["requiredDimensions"]:
                identity = (dimension, normalize(str(row[f"raw_{dimension}"])))
                represented.add(identity)
                assert (
                    row[product_prototype_module._DIMENSION_FIELDS[dimension]]
                    == (aliases[identity])
                )

    assert aggregate_alias_count == 759
    workbook_strings: set[str] = set()
    for release in ("2021-22", "2022-23", "2023-24", "2024-25"):
        for kind in ("source", "normalized"):
            workbook = load_workbook(
                FIXTURES
                / "workbooks"
                / f"criminal-courts-{release}-cube-9-{kind}.xlsx",
                data_only=False,
                read_only=True,
            )
            try:
                workbook_strings.update(
                    normalize(cell.value)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str)
                )
            finally:
                workbook.close()
    assert {raw for _dimension, raw in aliases} <= workbook_strings
    assert set(aliases) - represented == set()


def test_tasmania_cluster_preserves_exact_semantics() -> None:
    membership = _load(MEMBERSHIP)
    families = {
        family["familyId"]: family
        for family in membership["families"]
        if family["familyId"] in TASMANIA_FAMILIES
    }
    assert set(families) == set(TASMANIA_FAMILIES)
    members = [member for family in families.values() for member in family["members"]]
    assert len(members) == 22
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 5,
        "2022-23": 5,
        "2023-24": 6,
        "2024-25": 6,
    }
    assert all(member["registered"] for member in members)
    assert {member["cubeId"] for member in members} == {"defendants-finalised-tasmania"}
    legacy = families[TASMANIA_FAMILIES[5]]["members"]
    proper = families[TASMANIA_FAMILIES[6]]["members"]
    assert [item["releaseId"] for item in legacy] == ["2021-22", "2022-23"]
    en_dash = "\N{EN DASH}"
    assert all(
        f"Magistrates' Courts {en_dash}Tasmania" in item["publishedTitle"]
        for item in legacy
    )
    assert [item["releaseId"] for item in proper] == ["2023-24"]
    assert f"Magistrates' Courts {en_dash} Tasmania" in proper[0]["publishedTitle"]
    assert families[TASMANIA_FAMILIES[8]]["semanticTitle"] == (
        "Defendants finalised, summary characteristics by court level — "
        "mixed concorded history — Tasmania"
    )
    assert (
        "concorded from ANZSOC 2011"
        in families[TASMANIA_FAMILIES[8]]["members"][0]["publishedTitle"]
    )

    workbooks = {
        release: load_workbook(
            FIXTURES
            / "workbooks"
            / f"criminal-courts-{release}-cube-9-normalized.xlsx",
            data_only=False,
            read_only=False,
        )
        for release in ("2021-22", "2022-23", "2023-24", "2024-25")
    }
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str, str]] = set()
    warning_count = 0
    for family_id in TASMANIA_FAMILIES:
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        cohort = _load(FIXTURES / f"{family_id}.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
        assert contract["trainingEligibility"] is False
        assert contract["totalEquations"] == []
        assert contract["totalValidation"] == "not_applicable"
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert manifest["providerCalls"] == run["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["historicalReplayIsAcceptanceAuthority"] is False
        assert all(
            workbook["replayResponse"]["acceptanceAuthority"] is False
            for workbook in cohort["workbooks"]
        )
        observed_warnings = {
            str(workbook["year"]): workbook["executionWarningCount"]
            for workbook in run["workbooks"]
        }
        assert observed_warnings == contract["expectedWarningCountsByYear"]
        assert observed_warnings == manifest["warningCountsByYear"]
        warning_count += sum(observed_warnings.values())
        decisions = {
            (workbook["workbookDigest"], workbook["sheet"]): workbook["decisionId"]
            for workbook in run["workbooks"]
        }
        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", str(row["source_cell"]))
            assert match is not None
            release_start = int(str(row["publication_vintage_date"])[:4]) - 1
            release = f"{release_start}-{str(release_start + 1)[-2:]}"
            cell = workbooks[release][str(row["source_sheet"])].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"]
            assert cell.data_type != "f"
            key = (
                str(row["source_workbook_digest"]),
                str(row["source_sheet"]),
                str(row["source_cell"]),
            )
            assert key not in source_cells
            source_cells.add(key)
            assert row["acceptance_decision_digest"] == decisions[key[:2]]
            assert row["reference_date"] == row["observation_period_id"]
            assert row["jurisdiction_id"] == "TAS"
        rows.extend(family_rows)

    assert len(rows) == len(source_cells) == 16_545
    assert Counter(str(row["measure_id"]) for row in rows) == {
        "defendant-count": 15_681,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter((str(row["measure_id"]), str(row["unit_id"])) for row in rows) == {
        ("defendant-count", "person"): 15_681,
        ("mean-case-duration", "week"): 216,
        ("mean-defendant-age", "year"): 216,
        ("median-case-duration", "week"): 216,
        ("median-defendant-age", "year"): 216,
    }
    assert Counter(str(row["value_status"]) for row in rows) == {
        "observed": 16_324,
        "suppressed": 129,
        "not_applicable": 53,
        "not_available": 39,
    }
    assert Counter(
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    ) == {
        ("np", "suppressed"): 129,
        ("..", "not_applicable"): 53,
        ("na", "not_available"): 39,
    }
    assert sum(row["value"] == 0 for row in rows) == 2_342
    assert Counter(str(row["classification_context_id"]) for row in rows) == {
        "ANZSOC_2011": 11_622,
        "ANZSOC_2023": 2_268,
        "ANZSOC_2023_WITH_CONCORDED_ANZSOC_2011_SERIES": 2_655,
    }
    assert warning_count == 10_231
    assert Counter(
        (row["raw_principal_offence"], row["principal_offence_id"])
        for row in rows
        if "regulaton" in str(row.get("raw_principal_offence", "")).lower()
    ) == {
        (
            "163 Commercial/industry/financial regulaton",
            "OFFENCE_163_COMMERCIAL_INDUSTRY_FINANCIAL_REGULATION",
        ): 6
    }
    for row in rows:
        raw = str(row.get("raw_characteristic_category", "")).lower()
        if "excluding transfer" not in raw and any(
            term in raw for term in ("guilty ex-parte", "transfer", "withdraw")
        ):
            assert row["characteristic_group_id"] == "GROUP_METHOD_OF_FINALISATION"
        if raw.startswith("total finalised") and "excluding transfer" not in raw:
            assert row["characteristic_group_id"] == "GROUP_METHOD_OF_FINALISATION"
    assert all(
        row["characteristic_group_id"].startswith("GROUP_PRINCIPAL_OFFENCE")
        for row in rows
        if str(row.get("raw_characteristic_category", "")).startswith(
            "Total finalised (excluding transfer"
        )
    )


def test_northern_territory_aliases_are_exhaustively_source_bound() -> None:
    def normalize(value: str) -> str:
        return " ".join(value.strip().split())

    aliases: dict[tuple[str, str], str] = {}
    aggregate_alias_count = 0
    represented_aliases: set[tuple[str, str]] = set()
    for family_id in NORTHERN_TERRITORY_FAMILIES:
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        for dimension, dimension_aliases in contract["aliases"].items():
            aggregate_alias_count += len(dimension_aliases)
            for raw, target in dimension_aliases.items():
                identity = (dimension, normalize(raw))
                previous = aliases.setdefault(identity, target)
                assert previous == target, f"normalized alias collision: {identity}"

        rows = json.loads(
            (FIXTURES / f"{family_id}-evidence/canonical-observations.json").read_text()
        )
        for row in rows:
            for dimension in contract["requiredDimensions"]:
                raw_field = f"raw_{dimension}"
                assert raw_field in row
                identity = (dimension, normalize(str(row[raw_field])))
                represented_aliases.add(identity)
                assert identity in aliases
                target_field = product_prototype_module._DIMENSION_FIELDS[dimension]
                assert row[target_field] == aliases[identity]

    assert aggregate_alias_count == 750

    workbook_paths = {
        FIXTURES / "workbooks" / f"criminal-courts-{release}-cube-10-{kind}.xlsx"
        for release in ("2021-22", "2022-23", "2023-24", "2024-25")
        for kind in ("source", "normalized")
    }
    assert len(workbook_paths) == 8
    workbook_strings: set[str] = set()
    for path in workbook_paths:
        workbook = load_workbook(path, data_only=False, read_only=True)
        try:
            workbook_strings.update(
                normalize(cell.value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str)
            )
        finally:
            workbook.close()
    assert {raw for _dimension, raw in aliases} <= workbook_strings

    # No NT alias is justified only by a row excluded from canonical evidence. If
    # that changes, the excluded-row source cells must be added as a separate
    # authority rather than silently treating the contract as self-authorizing.
    source_only_excluded_aliases = set(aliases) - represented_aliases
    assert source_only_excluded_aliases == set()


def test_northern_territory_cluster_preserves_exact_source_semantics() -> None:
    membership = _load(MEMBERSHIP)
    nt_families = {
        family["familyId"]: family
        for family in membership["families"]
        if family["familyId"] in NORTHERN_TERRITORY_FAMILIES
    }
    assert set(nt_families) == set(NORTHERN_TERRITORY_FAMILIES)
    members = [
        member for family in nt_families.values() for member in family["members"]
    ]
    assert len(members) == 22
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 5,
        "2022-23": 5,
        "2023-24": 6,
        "2024-25": 6,
    }
    assert all(member["registered"] for member in members)
    assert {member["cubeId"] for member in members} == {
        "defendants-finalised-northern-territory"
    }
    assert all("Northern Territory" in member["publishedTitle"] for member in members)
    assert nt_families[NORTHERN_TERRITORY_FAMILIES[-1]]["semanticTitle"] == (
        "Defendants finalised, summary characteristics by court level — "
        "mixed concorded history — Northern Territory"
    )

    expected_warning_counts = (
        {"2024": 342},
        {"2024": 342},
        {"2021": 306, "2022": 323, "2023": 323},
        {"2021": 144, "2022": 152, "2023": 152, "2024": 168},
        {"2021": 85, "2022": 90, "2023": 95, "2024": 95},
        {"2021": 306, "2022": 323, "2023": 323},
        {"2023": 0, "2024": 0},
        {"2021": 1308, "2022": 1690, "2023": 1862},
        {"2024": 1995},
    )
    source_books = {
        release: load_workbook(
            FIXTURES
            / next(
                member["sourcePath"]
                for member in members
                if member["releaseId"] == release
            ),
            data_only=False,
            read_only=False,
        )
        for release in {member["releaseId"] for member in members}
    }
    rows: list[dict[str, object]] = []
    source_cells: set[tuple[str, str, str]] = set()
    warning_sources: dict[str, set[str]] = {}
    warning_count = 0
    for family_id, expected_counts in zip(
        NORTHERN_TERRITORY_FAMILIES, expected_warning_counts, strict=True
    ):
        evidence = FIXTURES / f"{family_id}-evidence"
        manifest = _load(evidence / "manifest.json")
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        cohort = _load(FIXTURES / f"{family_id}.json")
        run = _load(evidence / "run.json")
        family_rows = json.loads((evidence / "canonical-observations.json").read_text())
        assert manifest["canonicalObservationCount"] == len(family_rows)
        assert manifest["providerCalls"] == run["providerCalls"] == 0
        assert manifest["exceptionWorkbookCount"] == 0
        assert run["historicalReplayIsAcceptanceAuthority"] is False
        assert all(
            workbook["decision"] == "prototype_auto_accepted"
            for workbook in run["workbooks"]
        )
        assert all(
            workbook["replayResponse"]["acceptanceAuthority"] is False
            for workbook in cohort["workbooks"]
        )
        observed_warning_counts = {
            str(workbook["year"]): workbook["executionWarningCount"]
            for workbook in run["workbooks"]
        }
        assert observed_warning_counts == expected_counts
        assert observed_warning_counts == contract["expectedWarningCountsByYear"]
        assert observed_warning_counts == manifest["warningCountsByYear"]
        warning_count += sum(observed_warning_counts.values())
        for rule in contract["allowedExecutionWarnings"]:
            assert rule["code"] == "AMBIGUOUS_HEADER"
            assert rule["requireCanonicalOutputEquivalence"] is True
            assert set(rule["expectedHeaderSourcesByYear"]) == set(expected_counts)
            warning_sources.setdefault(rule["dimension"], set()).update(
                source
                for resolved_headers in rule["expectedHeaderSourcesByYear"].values()
                for bound_sources in resolved_headers.values()
                for source in bound_sources
            )

        decisions = {
            (workbook["workbookDigest"], workbook["sheet"]): workbook["decisionId"]
            for workbook in run["workbooks"]
        }
        for row in family_rows:
            match = re.fullmatch(r"R([0-9]+)C([0-9]+)", str(row["source_cell"]))
            assert match is not None
            release_start = int(str(row["publication_vintage_date"])[:4]) - 1
            release = f"{release_start}-{str(release_start + 1)[-2:]}"
            cell = source_books[release][str(row["source_sheet"])].cell(
                row=int(match.group(1)), column=int(match.group(2))
            )
            assert cell.value == row["raw_value"]
            assert cell.data_type != "f"
            source_key = (
                str(row["source_workbook_digest"]),
                str(row["source_sheet"]),
                str(row["source_cell"]),
            )
            assert source_key not in source_cells
            source_cells.add(source_key)
            assert row["acceptance_decision_digest"] == decisions[source_key[:2]]
            assert row["reference_date"] == row["observation_period_id"]
            assert row["jurisdiction_id"] == "NT"
        rows.extend(family_rows)

    assert len(rows) == len(source_cells) == 16_931
    assert Counter(str(row["measure_id"]) for row in rows) == {
        "defendant-count": 16_067,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter(str(row["value_status"]) for row in rows) == {
        "observed": 16_794,
        "not_available": 88,
        "not_applicable": 49,
    }
    assert {
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    } == {("..", "not_applicable"), ("na", "not_available")}
    assert sum(row["value"] == 0 for row in rows) == 2_913
    assert {str(row["classification_context_id"]) for row in rows} == {
        "ANZSOC_2011",
        "ANZSOC_2023",
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
    }
    assert warning_count == 10_424
    assert warning_sources == {
        "court_level": {
            "R55C2",
            "R61C2",
            "R62C2",
            "R104C2",
            "R117C2",
            "R118C2",
            "R153C2",
            "R173C2",
            "R174C2",
        },
        "observation_period": {"R28C2", "R29C2", "R30C2", "R31C2"},
    }

    nbsp_offence = "14\u00a0Offences against justice procedures and orders(c)"
    assert source_books["2024-25"]["Table 59"]["O5"].value == nbsp_offence
    assert any(row.get("raw_principal_offence") == nbsp_offence for row in rows)
    all_courts_contract = _load(
        FIXTURES / "acceptance" / f"{NORTHERN_TERRITORY_FAMILIES[0]}-v1.json"
    )
    assert (
        all_courts_contract["aliases"]["principal_offence"][nbsp_offence]
        == "OFFENCE_14_OFFENCES_AGAINST_JUSTICE_PROCEDURES_AND_ORDERS"
    )
    mixed_contract = _load(
        FIXTURES / "acceptance" / f"{NORTHERN_TERRITORY_FAMILIES[-1]}-v1.json"
    )
    mixed_sheet = source_books["2024-25"]["Table 57"]
    mixed_variants = {
        "A17": ("01 Homicide(e)", "CHAR_01_HOMICIDE"),
        "A20": (
            "04 Harm or endanger persons(f)",
            "CHAR_04_HARM_OR_ENDANGER_PERSONS",
        ),
        "A29": (
            "13 Traffic and vehicle offences(g)",
            "CHAR_13_TRAFFIC_AND_VEHICLE_OFFENCES",
        ),
        "A30": (
            "14 Offences against justice procedures and orders(h)",
            "CHAR_14_OFFENCES_AGAINST_JUSTICE_PROCEDURES_AND_ORDERS",
        ),
        "A34": (
            "Total finalised (excluding transfer to other court levels)(i)",
            "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
        ),
        "A37": ("Mean (weeks)(k)", "CHAR_MEAN_WEEKS"),
        "A38": ("Median (weeks)(k)", "CHAR_MEDIAN_WEEKS"),
        "A55": ("Community service / work(t)", "CHAR_COMMUNITY_SERVICE_WORK"),
        "A56": (
            "Moderate penalty in the community(t)",
            "CHAR_MODERATE_PENALTY_IN_THE_COMMUNITY",
        ),
        "A57": ("Monetary penalties(u)", "CHAR_MONETARY_PENALTIES"),
        "A58": ("Fines(g)", "CHAR_FINES"),
        "A59": (
            "Good behaviour (incl. bonds)(t)",
            "CHAR_GOOD_BEHAVIOUR_INCL_BONDS",
        ),
        "A61": ("Total guilty outcome(v)", "CHAR_TOTAL_GUILTY_OUTCOME"),
    }
    for cell, (source_label, canonical_label) in mixed_variants.items():
        assert mixed_sheet[cell].value == source_label
        assert (
            mixed_contract["aliases"]["characteristic_category"][source_label]
            == canonical_label
        )
    assert mixed_sheet["A16"].value == "Principal offence (ANZSOC 2023)(d)"
    assert (
        mixed_contract["aliases"]["characteristic_group"][
            "Principal offence (ANZSOC 2023)(d)"
        ]
        == "GROUP_PRINCIPAL_OFFENCE_ANZSOC_2023_WITH_CONCORDED_ANZSOC_2011_SERIES"
    )
    assert mixed_sheet["A36"].value == "Duration(j)"
    assert (
        mixed_contract["aliases"]["characteristic_group"]["Duration(j)"]
        == "GROUP_DURATION"
    )
    assert not any(
        row.get("characteristic_group_id") == "GROUP_GUILTY_EX_PARTE" for row in rows
    )
    method_rows = [
        row
        for row in rows
        if row.get("characteristic_group_id") == "GROUP_METHOD_OF_FINALISATION"
    ]
    assert len(method_rows) == 5_372
    assert len({str(row["characteristic_category_id"]) for row in method_rows}) == 10
    assert (
        sum(
            row.get("characteristic_category_id") == "CHAR_GUILTY_EX_PARTE"
            for row in method_rows
        )
        == 551
    )
    assert Counter(
        str(row["method_of_finalisation_id"])
        for row in rows
        if "method_of_finalisation_id" in row
    ) == {
        "METHOD_ACQUITTED": 105,
        "METHOD_GUILTY_OUTCOME": 105,
        "METHOD_TOTAL": 105,
        "METHOD_TOTAL_ADJUDICATED": 105,
        "METHOD_TRANSFER_TO_OTHER_COURT_LEVELS": 105,
        "METHOD_WITHDRAWN_BY_PROSECUTION": 105,
    }


def test_download_digest_mutation_fails_closed(tmp_path: Path) -> None:
    _copy_release_project(tmp_path)
    path = tmp_path / "fixtures" / "product-prototype" / DOWNLOADS.name
    downloads = _load(path)
    changed = copy.deepcopy(downloads)
    changed["downloads"][0]["contentDigest"] = "sha256:" + "0" * 64
    path.write_text(json.dumps(changed))

    with pytest.raises(CriminalCourtsReleaseError, match="bytes differ"):
        build_source_inventory(tmp_path)


def test_incomplete_or_retitled_family_cover_fails_closed(tmp_path: Path) -> None:
    _copy_release_project(tmp_path)
    path = tmp_path / "fixtures" / "product-prototype" / CROSSWALK.name
    crosswalk = _load(path)
    incomplete = copy.deepcopy(crosswalk)
    multi_release = next(
        family for family in incomplete["families"] if len(family["members"]) > 1
    )
    multi_release["members"].pop()
    path.write_text(json.dumps(incomplete))
    inventory = build_source_inventory(tmp_path)
    with pytest.raises(CriminalCourtsReleaseError, match="exact cover"):
        build_family_membership(tmp_path, inventory)

    retitled = copy.deepcopy(crosswalk)
    retitled["families"][0]["members"][0]["publishedTitle"] += " fabricated"
    path.write_text(json.dumps(retitled))
    with pytest.raises(CriminalCourtsReleaseError, match="exact source identity"):
        build_family_membership(tmp_path, inventory)


def test_classification_context_is_exactly_custodied(tmp_path: Path) -> None:
    _copy_release_project(tmp_path)
    path = tmp_path / "fixtures" / "product-prototype" / CROSSWALK.name
    crosswalk = _load(path)
    changed = copy.deepcopy(crosswalk)
    changed["families"][0]["members"][0]["classificationContext"] = "generic-anzsoc"
    path.write_text(json.dumps(changed))
    inventory = build_source_inventory(tmp_path)
    with pytest.raises(CriminalCourtsReleaseError, match="classification context"):
        build_family_membership(tmp_path, inventory)


DEFENDANT_RATE_CUBE_FAMILIES = (
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-7f1899e604",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-and-04a7eabfc9",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-australian-capital-ter-5b71625ad1",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-australian-capital-ter-deef593a0e",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-new-south-wales-and-be5f5e8837",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-new-south-wales-cc3f4ad25c",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-northern-territory-3407ddb4e5",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-northern-territory-and-88fd1c42eb",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-queensland-and-c2ba08e992",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-queensland-ca54310c6d",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-south-australia-291078ff14",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-south-australia-and-12cc29caf3",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-tasmania-05e6aa34d2",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-tasmania-and-a561865464",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-victoria-and-94dc135eb2",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-victoria-e342ba2cd4",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-western-australia-680a53f26b",
    "criminal-courts-main-defendants-finalised-excluding-transfers-and-organisations-summary-characteristics-western-australia-and-fddeb560ad",
)

# These expectations are deliberately literal and independent of the generated
# contracts and evidence.  They freeze the source custody matrix that Cube 12
# is permitted to represent.
DEFENDANT_RATE_SOURCE_MATRIX = {
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-7f1899e604"
    ): (
        "AUS",
        (2021, 2022, 2023),
        ("Table 56", "Table 56", "Table 64"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-and-04a7eabfc9"
    ): (
        "AUS",
        (2024,),
        ("Table 69",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-australian-capital-ter-5b71625ad1"
    ): (
        "ACT",
        (2021, 2022, 2023),
        ("Table 64", "Table 64", "Table 72"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-australian-capital-ter-deef593a0e"
    ): (
        "ACT",
        (2024,),
        ("Table 77",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-new-south-wales-and-be5f5e8837"
    ): (
        "NSW",
        (2024,),
        ("Table 70",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-new-south-wales-cc3f4ad25c"
    ): (
        "NSW",
        (2021, 2022, 2023),
        ("Table 57", "Table 57", "Table 65"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-northern-territory-3407ddb4e5"
    ): (
        "NT",
        (2021, 2022, 2023),
        ("Table 63", "Table 63", "Table 71"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-northern-territory-and-88fd1c42eb"
    ): (
        "NT",
        (2024,),
        ("Table 76",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-queensland-and-c2ba08e992"
    ): (
        "QLD",
        (2024,),
        ("Table 72",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-queensland-ca54310c6d"
    ): (
        "QLD",
        (2021, 2022, 2023),
        ("Table 59", "Table 59", "Table 67"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-south-australia-291078ff14"
    ): (
        "SA",
        (2021, 2022, 2023),
        ("Table 60", "Table 60", "Table 68"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-south-australia-and-12cc29caf3"
    ): (
        "SA",
        (2024,),
        ("Table 73",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-tasmania-05e6aa34d2"
    ): (
        "TAS",
        (2021, 2022, 2023),
        ("Table 62", "Table 62", "Table 70"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-tasmania-and-a561865464"
    ): (
        "TAS",
        (2024,),
        ("Table 75",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-victoria-and-94dc135eb2"
    ): (
        "VIC",
        (2024,),
        ("Table 71",),
        (870,),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-victoria-e342ba2cd4"
    ): (
        "VIC",
        (2021, 2022, 2023),
        ("Table 58", "Table 58", "Table 66"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-western-australia-680a53f26b"
    ): (
        "WA",
        (2021, 2022, 2023),
        ("Table 61", "Table 61", "Table 69"),
        (672, 728, 784),
    ),
    (
        "criminal-courts-main-defendants-finalised-excluding-transfers-and-"
        "organisations-summary-characteristics-western-australia-and-fddeb560ad"
    ): (
        "WA",
        (2024,),
        ("Table 74",),
        (870,),
    ),
}
DEFENDANT_RATE_SOURCE_FILES = {
    2021: (
        "workbooks/criminal-courts-2021-22-cube-12-source.xlsx",
        "sha256:5d7734110d0f6348c017e4b3ec2fad4118b95de47cb48b6e7b2cf51f4f57bcee",
        "sha256:5d7734110d0f6348c017e4b3ec2fad4118b95de47cb48b6e7b2cf51f4f57bcee",
    ),
    2022: (
        "workbooks/criminal-courts-2022-23-cube-12-source.xlsx",
        "sha256:4412ad57f73fe112e3c8d546bbd348f720aa72cd5abb422d54e5ee19fdd445ec",
        "sha256:4412ad57f73fe112e3c8d546bbd348f720aa72cd5abb422d54e5ee19fdd445ec",
    ),
    2023: (
        "workbooks/criminal-courts-2023-24-cube-12-normalized.xlsx",
        "sha256:047a4e140fbddb5e48210be32c2fc0dc0baf8cfa1d2d181fed73d66ba3b67cc0",
        "sha256:1c800c30cf50594a0ece895882981cf109d36601cafcfcaaa9fd0aaece38d0f6",
    ),
    2024: (
        "workbooks/criminal-courts-2024-25-cube-12-source.xlsx",
        "sha256:f0411a3531d5c74ef68f24c6a9a57edfc24fdbf5d95faab0a8c73690f57bc8de",
        "sha256:f0411a3531d5c74ef68f24c6a9a57edfc24fdbf5d95faab0a8c73690f57bc8de",
    ),
}
DEFENDANT_RATE_COMMON_CATEGORIES = {
    "GROUP_SEX": ("CHAR_MALES", "CHAR_FEMALES"),
    "GROUP_AGE": (
        "CHAR_10_19_YEARS",
        "CHAR_20_24_YEARS",
        "CHAR_25_29_YEARS",
        "CHAR_30_34_YEARS",
        "CHAR_35_39_YEARS",
        "CHAR_40_44_YEARS",
        "CHAR_45_49_YEARS",
        "CHAR_50_54_YEARS",
        "CHAR_55_YEARS_AND_OVER",
    ),
}
DEFENDANT_RATE_LEGACY_OFFENCES = (
    "CHAR_01_HOMICIDE_AND_RELATED_OFFENCES",
    "CHAR_02_ACTS_INTENDED_TO_CAUSE_INJURY",
    "CHAR_03_SEXUAL_ASSAULT_AND_RELATED_OFFENCES",
    "CHAR_04_DANGEROUS_NEGLIGENT_ACTS",
    "CHAR_05_ABDUCTION_HARASSMENT",
    "CHAR_06_ROBBERY_EXTORTION",
    "CHAR_07_UNLAWFUL_ENTRY_WITH_INTENT",
    "CHAR_08_THEFT",
    "CHAR_09_FRAUD_DECEPTION",
    "CHAR_10_ILLICIT_DRUG_OFFENCES",
    "CHAR_11_WEAPONS_EXPLOSIVES",
    "CHAR_12_PROPERTY_DAMAGE_AND_ENVIRONMENTAL_POLLUTION",
    "CHAR_13_PUBLIC_ORDER_OFFENCES",
    "CHAR_14_TRAFFIC_AND_VEHICLE_REGULATORY_OFFENCES",
    "CHAR_15_OFFENCES_AGAINST_JUSTICE",
    "CHAR_16_MISCELLANEOUS_OFFENCES",
    "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
)
DEFENDANT_RATE_CONCORDED_OFFENCES = (
    "CHAR_01_HOMICIDE",
    "CHAR_02_ASSAULT",
    "CHAR_03_SEXUAL_OFFENCES",
    "CHAR_04_HARM_OR_ENDANGER_PERSONS",
    "CHAR_05_ROBBERY_BLACKMAIL_AND_EXTORTION",
    "CHAR_06_BURGLARY",
    "CHAR_07_THEFT",
    "CHAR_08_FRAUD_AND_RELATED_OFFENCES",
    "CHAR_09_DRUG_OFFENCES",
    "CHAR_10_WEAPONS_AND_EXPLOSIVES_OFFENCES",
    "CHAR_11_PROPERTY_DAMAGE",
    "CHAR_12_PUBLIC_ORDER_HEALTH_AND_SAFETY_OFFENCES",
    "CHAR_13_TRAFFIC_AND_VEHICLE_OFFENCES",
    "CHAR_14_OFFENCES_AGAINST_JUSTICE_PROCEDURES_AND_ORDERS",
    "CHAR_15_OFFENCES_AGAINST_GOVERNMENT",
    "CHAR_16_ENVIRONMENTAL_OFFENCES",
    "CHAR_17_MISCELLANEOUS_OFFENCES",
    "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
)
DEFENDANT_RATE_SOURCE_GEOMETRY = {
    2021: (
        4,
        5,
        (
            (
                "COUNT",
                6,
                (
                    ("GROUP_SEX", 7, (8, 9)),
                    ("GROUP_AGE", 11, tuple(range(12, 21))),
                    ("GROUP_PRINCIPAL_OFFENCE", 22, tuple(range(23, 40))),
                ),
            ),
            (
                "RATE",
                40,
                (
                    ("GROUP_SEX", 41, (42, 43)),
                    ("GROUP_AGE", 45, tuple(range(46, 55))),
                    ("GROUP_PRINCIPAL_OFFENCE", 56, tuple(range(57, 74))),
                ),
            ),
        ),
    ),
    2022: (
        4,
        5,
        (
            (
                "COUNT",
                6,
                (
                    ("GROUP_SEX", 7, (8, 9)),
                    ("GROUP_AGE", 11, tuple(range(12, 21))),
                    ("GROUP_PRINCIPAL_OFFENCE", 22, tuple(range(23, 40))),
                ),
            ),
            (
                "RATE",
                40,
                (
                    ("GROUP_SEX", 41, (42, 43)),
                    ("GROUP_AGE", 45, tuple(range(46, 55))),
                    ("GROUP_PRINCIPAL_OFFENCE", 56, tuple(range(57, 74))),
                ),
            ),
        ),
    ),
    2023: (
        3,
        4,
        (
            (
                "COUNT",
                5,
                (
                    ("GROUP_SEX", 6, (7, 8)),
                    ("GROUP_AGE", 10, tuple(range(11, 20))),
                    ("GROUP_PRINCIPAL_OFFENCE", 21, tuple(range(22, 39))),
                ),
            ),
            (
                "RATE",
                39,
                (
                    ("GROUP_SEX", 40, (41, 42)),
                    ("GROUP_AGE", 44, tuple(range(45, 54))),
                    ("GROUP_PRINCIPAL_OFFENCE", 55, tuple(range(56, 73))),
                ),
            ),
        ),
    ),
    2024: (
        2,
        5,
        (
            (
                "RATE",
                6,
                (
                    ("GROUP_SEX", 7, (8, 9)),
                    ("GROUP_AGE", 11, tuple(range(12, 21))),
                    ("GROUP_PRINCIPAL_OFFENCE", 22, tuple(range(23, 41))),
                ),
            ),
            (
                "COUNT",
                41,
                (
                    ("GROUP_SEX", 42, (43, 44)),
                    ("GROUP_AGE", 46, tuple(range(47, 56))),
                    ("GROUP_PRINCIPAL_OFFENCE", 57, tuple(range(58, 76))),
                ),
            ),
        ),
    ),
}
DEFENDANT_RATE_CSV_FIELDS = (
    "publication_vintage_date",
    "reference_date",
    "characteristic_category_id",
    "characteristic_group_id",
    "observation_period_id",
    "statistic_basis_id",
    "jurisdiction_id",
    "classification_context_id",
    "measure_id",
    "unit_id",
    "value",
    "value_status",
    "raw_value",
    "source_workbook_digest",
    "source_sheet",
    "source_cell",
    "recipe_digest",
    "publication_id",
    "execution_digest",
    "acceptance_policy_version",
    "acceptance_policy_digest",
    "acceptance_decision_digest",
    "prompt_package_digest",
    "generation_model",
    "generation_attempt_id",
    "raw_characteristic_category",
    "raw_characteristic_group",
    "raw_observation_period",
    "raw_statistic_basis",
    "raw_jurisdiction",
    "raw_classification_context",
)


def _defendant_rate_cube_members() -> dict[str, list[dict[str, object]]]:
    membership = _load(MEMBERSHIP)
    selected = {}
    for family in membership["families"]:
        members = [
            member
            for member in family["members"]
            if member["cubeId"] == "rate-of-defendants-finalised-australia"
        ]
        if members:
            selected[family["familyId"]] = members
    return selected


def _r1c1_parts(address: str) -> tuple[int, int]:
    match = re.fullmatch(r"R([1-9][0-9]*)C([1-9][0-9]*)", address)
    assert match is not None
    return int(match.group(1)), int(match.group(2))


def _defendant_rate_evidence_spec(family_id: str) -> LargeBatchSpec:
    _, years, _, year_counts = DEFENDANT_RATE_SOURCE_MATRIX[family_id]
    canonical_count = sum(year_counts)
    assert canonical_count % 2 == 0
    return LargeBatchSpec(
        family_id=family_id,
        label="Criminal Courts defendant counts and published rates",
        cohort_path=f"fixtures/product-prototype/{family_id}.json",
        evidence_manifest_path=(
            f"fixtures/product-prototype/{family_id}-evidence/manifest.json"
        ),
        dagster_asset=f"cube_12_{years[-1]}_{len(years)}",
        dagster_job=f"cube_12_{years[-1]}_{len(years)}_job",
        output_directory=f".tmp/cube-12/{family_id}",
        expected_years=years,
        expected_year_counts=year_counts,
        expected_canonical_count=canonical_count,
        expected_excluded_observation_count=0,
        expected_excluded_observation_counts_by_year={year: 0 for year in years},
        expected_measure_counts={
            "defendant-count": canonical_count // 2,
            "defendant-rate": canonical_count // 2,
        },
        expected_value_status_counts={"observed": canonical_count},
        expected_manual_replay_years=years,
        preserves_publication_vintage=True,
        acceptance_policy_version="tidy.table-family-acceptance/v2",
        replay_recorded_at="2026-08-23T09:00:00+00:00",
    )


def _canonical_csv_scalar(field: str, value: object) -> str:
    assert isinstance(value, str | int | float) and not isinstance(value, bool)
    if isinstance(value, str) and field != "source_sheet":
        return value.rstrip()
    return str(value)


def _assert_canonical_csv_matches_json(
    csv_path: Path, rows: list[dict[str, object]]
) -> None:
    assert rows
    with csv_path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        expected_fields = reader.fieldnames
        assert expected_fields == list(DEFENDANT_RATE_CSV_FIELDS)
        csv_rows = list(reader)
    assert all(set(row) == set(DEFENDANT_RATE_CSV_FIELDS) for row in rows)
    assert csv_rows == [
        {field: _canonical_csv_scalar(field, row[field]) for field in expected_fields}
        for row in rows
    ]


def _defendant_rate_source_dimensions(
    *,
    year: int,
    source_row: int,
    source_column: int,
    values: dict[tuple[int, int], object],
    jurisdiction_id: str,
) -> dict[str, tuple[object, str]]:
    title_row, period_row, panels = DEFENDANT_RATE_SOURCE_GEOMETRY[year]
    selected = None
    for basis_id, basis_row, groups in panels:
        for group_id, group_row, category_rows in groups:
            if source_row in category_rows:
                selected = (basis_id, basis_row, group_id, group_row, category_rows)
                break
        if selected is not None:
            break
    assert selected is not None
    basis_id, basis_row, group_id, group_row, category_rows = selected
    if group_id in DEFENDANT_RATE_COMMON_CATEGORIES:
        category_ids = DEFENDANT_RATE_COMMON_CATEGORIES[group_id]
    else:
        category_ids = (
            DEFENDANT_RATE_CONCORDED_OFFENCES
            if year == 2024
            else DEFENDANT_RATE_LEGACY_OFFENCES
        )
    category_id = category_ids[category_rows.index(source_row)]
    period_id = f"{2009 + source_column:04d}-06-30"
    title = values[(title_row, 1)]
    return {
        "characteristic_category": (values[(source_row, 1)], category_id),
        "characteristic_group": (values[(group_row, 1)], group_id),
        "observation_period": (values[(period_row, source_column)], period_id),
        "statistic_basis": (values[(basis_row, 2)], basis_id),
        "jurisdiction": (title, jurisdiction_id),
        "classification_context": (
            title,
            (
                "ANZSOC_2023_WITH_CONCORDED_ANZSOC_2011_SERIES"
                if year == 2024
                else "ANZSOC_2011"
            ),
        ),
    }


def _assert_defendant_rate_source_dimensions(
    family_id: str,
    cohort: dict[str, object],
    contract: dict[str, object],
    rows: list[dict[str, object]],
) -> None:
    jurisdiction_id, expected_years, expected_sheets, expected_counts = (
        DEFENDANT_RATE_SOURCE_MATRIX[family_id]
    )
    expected_by_year = dict(zip(expected_years, expected_sheets, strict=True))
    expected_count_by_year = dict(zip(expected_years, expected_counts, strict=True))
    cohort_by_year = {workbook["year"]: workbook for workbook in cohort["workbooks"]}
    assert tuple(cohort_by_year) == expected_years
    tables: dict[int, dict[tuple[int, int], object]] = {}
    for year in expected_years:
        workbook = cohort_by_year[year]
        expected_path, expected_digest, _ = DEFENDANT_RATE_SOURCE_FILES[year]
        assert workbook["path"] == expected_path
        assert workbook["contentDigest"] == expected_digest
        assert workbook["sheet"] == expected_by_year[year]
        source = load_workbook(
            FIXTURES / expected_path, read_only=True, data_only=False
        )
        try:
            worksheet = source[expected_by_year[year]]
            tables[year] = {
                (row_number, column_number): cell.value
                for row_number, source_row in enumerate(
                    worksheet.iter_rows(max_row=75, max_col=16), start=1
                )
                for column_number, cell in enumerate(source_row, start=1)
            }
        finally:
            source.close()
    rows_by_year = Counter()
    raw_fields = {
        "characteristic_category": "raw_characteristic_category",
        "characteristic_group": "raw_characteristic_group",
        "observation_period": "raw_observation_period",
        "statistic_basis": "raw_statistic_basis",
        "jurisdiction": "raw_jurisdiction",
        "classification_context": "raw_classification_context",
    }
    publication_dates = {
        2021: "2022-06-30",
        2022: "2023-06-30",
        2023: "2024-06-30",
        2024: "2025-06-30",
    }
    digest_to_year = {
        DEFENDANT_RATE_SOURCE_FILES[year][1]: year for year in expected_years
    }
    for row in rows:
        year = digest_to_year[row["source_workbook_digest"]]
        assert row["publication_vintage_date"] == publication_dates[year]
        assert row["source_sheet"] == expected_by_year[year]
        source_row, source_column = _r1c1_parts(row["source_cell"])
        assert source_column >= 2
        expected = _defendant_rate_source_dimensions(
            year=year,
            source_row=source_row,
            source_column=source_column,
            values=tables[year],
            jurisdiction_id=jurisdiction_id,
        )
        for dimension, (raw, canonical_id) in expected.items():
            assert row[raw_fields[dimension]] == raw
            assert row[f"{dimension}_id"] == canonical_id
            alias_key = " ".join(str(raw).strip().split())
            assert contract["aliases"][dimension][alias_key] == canonical_id
        assert row["reference_date"] == expected["observation_period"][1]
        assert row["measure_id"] == (
            "defendant-count"
            if expected["statistic_basis"][1] == "COUNT"
            else "defendant-rate"
        )
        assert row["unit_id"] == (
            "person"
            if expected["statistic_basis"][1] == "COUNT"
            else "per-100000-persons-aged-10-plus"
        )
        rows_by_year[year] += 1
    assert rows_by_year == expected_count_by_year


def test_defendant_rate_cube_topology_contracts_and_frozen_evidence() -> None:
    members_by_family = _defendant_rate_cube_members()
    assert set(members_by_family) == set(DEFENDANT_RATE_CUBE_FAMILIES)
    members = [member for family in members_by_family.values() for member in family]
    assert len(members_by_family) == 18
    assert Counter(len(family) for family in members_by_family.values()) == {3: 9, 1: 9}
    assert Counter(member["releaseId"] for member in members) == {
        "2021-22": 9,
        "2022-23": 9,
        "2023-24": 9,
        "2024-25": 9,
    }
    assert {
        release: sorted(
            member["physicalTableNumber"]
            for member in members
            if member["releaseId"] == release
        )
        for release in ("2021-22", "2022-23", "2023-24", "2024-25")
    } == {
        "2021-22": list(range(56, 65)),
        "2022-23": list(range(56, 65)),
        "2023-24": list(range(64, 73)),
        "2024-25": list(range(69, 78)),
    }
    source_digests = {
        "2021-22": (
            "sha256:5d7734110d0f6348c017e4b3ec2fad4118b95de47cb48b6e7b2cf51f4f57bcee"
        ),
        "2022-23": (
            "sha256:4412ad57f73fe112e3c8d546bbd348f720aa72cd5abb422d54e5ee19fdd445ec"
        ),
        "2023-24": (
            "sha256:1c800c30cf50594a0ece895882981cf109d36601cafcfcaaa9fd0aaece38d0f6"
        ),
        "2024-25": (
            "sha256:f0411a3531d5c74ef68f24c6a9a57edfc24fdbf5d95faab0a8c73690f57bc8de"
        ),
    }
    assert all(
        member["physicalSheetName"] == f"Table {member['physicalTableNumber']}"
        and member["sourceDigest"] == source_digests[member["releaseId"]]
        and member["registered"] is True
        for member in members
    )
    assert {
        member["classificationContext"]
        for member in members
        if member["releaseId"] != "2024-25"
    } == {"anzsoc-2011"}
    assert {
        member["classificationContext"]
        for member in members
        if member["releaseId"] == "2024-25"
    } == {"anzsoc-2023-with-concorded-anzsoc-2011-series"}
    release_years = {
        "2021-22": 2021,
        "2022-23": 2022,
        "2023-24": 2023,
        "2024-25": 2024,
    }
    for family_id, (
        _,
        expected_years,
        expected_sheets,
        _,
    ) in DEFENDANT_RATE_SOURCE_MATRIX.items():
        family_members = members_by_family[family_id]
        assert tuple(
            release_years[member["releaseId"]] for member in family_members
        ) == (expected_years)
        assert tuple(member["physicalSheetName"] for member in family_members) == (
            expected_sheets
        )
        assert tuple(member["sourceDigest"] for member in family_members) == tuple(
            DEFENDANT_RATE_SOURCE_FILES[year][2] for year in expected_years
        )

    all_rows = []
    workbook_paths: dict[str, Path] = {}
    title_by_source: dict[tuple[str, str], str] = {}
    warning_counts = Counter()
    aliases_declared = 0
    for family_id in DEFENDANT_RATE_CUBE_FAMILIES:
        cohort_path = FIXTURES / f"{family_id}.json"
        contract_path = FIXTURES / "acceptance" / f"{family_id}-v1.json"
        evidence = FIXTURES / f"{family_id}-evidence"
        cohort = _load(cohort_path)
        contract = _load(contract_path)
        run = _load(evidence / "run.json")
        rows = json.loads((evidence / "canonical-observations.json").read_text())
        _assert_canonical_csv_matches_json(
            evidence / "canonical-observations.csv", rows
        )
        _assert_defendant_rate_source_dimensions(family_id, cohort, contract, rows)
        product_prototype_module._validate_cohort(cohort)
        product_prototype_module._validate_contract(contract, cohort)
        assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
        assert contract["strictAliasMatching"] is True
        assert contract["trainingEligibility"] is False
        assert contract["totalEquations"] == []
        assert contract["totalValidation"] == "not_applicable"
        assert contract["preservePublicationVintage"] is True
        assert {measure["id"] for measure in contract["measures"]} == {
            "defendant-count",
            "defendant-rate",
        }
        assert {
            (measure["id"], measure["unitId"]) for measure in contract["measures"]
        } == {
            ("defendant-count", "person"),
            ("defendant-rate", "per-100000-persons-aged-10-plus"),
        }
        assert all(
            workbook["replayResponse"]["acceptanceAuthority"] is False
            for workbook in cohort["workbooks"]
        )
        assert run["providerCalls"] == 0
        assert run["exceptionWorkbookCount"] == 0
        assert run["trainingEligibility"] is False
        assert run["historicalReplayIsAcceptanceAuthority"] is False
        assert all(
            workbook["issues"] == [] and all(workbook["checks"].values())
            for workbook in run["workbooks"]
        )
        verify_large_batch_evidence(PROJECT, _defendant_rate_evidence_spec(family_id))
        manifest = _load(evidence / "manifest.json")
        assert manifest["recordedAt"] == "2026-08-23T09:00:00+00:00"
        assert manifest["acceptanceContractDigest"] == sha256_digest(
            contract_path.read_bytes()
        )
        assert manifest["cohortDigest"] == sha256_digest(cohort_path.read_bytes())
        assert {path.name for path in evidence.iterdir()} == {
            "README.md",
            "canonical-observations.csv",
            "canonical-observations.json",
            "collation-report.json",
            "exceptions.json",
            "run.json",
            "manifest.json",
        }
        assert (
            contract["expectedWarningCountsByYear"] == manifest["warningCountsByYear"]
        )
        warning_counts.update(
            {
                int(year): count
                for year, count in manifest["warningCountsByYear"].items()
            }
        )
        run_by_year = {workbook["year"]: workbook for workbook in run["workbooks"]}
        member_by_release = {
            member["releaseId"]: member for member in members_by_family[family_id]
        }
        release_by_year = {
            2021: "2021-22",
            2022: "2022-23",
            2023: "2023-24",
            2024: "2024-25",
        }
        for workbook in cohort["workbooks"]:
            year = workbook["year"]
            member = member_by_release[release_by_year[year]]
            assert workbook["sheet"] == member["physicalSheetName"]
            assert run_by_year[year]["workbookDigest"] == workbook["contentDigest"]
            assert run_by_year[year]["sheet"] == member["physicalSheetName"]
            assert run_by_year[year]["referenceDate"] == workbook["referenceDate"]
            response_path = FIXTURES / workbook["replayResponse"]["path"]
            response = json.loads(response_path.read_text())
            assert response["version"] == "semantic-table-map-v1"
            assert response["table"]["values"]["regions"] == ["region-001"]
            assert (
                sha256_digest(response_path.read_bytes())
                == workbook["replayResponse"]["contentDigest"]
            )
            assert (
                len(response_path.read_bytes())
                == workbook["replayResponse"]["byteLength"]
            )
            workbook_path = FIXTURES / workbook["path"]
            assert (
                sha256_digest(workbook_path.read_bytes()) == workbook["contentDigest"]
            )
            workbook_paths[workbook["contentDigest"]] = workbook_path
            title_by_source[(workbook["contentDigest"], workbook["sheet"])] = member[
                "publishedTitle"
            ]
        aliases_by_dimension = {}
        raw_fields = {
            "characteristic_category": "raw_characteristic_category",
            "characteristic_group": "raw_characteristic_group",
            "observation_period": "raw_observation_period",
            "statistic_basis": "raw_statistic_basis",
            "jurisdiction": "raw_jurisdiction",
            "classification_context": "raw_classification_context",
        }
        for dimension, aliases in contract["aliases"].items():
            normalized = {}
            for raw, canonical in aliases.items():
                key = " ".join(raw.strip().split())
                assert " ".join(f"  {raw}\t ".split()) == key
                assert key not in normalized or normalized[key] == canonical
                normalized[key] = canonical
                aliases_declared += 1
            assert set(normalized) == {
                " ".join(str(row[raw_fields[dimension]]).strip().split())
                for row in rows
            }
            aliases_by_dimension[dimension] = normalized
        for row in rows:
            for dimension, raw_field in raw_fields.items():
                key = " ".join(str(row[raw_field]).strip().split())
                assert row[f"{dimension}_id"] == aliases_by_dimension[dimension][key]
            year = int(row["publication_vintage_date"][:4]) - 1
            assert (
                row["recipe_digest"]
                == contract["expectedRecipeDigestsByYear"][str(year)]
            )
            assert (
                row["raw_jurisdiction"]
                == title_by_source[(row["source_workbook_digest"], row["source_sheet"])]
            )
        all_rows.extend(rows)

    assert aliases_declared == 999
    assert len(all_rows) == 27486
    assert warning_counts == {2021: 3024, 2022: 3276, 2023: 3528, 2024: 3915}
    assert Counter((row["measure_id"], row["unit_id"]) for row in all_rows) == {
        ("defendant-count", "person"): 13743,
        ("defendant-rate", "per-100000-persons-aged-10-plus"): 13743,
    }
    assert Counter(row["characteristic_group_id"] for row in all_rows) == {
        "GROUP_PRINCIPAL_OFFENCE": 16794,
        "GROUP_AGE": 8748,
        "GROUP_SEX": 1944,
    }
    assert Counter(row["classification_context_id"] for row in all_rows) == {
        "ANZSOC_2011": 19656,
        "ANZSOC_2023_WITH_CONCORDED_ANZSOC_2011_SERIES": 7830,
    }
    assert Counter(row["publication_vintage_date"] for row in all_rows) == {
        "2022-06-30": 6048,
        "2023-06-30": 6552,
        "2024-06-30": 7056,
        "2025-06-30": 7830,
    }
    assert Counter(row["value_status"] for row in all_rows) == {"observed": 27486}
    assert all(
        isinstance(row["raw_value"], int | float)
        and not isinstance(row["raw_value"], bool)
        and row["value"] == row["raw_value"]
        for row in all_rows
    )
    assert sum(row["value"] == 0 for row in all_rows) == 20
    assert {
        row["statistic_basis_id"]
        for row in all_rows
        if row["measure_id"] == "defendant-count"
    } == {"COUNT"}
    assert {
        row["statistic_basis_id"]
        for row in all_rows
        if row["measure_id"] == "defendant-rate"
    } == {"RATE"}
    assert (
        len(
            {
                (row["source_workbook_digest"], row["source_sheet"], row["source_cell"])
                for row in all_rows
            }
        )
        == 27486
    )

    total_id = "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS"
    total_rows = [
        row for row in all_rows if row["characteristic_category_id"] == total_id
    ]
    assert len(total_rows) == 972
    assert {row["characteristic_group_id"] for row in total_rows} == {
        "GROUP_PRINCIPAL_OFFENCE"
    }
    assert all(
        row["raw_characteristic_category"].startswith(
            "Total finalised (excluding transfer to other court levels)"
        )
        for row in total_rows
    )
    markers = {
        marker
        for row in all_rows
        for field, raw in row.items()
        if field.startswith("raw_")
        for marker in re.findall(r"\([a-z]\)", str(raw))
    }
    assert markers == {f"({letter})" for letter in "abcdefghijkl"}
    legacy_offences = {
        row["characteristic_category_id"]
        for row in all_rows
        if row["classification_context_id"] == "ANZSOC_2011"
        and row["characteristic_group_id"] == "GROUP_PRINCIPAL_OFFENCE"
    }
    concorded_offences = {
        row["characteristic_category_id"]
        for row in all_rows
        if row["classification_context_id"]
        == "ANZSOC_2023_WITH_CONCORDED_ANZSOC_2011_SERIES"
        and row["characteristic_group_id"] == "GROUP_PRINCIPAL_OFFENCE"
    }
    assert len(legacy_offences) == 17
    assert len(concorded_offences) == 18
    assert legacy_offences & concorded_offences == {total_id}

    rows_by_workbook_sheet: dict[tuple[str, str], list[dict[str, object]]] = {}
    for row in all_rows:
        rows_by_workbook_sheet.setdefault(
            (row["source_workbook_digest"], row["source_sheet"]), []
        ).append(row)
    open_workbooks = {
        digest: load_workbook(path, read_only=True, data_only=False)
        for digest, path in workbook_paths.items()
    }
    try:
        for (digest, sheet), source_rows in rows_by_workbook_sheet.items():
            worksheet = open_workbooks[digest][sheet]
            requested = {_r1c1_parts(row["source_cell"]) for row in source_rows}
            max_row = max(row for row, _ in requested)
            max_col = max(column for _, column in requested)
            cells = {
                (row_number, column_number): cell
                for row_number, row in enumerate(
                    worksheet.iter_rows(max_row=max_row, max_col=max_col), start=1
                )
                for column_number, cell in enumerate(row, start=1)
                if (row_number, column_number) in requested
            }
            assert len(cells) == len(requested)
            title = title_by_source[(digest, sheet)]
            assert (
                sum(
                    cell.value == title
                    for row in worksheet.iter_rows(min_row=1, max_row=6, max_col=1)
                    for cell in row
                )
                == 1
            )
            for row in source_rows:
                cell = cells[_r1c1_parts(row["source_cell"])]
                assert cell.data_type == "n"
                assert cell.value == row["raw_value"]
                assert type(cell.value) is type(row["raw_value"])
    finally:
        for workbook in open_workbooks.values():
            workbook.close()


def test_defendant_rate_source_and_csv_checks_reject_coordinated_mutations(
    tmp_path: Path,
) -> None:
    family_id = DEFENDANT_RATE_CUBE_FAMILIES[0]
    cohort = _load(FIXTURES / f"{family_id}.json")
    contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
    evidence = FIXTURES / f"{family_id}-evidence"
    rows = json.loads((evidence / "canonical-observations.json").read_text())

    mutated_contract = copy.deepcopy(contract)
    mutated_rows = copy.deepcopy(rows)
    first = "CHAR_01_HOMICIDE_AND_RELATED_OFFENCES"
    second = "CHAR_02_ACTS_INTENDED_TO_CAUSE_INJURY"
    for raw, target in mutated_contract["aliases"]["characteristic_category"].items():
        if target == first:
            mutated_contract["aliases"]["characteristic_category"][raw] = second
        elif target == second:
            mutated_contract["aliases"]["characteristic_category"][raw] = first
    for row in mutated_rows:
        if row["characteristic_category_id"] == first:
            row["characteristic_category_id"] = second
        elif row["characteristic_category_id"] == second:
            row["characteristic_category_id"] = first
    with pytest.raises(AssertionError):
        _assert_defendant_rate_source_dimensions(
            family_id, cohort, mutated_contract, mutated_rows
        )

    csv_path = evidence / "canonical-observations.csv"
    with csv_path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames
        assert fieldnames is not None
        csv_rows = list(reader)
    csv_rows[0]["value"] = "640" if csv_rows[0]["value"] != "640" else "641"
    mutated_csv = tmp_path / "canonical-observations.csv"
    with mutated_csv.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(csv_rows)
    with pytest.raises(AssertionError):
        _assert_canonical_csv_matches_json(mutated_csv, rows)


def test_defendant_rate_warning_panels_are_exact_and_reverse_in_2024() -> None:
    expected = {
        2021: ("RATE", "R40C2", 3024),
        2022: ("RATE", "R40C2", 3276),
        2023: ("RATE", "R39C2", 3528),
        2024: ("COUNT", "R41C2", 3915),
    }
    by_year = Counter()
    header_values: dict[tuple[int, str, str], set[str]] = {}
    for family_id in DEFENDANT_RATE_CUBE_FAMILIES:
        contract = _load(FIXTURES / "acceptance" / f"{family_id}-v1.json")
        cohort = _load(FIXTURES / f"{family_id}.json")
        evidence = FIXTURES / f"{family_id}-evidence"
        run = _load(evidence / "run.json")
        rows = json.loads((evidence / "canonical-observations.json").read_text())
        rule = contract["allowedExecutionWarnings"][0]
        assert rule["code"] == "AMBIGUOUS_HEADER"
        assert rule["dimension"] == "statistic_basis"
        assert rule["requireCanonicalOutputEquivalence"] is True
        run_by_year = {workbook["year"]: workbook for workbook in run["workbooks"]}
        cohort_by_year = {
            workbook["year"]: workbook for workbook in cohort["workbooks"]
        }
        for year in run_by_year:
            ambiguous_code, header_source, count = expected[year]
            assert contract["expectedWarningCountsByYear"][str(year)] == count // 9
            assert rule["expectedHeaderSourcesByYear"][str(year)][ambiguous_code] == [
                header_source
            ]
            year_rows = [
                row
                for row in rows
                if int(row["publication_vintage_date"][:4]) - 1 == year
            ]
            ambiguous = [
                row for row in year_rows if row["statistic_basis_id"] == ambiguous_code
            ]
            assert len(ambiguous) == run_by_year[year]["executionWarningCount"]
            assert all(
                _r1c1_parts(row["source_cell"])[0]
                > int(header_source[1:].split("C")[0])
                for row in ambiguous
            )
            other = [
                row for row in year_rows if row["statistic_basis_id"] != ambiguous_code
            ]
            assert all(
                _r1c1_parts(row["source_cell"])[0]
                < int(header_source[1:].split("C")[0])
                for row in other
            )
            workbook_path = FIXTURES / cohort_by_year[year]["path"]
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            try:
                row_number, column = _r1c1_parts(header_source)
                header_value = (
                    workbook[cohort_by_year[year]["sheet"]]
                    .cell(row_number, column)
                    .value
                )
            finally:
                workbook.close()
            normalized = " ".join(str(header_value).strip().split())
            assert contract["aliases"]["statistic_basis"][normalized] == ambiguous_code
            header_values.setdefault((year, ambiguous_code, header_source), set()).add(
                normalized
            )
            by_year[year] += len(ambiguous)
    assert by_year == {year: declaration[2] for year, declaration in expected.items()}
    assert header_values == {
        (2021, "RATE", "R40C2"): {"Rate of defendants finalised"},
        (2022, "RATE", "R40C2"): {"Rate of defendants finalised"},
        (2023, "RATE", "R39C2"): {
            "Rate of defendants finalised(e)(f)",
            "Rate of defendants finalised(f)(g)",
            "Rate of defendants finalised(g)(h)",
            "Rate of defendants finalised(h)(i)",
        },
        (2024, "COUNT", "R41C2"): {"Number of defendants finalised"},
    }


def _workbook_sheet_xml(
    archive: zipfile.ZipFile, sheet_name: str
) -> ElementTree.Element:
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    sheet = next(
        item
        for item in workbook.findall(f".//{{{main}}}sheet")
        if item.attrib["name"] == sheet_name
    )
    relationship_id = sheet.attrib[f"{{{relationships}}}id"]
    rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationship = next(
        item
        for item in rels.findall(f"{{{package_relationships}}}Relationship")
        if item.attrib["Id"] == relationship_id
    )
    target = relationship.attrib["Target"]
    path = target.lstrip("/") if target.startswith("/xl/") else f"xl/{target}"
    return ElementTree.fromstring(archive.read(path))


def test_defendant_rate_normalization_is_exact_and_reproducible(tmp_path: Path) -> None:
    manifest = _load(FIXTURES / "batch-workbook-normalization-v1.json")
    semantic = {
        key: value for key, value in manifest.items() if key != "manifestDigest"
    }
    assert (
        domain_digest(manifest["schemaVersion"], semantic) == manifest["manifestDigest"]
    )
    entries = [
        entry for entry in manifest["entries"] if "cube-12" in entry["sourcePath"]
    ]
    assert len(entries) == 1
    entry = entries[0]
    assert entry["correction"] is None
    assert entry["trimmedSheets"] == [
        {"sheet": "Table 68", "retainedRange": "A1:O84"},
        {"sheet": "Table 71", "retainedRange": "A1:O82"},
    ]
    source = PROJECT / entry["sourcePath"]
    normalized = PROJECT / entry["outputPath"]
    assert sha256_digest(source.read_bytes()) == entry["sourceDigest"]
    assert sha256_digest(normalized.read_bytes()) == entry["outputDigest"]
    reproduced = tmp_path / "cube-12-normalized.xlsx"
    completed = subprocess.run(
        [
            str(PROJECT / manifest["scriptPath"]),
            str(source),
            str(reproduced),
            "--sheet",
            "Table 68=A1:O84",
            "--sheet",
            "Table 71=A1:O82",
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=120,
    )
    assert completed.returncode == 0, completed.stderr
    assert reproduced.read_bytes() == normalized.read_bytes()

    source_workbook = load_workbook(source, read_only=True, data_only=False)
    normalized_workbook = load_workbook(normalized, read_only=True, data_only=False)
    limits = {"Table 68": (84, 15), "Table 71": (82, 15)}
    sheet_names = source_workbook.sheetnames
    try:
        assert sheet_names == normalized_workbook.sheetnames
        for sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet_name]
            normalized_sheet = normalized_workbook[sheet_name]
            max_row, max_column = limits.get(
                sheet_name,
                (
                    max(source_sheet.max_row, normalized_sheet.max_row),
                    max(source_sheet.max_column, normalized_sheet.max_column),
                ),
            )
            source_rows = source_sheet.iter_rows(max_row=max_row, max_col=max_column)
            normalized_rows = normalized_sheet.iter_rows(
                max_row=max_row, max_col=max_column
            )
            for source_row, normalized_row in zip(
                source_rows, normalized_rows, strict=True
            ):
                assert [
                    (
                        cell.value,
                        getattr(cell, "data_type", "n"),
                        getattr(cell, "_style_id", 0),
                        getattr(cell, "number_format", "General"),
                    )
                    for cell in source_row
                ] == [
                    (
                        cell.value,
                        getattr(cell, "data_type", "n"),
                        getattr(cell, "_style_id", 0),
                        getattr(cell, "number_format", "General"),
                    )
                    for cell in normalized_row
                ]
    finally:
        source_workbook.close()
        normalized_workbook.close()

    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    with (
        zipfile.ZipFile(source) as source_archive,
        zipfile.ZipFile(normalized) as normalized_archive,
    ):
        for sheet_name in sheet_names:
            source_xml = _workbook_sheet_xml(source_archive, sheet_name)
            normalized_xml = _workbook_sheet_xml(normalized_archive, sheet_name)
            source_merges = {
                item.attrib["ref"]
                for item in source_xml.findall(f".//{{{spreadsheet}}}mergeCell")
            }
            normalized_merges = {
                item.attrib["ref"]
                for item in normalized_xml.findall(f".//{{{spreadsheet}}}mergeCell")
            }
            if sheet_name == "Table 71":
                assert source_merges - normalized_merges == {"A82:O1048576"}
                assert normalized_merges == source_merges - {"A82:O1048576"}
            else:
                assert normalized_merges == source_merges
        source_outside_counts = {}
        normalized_outside_counts = {}
        for sheet_name, retained_range in {
            "Table 68": "A1:O84",
            "Table 71": "A1:O82",
        }.items():
            min_column, min_row, max_column, max_row = range_boundaries(retained_range)
            outside_by_kind = {}
            for kind, archive in (
                ("source", source_archive),
                ("normalized", normalized_archive),
            ):
                xml = _workbook_sheet_xml(archive, sheet_name)
                outside = []
                for cell in xml.findall(f".//{{{spreadsheet}}}c"):
                    letters, row = coordinate_from_string(cell.attrib["r"])
                    column = 0
                    for character in letters:
                        column = column * 26 + ord(character) - ord("A") + 1
                    if not (
                        min_row <= row <= max_row and min_column <= column <= max_column
                    ):
                        outside.append(cell)
                outside_by_kind[kind] = outside
            source_outside = outside_by_kind["source"]
            normalized_outside = outside_by_kind["normalized"]
            source_outside_counts[sheet_name] = len(source_outside)
            normalized_outside_counts[sheet_name] = len(normalized_outside)
            assert all(
                cell.find(f"{{{spreadsheet}}}v") is None
                and cell.find(f"{{{spreadsheet}}}f") is None
                and cell.find(f"{{{spreadsheet}}}is") is None
                for cell in source_outside
            )
            assert normalized_outside == []
        assert source_outside_counts == {"Table 68": 186, "Table 71": 0}
        assert normalized_outside_counts == {"Table 68": 0, "Table 71": 0}


FDV_BREACH_FAMILIES = (
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-00438881d8",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-0964025147",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-0e2d3059c2",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-4f8a6d549a",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-838adf447e",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-85ca3806a6",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-8d69cd5de3",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-9bdd94cf9f",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-ae9d5bfaeb",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b",
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645",
)
FDV_BREACH_SOURCE_MATRIX = (
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-00438881d8",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 25",
        71,
        7,
        240,
        228,
        6,
        6,
        235,
        5,
        0,
        9,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-00438881d8-2024.response.txt",
        "sha256:8b27cd2d82812e7c7e4e3eef2042deaa8d21fe7fa7fd027be412abe952b76e4f",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-0964025147",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 26",
        71,
        7,
        240,
        228,
        6,
        6,
        225,
        15,
        0,
        27,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-0964025147-2024.response.txt",
        "sha256:a36ea4266ca1c594ec38b148624c47c91a6c1ee711f13943aa09260da2e11e5f",
        1038,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-0e2d3059c2",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 21",
        69,
        7,
        240,
        228,
        6,
        6,
        234,
        0,
        6,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-0e2d3059c2-2024.response.txt",
        "sha256:aac8ff9e563ce336a9e96e590fbed2092aa96bcdf96f59583f2a945065e13b2f",
        1038,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-4f8a6d549a",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 24",
        72,
        7,
        240,
        228,
        6,
        6,
        230,
        10,
        0,
        16,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-4f8a6d549a-2024.response.txt",
        "sha256:6b24d4e091b205db38a482df8a5e4617a23a33d0a76afb275ea59d38d1815f3b",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-838adf447e",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 23",
        67,
        7,
        222,
        210,
        6,
        6,
        218,
        4,
        0,
        9,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-838adf447e-2024.response.txt",
        "sha256:fc27ae6cd8af3d98be4ca77353147604c24faec7c6b6507ded9bb8b560f24a9a",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-85ca3806a6",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 22",
        74,
        7,
        240,
        228,
        6,
        6,
        236,
        4,
        0,
        12,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-85ca3806a6-2024.response.txt",
        "sha256:c13839feea69472a4d22b623bddaed3f4ccf3fb344b267ddd5744e04ea8d01bd",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-8d69cd5de3",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 20",
        66,
        7,
        222,
        210,
        6,
        6,
        218,
        4,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-8d69cd5de3-2024.response.txt",
        "sha256:fcc7bf3cedabe1227b0c4b948191d11b8851d498f40004622060cb48c035c760",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-9bdd94cf9f",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 18",
        69,
        7,
        234,
        222,
        6,
        6,
        226,
        8,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-9bdd94cf9f-2024.response.txt",
        "sha256:6999c74ea0a9effe33b9771304a05b3f22521d547f13b0b5defb4b01a5f10aa2",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-ae9d5bfaeb",
        2024,
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
        96631,
        "FDV Table 19",
        72,
        7,
        240,
        228,
        6,
        6,
        236,
        4,
        0,
        3,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-restraining-order-violenc-ae9d5bfaeb-2024.response.txt",
        "sha256:e5da8ce617622234b82e8d5c799c2db9cb2e071a04e914b91b74d47a3226011f",
        999,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 21",
        76,
        13,
        128,
        120,
        4,
        4,
        128,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795-2021.response.txt",
        "sha256:9fafda04051f59f0ca6f8997ba0f7ba84f7c974fff63b976ad0867a5352e43e0",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 21",
        54,
        13,
        175,
        165,
        5,
        5,
        175,
        0,
        0,
        10,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795-2022.response.txt",
        "sha256:9b7f892f1872c8fdd4f8e51df901ec7b32847a39c2fcabb67f5292eaecd3373c",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 21",
        66,
        7,
        210,
        198,
        6,
        6,
        210,
        0,
        0,
        10,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-27caf21795-2023.response.txt",
        "sha256:d57837d639c53ab3ad8a7684cdc71bec79d62f86936ca01ddf3d0ba7f542c58c",
        1038,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 16",
        72,
        37,
        120,
        112,
        4,
        4,
        120,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571-2021.response.txt",
        "sha256:ca1c0e2d6a6082c42e09e0e34a46ebb6f8c8816c9534cba5d6ac862ae66a7a86",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 16",
        50,
        37,
        165,
        155,
        5,
        5,
        165,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571-2022.response.txt",
        "sha256:5b05120305de5dda7d07dfecbb4e753ef85d711c2e79f363b37a4bf946b8d707",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 16",
        63,
        7,
        204,
        192,
        6,
        6,
        199,
        5,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-37a9caf571-2023.response.txt",
        "sha256:160279c44789255adf3691dce83fca597e8f0f67824c7c677de9e314ed0f3553",
        1077,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 19",
        63,
        39,
        120,
        112,
        4,
        4,
        120,
        0,
        0,
        4,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d-2021.response.txt",
        "sha256:e96242470d78b1a42b298eccb12fbdc558fde735e2321869de874dc1894406bc",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 19",
        51,
        39,
        165,
        155,
        5,
        5,
        165,
        0,
        0,
        10,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d-2022.response.txt",
        "sha256:63c11384ba6dcac6b29b2177bb7a0030f323d7a2b4ff5771e9cdfca01bd54773",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 19",
        64,
        7,
        204,
        192,
        6,
        6,
        199,
        5,
        0,
        11,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-59939c8e3d-2023.response.txt",
        "sha256:26131be55f901763e755f62de87ad6ce0a396ae565b8c2c44276c3c78954af38",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 22",
        66,
        14,
        122,
        114,
        4,
        4,
        122,
        0,
        0,
        8,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831-2021.response.txt",
        "sha256:dd37d4c08080b854cec16929d67037bbf2f39d977f2a7f9b7ff8cd79140558be",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 22",
        54,
        14,
        169,
        159,
        5,
        5,
        169,
        0,
        0,
        18,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831-2022.response.txt",
        "sha256:c79d759d1e527e27949e472cf3ce933d81eecba17f62ab5643cf6837a6737bf7",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 22",
        67,
        7,
        210,
        198,
        6,
        6,
        204,
        6,
        0,
        26,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831-2023.response.txt",
        "sha256:10aa4f2edf2678bc29dd859783645b25a8e7a139ed50734406cf19320a7ece81",
        1038,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 18",
        67,
        37,
        96,
        90,
        3,
        3,
        96,
        0,
        0,
        2,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33-2021.response.txt",
        "sha256:4e8ed01209ba1189737879600885d5129038f820669db6b76a8a1fba5d3d98c9",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 18",
        55,
        37,
        140,
        132,
        4,
        4,
        140,
        0,
        0,
        3,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33-2022.response.txt",
        "sha256:60358ad94b5c65568f417dbf6847fc17055e7e9bafc94f7efbcdf4efa011830a",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 18",
        71,
        6,
        180,
        170,
        5,
        5,
        176,
        4,
        0,
        11,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-7ced84ab33-2023.response.txt",
        "sha256:39ea57bf3a0af3715a7e00e483740f8a1664d49f63c48977199e0b2fad0a276b",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 14",
        71,
        25,
        90,
        84,
        3,
        3,
        90,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079-2021.response.txt",
        "sha256:94dd25ee3edf38fbe16c7085c67856712cbbbe1cd0b77896f6a254b0174e0062",
        1038,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 14",
        59,
        25,
        132,
        124,
        4,
        4,
        132,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079-2022.response.txt",
        "sha256:bbde412d85452241b536a6b58bd939ae1b030e270cb0853f4c84dd20706092eb",
        1038,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 14",
        62,
        7,
        170,
        160,
        5,
        5,
        166,
        4,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-827a330079-2023.response.txt",
        "sha256:bd1ff89b0edb7ac93c5d174a3073ef9392c996b34a7b54619b9938cd9d5054c5",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 15",
        76,
        38,
        128,
        120,
        4,
        4,
        128,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7-2021.response.txt",
        "sha256:bc8a58bcada62288da818d8a02d182c7f6cda796a819c266de79e06946d7c8a9",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 15",
        54,
        38,
        175,
        165,
        5,
        5,
        175,
        0,
        0,
        2,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7-2022.response.txt",
        "sha256:202ddbe9fdd0f1620a71a06e06613abb479ef5abdc30edaa7ddf99c2be75dfd5",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 15",
        69,
        7,
        216,
        204,
        6,
        6,
        211,
        5,
        0,
        2,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-829327a2a7-2023.response.txt",
        "sha256:8246e22756f81e5b3df0899e16416c5a8c3045f47eb7ebac636e646787315013",
        1077,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 20",
        72,
        38,
        122,
        114,
        4,
        4,
        122,
        0,
        0,
        8,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b-2021.response.txt",
        "sha256:499948a8f3681e237b0c92eb162ea228ba17e42ca8cc80a2615ef9b3940b04c4",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 20",
        54,
        38,
        169,
        159,
        5,
        5,
        169,
        0,
        0,
        10,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b-2022.response.txt",
        "sha256:6e7d98e5e27a679c593cd0ce5a912ac23f1bc1851dea2a81505dd0939156ad5c",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 20",
        70,
        7,
        216,
        204,
        6,
        6,
        205,
        11,
        0,
        16,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-88e189a36b-2023.response.txt",
        "sha256:f2e1cecfe70f52935c3c230fff1ceca54ac99fb30284345990d8436628ed4146",
        1064,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645",
        2021,
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
        125006,
        "FDV Table 17",
        76,
        39,
        128,
        120,
        4,
        4,
        128,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645-2021.response.txt",
        "sha256:58a2a459129e795313ec737dc2139d1bb5e1f4fb9b1e20a915856af373014f0c",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645",
        2022,
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
        125663,
        "FDV Table 17",
        65,
        39,
        175,
        165,
        5,
        5,
        175,
        0,
        0,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645-2022.response.txt",
        "sha256:e08b348602f7742d06ea0c9666157f072b022f96cdb5801e42045f7ada851736",
        1051,
    ),
    (
        "criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645",
        2023,
        "workbooks/criminal-courts-2023-24-cube-17-normalized.xlsx",
        "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580",
        76428,
        "FDV Table 17",
        66,
        7,
        216,
        204,
        6,
        6,
        210,
        0,
        6,
        0,
        "replay/criminal-courts-family-domestic-violence-family-and-domestic-violence-defendants-finalised-breach-of-violence-orders-summary-c-8c70b71645-2023.response.txt",
        "sha256:c1efdb1e6fa8472576f2d3b2580af97d3df3681ac39dc6aec79b9c0ed889773d",
        1077,
    ),
)


# These digests freeze the exact contract and cohort bytes independently of the
# evidence they govern. Keys are the collision-resistant family suffixes.
FDV_BREACH_IDENTITY_PINS = {
    "00438881d8": (
        "sha256:4cf92fde98037c0ac8327d3dff48c1c30a7918f25ffdf58c03d79a3204c9563d",
        "sha256:b6179aac24823611ddfd3ef31fd7b3fbfc2fd14688aae9f6172d9495813f6148",
    ),
    "0964025147": (
        "sha256:0b0e104b11ff8253a747277793de74405a7fdc386e16eb5fdbdb1070f987adff",
        "sha256:03c4afb9ac9b5b8cb6839a3a6aba16624122b3441fac46e526d1482def6ff123",
    ),
    "0e2d3059c2": (
        "sha256:6aaba7b9c8770314aad6f7d663c14dce42e067a4550db6b840376a30241bb0ee",
        "sha256:0c8351f2b455cee5ee1282dfa89dad338e23822256b2d21f6ee140d24c424e1c",
    ),
    "4f8a6d549a": (
        "sha256:6bcb73549e292f38077eae3c58460f7dfdba8fcca30b98dfca877f50df3db942",
        "sha256:b17ef8b15693d5d7bf442f0d687760acab8430a8f43c4dfa06b569bb150dc30b",
    ),
    "838adf447e": (
        "sha256:210929bec7bae4d7a543c757433ed17c899d3e4954287d62bc2d4abd79f66f27",
        "sha256:c7dfa44ae0c7bd167695e7bf9733f17f177123d6fd6f79737def15f457c6c309",
    ),
    "85ca3806a6": (
        "sha256:318a84e0a0c0538431c786d57e703fbedb250f73c827f64dd3a7bb9c10516f4c",
        "sha256:eed4a445f10443e888e209617360460515aec5853e2f6776d2902169e817ff4f",
    ),
    "8d69cd5de3": (
        "sha256:b1606be401951da72601121999e0b1d7c1cdff55233042034e1782d034b7be65",
        "sha256:d5c4badaa71b2e0276429d0f4eda91f4b7cd5bdb20575ed23add23b701807283",
    ),
    "9bdd94cf9f": (
        "sha256:4a66c27d8195de8f56c4bd56e7aa71a59c8f14af2871e0484e478ac2c2fd0ab7",
        "sha256:d5e35a3507d3a5158f771e05bd5d5ebf24145f54277768ba8476b329be83cafc",
    ),
    "ae9d5bfaeb": (
        "sha256:f4c15999645b331351e1e1566dd304471deaf2ce9ac5217f9797c9571cc9493e",
        "sha256:5db2555cebcac9faed03391b49a11fcbf6b09e263a71f4fda7c734104a07aa58",
    ),
    "27caf21795": (
        "sha256:c7b50fe5c764e257e4d6e804d463b82e1ff429a8873dab4efb9e00f322c73f88",
        "sha256:c1452f939e9fb729c0de092f737145343a49fcc95dc788391d32cd2d61b9fb3e",
    ),
    "37a9caf571": (
        "sha256:d3a0b61f613562a4a19624f08e0ecae4da325406b31b07f3a56509fef28ead93",
        "sha256:55666c7cf53ebe50beb34c6e1bfb008de9dce4ae84105fdec550c703bf4ca357",
    ),
    "59939c8e3d": (
        "sha256:2de02c2bdc7117c3aaebde749c419a718d441df284a2751b788cb266c883724a",
        "sha256:e98cf47619400bc63061984f4bd1744040a75a7f55834e969bfaba64dbcfcb53",
    ),
    "6e7986a831": (
        "sha256:32f7632f86e3cfbfba009d414a173675f7059ca96cd25f9d122da2942d497084",
        "sha256:e5ad43ffd1ba8187d1bd0d2bb9e5e66e33a3df47b69edf5af1f63298a88d078a",
    ),
    "7ced84ab33": (
        "sha256:bc2daf4b9d7aa45b4b2ba68498171a70565d13a31cf29be0e414fc691252959b",
        "sha256:192327cb40f629ed72155fc311ed0a8e54e2c5e29b22a66a318d88d485c2e281",
    ),
    "827a330079": (
        "sha256:6562de2bdb1c037b9021055d5418d29237988a9dcd5b1b81a457a1ac07f2f478",
        "sha256:b3c5b918e79d52bb8e4b3e01700df5d9558c5ef5ec27f5c85e6f88e8a6645845",
    ),
    "829327a2a7": (
        "sha256:64a0b56166e10965f75340afac3ef1cef1ceef0dfc88467755f23487fc77720f",
        "sha256:296384f8ab4cc1761327ec4553e572c2bfd9aa4ab330bd6d23514da684eaa68a",
    ),
    "88e189a36b": (
        "sha256:3e62541f3d828c191726b4bb77654d0f150648d3134786ad57fcdd568bb698d7",
        "sha256:8fafa8e23a5695a26a8efcdf406452f93230ea76863fedb1c4fb04ff7c13adf6",
    ),
    "8c70b71645": (
        "sha256:a6609027f2273e56b443cd3c01ca2f4a73f3aeb25bf5233394aabc370485f19f",
        "sha256:c4ba0785d4ae87f7453c685bf6286bd70b655244506dea29311b78ba3becede5",
    ),
}

FDV_BREACH_MEMBERSHIP_SOURCES = {
    2021: (
        "2021-22",
        "workbooks/criminal-courts-2021-22-cube-15-source.xlsx",
        "sha256:d6020cde88de403d1c86dee430db6000949d7eb87c29c05f0e9dd673f2149ac5",
    ),
    2022: (
        "2022-23",
        "workbooks/criminal-courts-2022-23-cube-16-source.xlsx",
        "sha256:bece9e301160f84a1c98b1d3dc8842a53812211954b049f8004d21579729162d",
    ),
    2023: (
        "2023-24",
        "workbooks/criminal-courts-2023-24-cube-17-source.xlsx",
        "sha256:f5780d562b078756add08d13afe3a27413c1dd1c9eb9d188d77596f6b6c43a73",
    ),
    2024: (
        "2024-25",
        "workbooks/criminal-courts-2024-25-cube-17-source.xlsx",
        "sha256:a7de3c6f4ee210460ffcadcdded5f4d1b654220c97b3037383b88f2e9d8d8289",
    ),
}

# Canonical JSON of every distinct (raw category, canonical category) pair for
# each era/group. These pins freeze the complete age, Indigenous-status,
# method, sentence, outcome, court-level, and sex taxonomies without deriving
# expected truth from accepted contracts or evidence.
FDV_BREACH_TAXONOMY_DIGESTS = {
    "historical": {
        "GROUP_AGE": (
            15,
            "sha256:d32405a80a8406e4222e14d5ff19b426ac52a25f6bfa471f445802d3616969fb",
        ),
        "GROUP_COURT_LEVEL": (
            3,
            "sha256:70de1ff358058c5a95dfd778482a388c5eb1187febb51f5e5dfc258045b432ec",
        ),
        "GROUP_INDIGENOUS_STATUS": (
            8,
            "sha256:4cb88e876490fcbdd81f281eaa58293b90da855868d775a11a5585528f5e72dc",
        ),
        "GROUP_METHOD_OF_FINALISATION": (
            27,
            "sha256:5b7ab5b52b0bcd5dae850629fc1dc62df960529ff9fac5bfb39bbb6645e2c08b",
        ),
        "GROUP_NON_INDIGENOUS": (
            2,
            "sha256:f611cce13b55c5cf55c349205925c93bc9eabb8a74f2aad1e19f2898271d2f2c",
        ),
        "GROUP_NON_INDIGENOUS_AND_NOT_STATED": (
            2,
            "sha256:caa182ccf79e1c2dc39954ba08eafe016be99a0fb479a23b51717d7ef0ee063c",
        ),
        "GROUP_PRINCIPAL_SENTENCE": (
            37,
            "sha256:958c03e3d16e2a34b6b81159fa6594da51c55497e57aa051b44d4fb34dc54c63",
        ),
        "GROUP_SEX": (
            2,
            "sha256:135173d2d0ec2c37e986678b5ade52a858e832f752ac3ba91888eccfe8f45eb1",
        ),
    },
    "2024": {
        "GROUP_AGE": (
            13,
            "sha256:ad5219e83aaaded3036bd888631573e74c2d6426cc9ce10f42014f5460954ba9",
        ),
        "GROUP_COURT_LEVEL": (
            3,
            "sha256:70de1ff358058c5a95dfd778482a388c5eb1187febb51f5e5dfc258045b432ec",
        ),
        "GROUP_INDIGENOUS_STATUS": (
            8,
            "sha256:def45b7b369cd7218f61838fc914a24d36780fd5c40fd6ad70b541934f5edde6",
        ),
        "GROUP_METHOD_OF_FINALISATION": (
            22,
            "sha256:6e9ab7cb81d3c38db202c955d31251b47deb1e2fe1a5cd59b98aabc93d5c3967",
        ),
        "GROUP_PRINCIPAL_SENTENCE": (
            33,
            "sha256:210fa5a6a21b79f0532a6c671c06c4981acbf82daa680dba6d14a7d27c5f304b",
        ),
        "GROUP_SEX": (
            2,
            "sha256:135173d2d0ec2c37e986678b5ade52a858e832f752ac3ba91888eccfe8f45eb1",
        ),
    },
}

FDV_BREACH_PHYSICAL_FORMULAS = {
    (2021, "FDV Table 14"): (("A3", "=Contents!A3"), ("A48", "=Contents!B30")),
    (2021, "FDV Table 15"): (("A3", "=Contents!A3"), ("A52", "=Contents!B30")),
    (2021, "FDV Table 16"): (("A3", "=Contents!A3"), ("A48", "=Contents!B30")),
    (2021, "FDV Table 17"): (("A3", "=Contents!A3"), ("A52", "=Contents!B30")),
    (2021, "FDV Table 18"): (("A3", "=Contents!A3"), ("A52", "=Contents!B30")),
    (2021, "FDV Table 19"): (("A3", "=Contents!A3"), ("A48", "=Contents!B30")),
    (2021, "FDV Table 20"): (("A3", "=Contents!A3"), ("A52", "=Contents!B30")),
    (2021, "FDV Table 21"): (("A3", "=Contents!A3"), ("A52", "=Contents!B30")),
    (2021, "FDV Table 22"): (("A3", "=Contents!A3"), ("A52", "=Contents!B30")),
    (2022, "FDV Table 14"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A50", "=Contents!B30"),
    ),
    (2022, "FDV Table 15"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A54", "=Contents!B30"),
    ),
    (2022, "FDV Table 16"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A50", "=Contents!B30"),
    ),
    (2022, "FDV Table 17"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A54", "=Contents!B30"),
    ),
    (2022, "FDV Table 18"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A54", "=Contents!B30"),
    ),
    (2022, "FDV Table 19"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A50", "=Contents!B30"),
    ),
    (2022, "FDV Table 20"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A54", "=Contents!B30"),
    ),
    (2022, "FDV Table 21"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A54", "=Contents!B30"),
    ),
    (2022, "FDV Table 22"): (
        ("A2", "=Contents!A2"),
        ("A3", "=Contents!A3"),
        ("A54", "=Contents!B30"),
    ),
    (2023, "FDV Table 14"): (),
    (2023, "FDV Table 15"): (),
    (2023, "FDV Table 16"): (),
    (2023, "FDV Table 17"): (),
    (2023, "FDV Table 18"): (),
    (2023, "FDV Table 19"): (),
    (2023, "FDV Table 20"): (),
    (2023, "FDV Table 21"): (),
    (2023, "FDV Table 22"): (),
    (2024, "FDV Table 18"): (),
    (2024, "FDV Table 19"): (),
    (2024, "FDV Table 20"): (),
    (2024, "FDV Table 21"): (),
    (2024, "FDV Table 22"): (),
    (2024, "FDV Table 23"): (),
    (2024, "FDV Table 24"): (),
    (2024, "FDV Table 25"): (),
    (2024, "FDV Table 26"): (),
}

FDV_BREACH_CSV_FIELDS = (
    "publication_vintage_date",
    "reference_date",
    "characteristic_category_id",
    "characteristic_group_id",
    "observation_period_id",
    "statistic_basis_id",
    "jurisdiction_id",
    "classification_context_id",
    "measure_id",
    "unit_id",
    "value",
    "value_status",
    "raw_value",
    "source_workbook_digest",
    "source_sheet",
    "source_cell",
    "recipe_digest",
    "publication_id",
    "execution_digest",
    "acceptance_policy_version",
    "acceptance_policy_digest",
    "acceptance_decision_digest",
    "prompt_package_digest",
    "generation_model",
    "generation_attempt_id",
    "raw_characteristic_category",
    "raw_characteristic_group",
    "raw_observation_period",
    "raw_statistic_basis",
    "raw_jurisdiction",
    "raw_classification_context",
)


def _fdv_breach_matrix_by_family() -> dict[str, list[tuple[object, ...]]]:
    result: dict[str, list[tuple[object, ...]]] = {
        family_id: [] for family_id in FDV_BREACH_FAMILIES
    }
    for member in FDV_BREACH_SOURCE_MATRIX:
        result[str(member[0])].append(member)
    return result


def _fdv_breach_spec(family_id: str) -> LargeBatchSpec:
    members = _fdv_breach_matrix_by_family()[family_id]
    years = tuple(int(item[1]) for item in members)
    counts = tuple(int(item[8]) for item in members)
    measures = {
        "defendant-count": sum(int(item[9]) for item in members),
        "mean-age": sum(int(item[10]) for item in members),
        "median-age": sum(int(item[11]) for item in members),
    }
    statuses = Counter()
    for item in members:
        statuses.update(
            {
                "observed": int(item[12]),
                "not_available": int(item[13]),
                "not_applicable": int(item[14]),
            }
        )
    return LargeBatchSpec(
        family_id=family_id,
        label="Criminal Courts FDV order breaches",
        cohort_path=f"fixtures/product-prototype/{family_id}.json",
        evidence_manifest_path=(
            f"fixtures/product-prototype/{family_id}-evidence/manifest.json"
        ),
        dagster_asset=f"fdv_breach_{family_id[-10:]}",
        dagster_job=f"fdv_breach_{family_id[-10:]}_job",
        output_directory=f"fdv-breach-{family_id[-10:]}",
        expected_years=years,
        expected_year_counts=counts,
        expected_canonical_count=sum(counts),
        expected_excluded_observation_count=0,
        expected_excluded_observation_counts_by_year={year: 0 for year in years},
        expected_measure_counts=measures,
        expected_value_status_counts={
            key: value for key, value in statuses.items() if value
        },
        expected_manual_replay_years=years,
        preserves_publication_vintage=True,
        acceptance_policy_version="tidy.table-family-acceptance/v2",
        replay_recorded_at="2026-08-24T09:00:00+00:00",
    )


FDV_BREACH_RAW_FIELDS = {
    "characteristic_category": "raw_characteristic_category",
    "characteristic_group": "raw_characteristic_group",
    "observation_period": "raw_observation_period",
    "statistic_basis": "raw_statistic_basis",
    "jurisdiction": "raw_jurisdiction",
    "classification_context": "raw_classification_context",
}
FDV_BREACH_ACT_2023_FAMILY = (
    "criminal-courts-family-domestic-violence-family-and-domestic-violence-"
    "defendants-finalised-breach-of-violence-orders-summary-c-6e7986a831"
)
FDV_BREACH_ACT_2023_TITLE = (
    "FDV Table 22 Experimental data \N{EN DASH} Family and domestic violence "
    "defendants finalised, Breach of violence orders, Summary characteristics, "
    "Australian Capital Territory, 2018\N{EN DASH}19 to 2023\N{EN DASH}24 "
)
FDV_BREACH_TITLE_ROWS = {2021: 4, 2022: 4, 2023: 3, 2024: 2}


def _fdv_normalize_label(value: str) -> str:
    return " ".join(value.strip().split())


def _fdv_without_terminal_footnote(value: str) -> str:
    normalized = _fdv_normalize_label(value)
    while re.search(r"\s*\([a-z]\)$", normalized, re.IGNORECASE):
        normalized = re.sub(r"\s*\([a-z]\)$", "", normalized, flags=re.IGNORECASE)
    return normalized


def _fdv_slug(prefix: str, value: str) -> str:
    normalized = _fdv_without_terminal_footnote(value)
    normalized = (
        normalized.replace("&", " and ")
        .replace("n.e.c.", "nec")
        .replace("N.E.C.", "NEC")
        .replace("'", "")
        .replace("\N{RIGHT SINGLE QUOTATION MARK}", "")
    )
    slug = re.sub(r"[^A-Za-z0-9]+", "_", normalized).strip("_").upper()
    return f"{prefix}{slug}"


def _fdv_expected_alias_target(dimension: str, raw: str) -> str:
    if dimension == "characteristic_category":
        if _fdv_normalize_label(raw) == "Children's":
            return "CHAR_CHILDREN_S"
        return _fdv_slug("CHAR_", raw)
    if dimension == "characteristic_group":
        return _fdv_slug("GROUP_", raw)
    if dimension == "observation_period":
        match = re.fullmatch(
            r"([0-9]{4})\N{EN DASH}([0-9]{2})(?:\([a-z]\))?",
            _fdv_normalize_label(raw),
            re.IGNORECASE,
        )
        assert match is not None
        return f"20{match.group(2)}-06-30"
    if dimension == "statistic_basis":
        normalized = _fdv_without_terminal_footnote(raw)
        if normalized == "Mean (years)":
            return "MEAN_AGE"
        if normalized == "Median (years)":
            return "MEDIAN_AGE"
        return "COUNT"
    if dimension in {"jurisdiction", "classification_context"}:
        title = _fdv_normalize_label(raw)
        if dimension == "classification_context":
            if "Breach of restraining order \N{EN DASH} violence" in title:
                return "FDV_BREACH_OF_RESTRAINING_ORDER_VIOLENCE_EXPERIMENTAL"
            assert "Breach of violence orders" in title
            return "FDV_BREACH_OF_VIOLENCE_ORDERS_EXPERIMENTAL"
        jurisdictions = {
            "Australia": "AUS",
            "New South Wales": "NSW",
            "Victoria": "VIC",
            "Queensland": "QLD",
            "South Australia": "SA",
            "Western Australia": "WA",
            "Tasmania": "TAS",
            "Northern Territory": "NT",
            "Australian Capital Territory": "ACT",
        }
        matches = [
            canonical
            for published, canonical in jurisdictions.items()
            if f", {published}," in title
        ]
        assert len(matches) == 1
        return matches[0]
    raise AssertionError(f"Unknown FDV breach dimension: {dimension}")


def _assert_fdv_aliases_against_source_truth(
    contract: dict[str, object], rows: list[dict[str, object]]
) -> int:
    alias_count = 0
    for dimension, raw_field in FDV_BREACH_RAW_FIELDS.items():
        aliases = contract["aliases"][dimension]
        normalized: dict[str, str] = {}
        for raw, target in aliases.items():
            key = _fdv_normalize_label(raw)
            assert key not in normalized
            assert target == _fdv_expected_alias_target(dimension, raw)
            normalized[key] = target
            alias_count += 1
        used = {_fdv_normalize_label(str(row[raw_field])) for row in rows}
        assert set(normalized) == used
        for row in rows:
            raw = str(row[raw_field])
            assert row[f"{dimension}_id"] == _fdv_expected_alias_target(dimension, raw)
            assert normalized[_fdv_normalize_label(raw)] == row[f"{dimension}_id"]
    return alias_count


def _fdv_csv_serialized_scalar(field: str, value: object) -> str:
    """Mirror the named production CSV serialization without claiming byte identity."""
    if value is None:
        return ""
    if isinstance(value, str) and field != "source_sheet":
        return value.rstrip()
    if value is True:
        return "true"
    if value is False:
        return "false"
    return str(value)


def _assert_fdv_csv_matches_json(
    csv_path: Path, rows: list[dict[str, object]]
) -> list[tuple[str, str, str, str]]:
    assert rows
    with csv_path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        fields = reader.fieldnames
        assert fields == list(FDV_BREACH_CSV_FIELDS)
        csv_rows = list(reader)
    assert len(csv_rows) == len(rows)
    assert all(set(row) == set(FDV_BREACH_CSV_FIELDS) for row in rows)
    whitespace_differences: list[tuple[str, str, str, str]] = []
    for csv_row, row in zip(csv_rows, rows, strict=True):
        for field in fields:
            expected = _fdv_csv_serialized_scalar(field, row[field])
            assert csv_row[field] == expected
            if isinstance(row[field], str) and csv_row[field] != row[field]:
                assert row[field] != row[field].rstrip()
                assert csv_row[field] == row[field].rstrip()
                whitespace_differences.append(
                    (
                        str(row["publication_vintage_date"]),
                        field,
                        row[field],
                        csv_row[field],
                    )
                )
    return whitespace_differences


def _fdv_taxonomy_digest(
    rows: list[dict[str, object]], group_id: str
) -> tuple[int, str]:
    pairs = sorted(
        {
            (
                str(row["raw_characteristic_category"]),
                str(row["characteristic_category_id"]),
            )
            for row in rows
            if row["characteristic_group_id"] == group_id
        }
    )
    payload = json.dumps(pairs, separators=(",", ":"), ensure_ascii=False).encode()
    return len(pairs), sha256_digest(payload)


def _assert_exact_sha(data: bytes, expected: str) -> None:
    assert sha256_digest(data) == expected


def test_fdv_order_breach_cube_source_matrix_and_frozen_evidence() -> None:
    matrix_by_family = _fdv_breach_matrix_by_family()
    assert set(matrix_by_family) == set(FDV_BREACH_FAMILIES)
    assert len(FDV_BREACH_SOURCE_MATRIX) == 36
    assert Counter(len(items) for items in matrix_by_family.values()) == {1: 9, 3: 9}
    membership = _load(MEMBERSHIP)
    membership_by_family = {
        family["familyId"]: family for family in membership["families"]
    }
    all_rows: list[dict[str, object]] = []
    rows_by_release = Counter()
    aliases_declared = 0
    selected_formula_count = 0
    physical_formula_count = 0
    csv_whitespace_difference_count = 0
    csv_whitespace_differences: set[tuple[str, str, str, str]] = set()
    workbook_cache: dict[str, object] = {}
    data_cache: dict[str, object] = {}
    membership_workbook_cache: dict[str, object] = {}
    normalization_manifest = _load(FIXTURES / "batch-workbook-normalization-v1.json")
    normalization_entry = next(
        entry
        for entry in normalization_manifest["entries"]
        if entry["sourcePath"] == "fixtures/product-prototype/workbooks/"
        "criminal-courts-2023-24-cube-17-source.xlsx"
    )
    assert {
        key: normalization_entry[key]
        for key in (
            "sourceDigest",
            "sourceByteLength",
            "outputPath",
            "outputDigest",
            "outputByteLength",
            "trimmedSheets",
            "correction",
            "normalization",
        )
    } == {
        "sourceDigest": (
            "sha256:f5780d562b078756add08d13afe3a27413c1dd1c9eb9d188d77596f6b6c43a73"
        ),
        "sourceByteLength": 85082,
        "outputPath": (
            "fixtures/product-prototype/workbooks/"
            "criminal-courts-2023-24-cube-17-normalized.xlsx"
        ),
        "outputDigest": (
            "sha256:ac87ce9a09d4da630aa02b1d7e711ca801e3ac022986584ef1864e8370f9d580"
        ),
        "outputByteLength": 76428,
        "trimmedSheets": [
            {"sheet": "FDV Table 15", "retainedRange": "A1:G69"},
            {"sheet": "FDV Table 16", "retainedRange": "A1:G63"},
            {"sheet": "FDV Table 17", "retainedRange": "A1:G66"},
            {"sheet": "FDV Table 18", "retainedRange": "A1:F71"},
            {"sheet": "FDV Table 19", "retainedRange": "A1:G64"},
        ],
        "correction": {
            "id": "criminal-courts-fdv-2023-24-duplicate-footnote-v1",
            "reason": (
                "Remove the duplicate far-right Table 16 footnote at XEX59 before "
                "format trimming; preserve the identical retained footnote at A58 "
                "and the exact source workbook separately."
            ),
            "removedCells": [
                {
                    "sheet": "FDV Table 16",
                    "cell": "XEX59",
                    "expectedStyle": "67",
                    "expectedValue": (
                        "(f) Includes defendants for whom method of finalisation "
                        "could not be determined, defendants deceased or unfit to "
                        "plead, transfers to non-court agencies and other "
                        "non-adjudicated finalisations n.e.c. "
                    ),
                    "insideRetainedRange": False,
                }
            ],
        },
        "normalization": "trim-pathological-styled-blank-cells-v1",
    }
    publication_dates = {
        2021: "2022-06-30",
        2022: "2023-06-30",
        2023: "2024-06-30",
        2024: "2025-06-30",
    }
    markers = {
        "na": "not_available",
        "n.a.": "not_available",
        "..": "not_applicable",
        "np": "suppressed",
        "n.p.": "suppressed",
    }
    try:
        for family_id in FDV_BREACH_FAMILIES:
            matrix = matrix_by_family[family_id]
            cohort_path = FIXTURES / f"{family_id}.json"
            contract_path = FIXTURES / "acceptance" / f"{family_id}-v1.json"
            evidence = FIXTURES / f"{family_id}-evidence"
            identity_pins = FDV_BREACH_IDENTITY_PINS[family_id[-10:]]
            _assert_exact_sha(contract_path.read_bytes(), identity_pins[0])
            _assert_exact_sha(cohort_path.read_bytes(), identity_pins[1])
            cohort = _load(cohort_path)
            contract = _load(contract_path)
            run = _load(evidence / "run.json")
            rows = json.loads((evidence / "canonical-observations.json").read_text())
            differences = _assert_fdv_csv_matches_json(
                evidence / "canonical-observations.csv", rows
            )
            csv_whitespace_difference_count += len(differences)
            for date, field, json_value, csv_value in differences:
                assert csv_value == json_value.rstrip()
                csv_whitespace_differences.add((family_id, date, field, json_value))
            aliases_declared += _assert_fdv_aliases_against_source_truth(contract, rows)
            verify_large_batch_evidence(PROJECT, _fdv_breach_spec(family_id))
            product_prototype_module._validate_cohort(cohort)
            product_prototype_module._validate_contract(contract, cohort)
            assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
            assert contract["strictAliasMatching"] is True
            assert contract["trainingEligibility"] is False
            assert contract["totalEquations"] == []
            assert contract["totalValidation"] == "not_applicable"
            assert contract["allowedExecutionWarnings"] == []
            assert contract["expectedWarningCountsByYear"] == {
                str(item[1]): 0 for item in matrix
            }
            assert run["providerCalls"] == 0
            assert run["exceptionWorkbookCount"] == 0
            assert run["trainingEligibility"] is False
            assert run["historicalReplayIsAcceptanceAuthority"] is False
            assert all(
                workbook["executionWarningCount"] == 0
                and workbook["issues"] == []
                and all(workbook["checks"].values())
                for workbook in run["workbooks"]
            )
            family_members = [
                member
                for member in membership_by_family[family_id]["members"]
                if member["cubeId"]
                == "family-and-domestic-violence-order-breaches-experimental-data"
            ]
            assert len(family_members) == len(matrix)
            membership_by_release = {
                member["releaseId"]: member for member in family_members
            }
            assert len(membership_by_release) == len(family_members)
            assert set(membership_by_release) == {
                FDV_BREACH_MEMBERSHIP_SOURCES[int(item[1])][0] for item in matrix
            }
            cohort_by_year = {item["year"]: item for item in cohort["workbooks"]}
            rows_by_year: dict[int, list[dict[str, object]]] = {
                int(item[1]): [] for item in matrix
            }
            digest_to_year = {str(item[3]): int(item[1]) for item in matrix}
            for row in rows:
                year = digest_to_year[str(row["source_workbook_digest"])]
                rows_by_year[year].append(row)
                assert row["publication_vintage_date"] == publication_dates[year]
                assert row["acceptance_policy_version"] == (
                    "tidy.table-family-acceptance/v2"
                )
                assert row["acceptance_policy_digest"] == identity_pins[0]
                assert (
                    row["recipe_digest"]
                    == contract["expectedRecipeDigestsByYear"][str(year)]
                )
            for item in matrix:
                (
                    _,
                    year,
                    workbook_path,
                    workbook_digest,
                    workbook_length,
                    sheet,
                    max_row,
                    max_column,
                    expected_rows,
                    count_rows,
                    mean_rows,
                    median_rows,
                    observed_rows,
                    na_rows,
                    not_applicable_rows,
                    zeros,
                    map_path,
                    map_digest,
                    map_length,
                ) = item
                release_id, member_source_path, member_source_digest = (
                    FDV_BREACH_MEMBERSHIP_SOURCES[year]
                )
                member = membership_by_release[release_id]
                assert {
                    "releaseId": member["releaseId"],
                    "cubeId": member["cubeId"],
                    "tableNamespace": member["tableNamespace"],
                    "physicalSheetName": member["physicalSheetName"],
                    "physicalTableNumber": member["physicalTableNumber"],
                    "sourcePath": member["sourcePath"],
                    "sourceDigest": member["sourceDigest"],
                    "classificationContext": member["classificationContext"],
                    "registered": member["registered"],
                } == {
                    "releaseId": release_id,
                    "cubeId": (
                        "family-and-domestic-violence-order-breaches-experimental-data"
                    ),
                    "tableNamespace": "family-domestic-violence",
                    "physicalSheetName": sheet,
                    "physicalTableNumber": int(sheet.rsplit(" ", 1)[1]),
                    "sourcePath": member_source_path,
                    "sourceDigest": member_source_digest,
                    "classificationContext": "experimental-fdv-publication-namespace",
                    "registered": True,
                }
                member_source = FIXTURES / member_source_path
                _assert_exact_sha(member_source.read_bytes(), member_source_digest)
                member_key = str(member_source)
                if member_key not in membership_workbook_cache:
                    membership_workbook_cache[member_key] = load_workbook(
                        member_source, read_only=False, data_only=False
                    )
                assert (
                    member["publishedTitle"]
                    == membership_workbook_cache[member_key][sheet]
                    .cell(FDV_BREACH_TITLE_ROWS[year], 1)
                    .value
                )

                workbook_decl = cohort_by_year[year]
                assert workbook_decl["path"] == workbook_path
                assert workbook_decl["contentDigest"] == workbook_digest
                assert workbook_decl["byteLength"] == workbook_length
                assert workbook_decl["sheet"] == sheet
                if year == 2023:
                    assert (
                        workbook_decl["normalization"]
                        == normalization_entry["normalization"]
                    )
                    assert (
                        workbook_decl["path"]
                        == Path(normalization_entry["outputPath"])
                        .relative_to("fixtures/product-prototype")
                        .as_posix()
                    )
                    assert (
                        workbook_decl["contentDigest"]
                        == normalization_entry["outputDigest"]
                    )
                    assert (
                        workbook_decl["byteLength"]
                        == normalization_entry["outputByteLength"]
                    )
                    assert (
                        member["sourcePath"]
                        == Path(normalization_entry["sourcePath"])
                        .relative_to("fixtures/product-prototype")
                        .as_posix()
                    )
                    assert member["sourceDigest"] == normalization_entry["sourceDigest"]
                else:
                    assert "normalization" not in workbook_decl
                assert workbook_decl["replayResponse"] == {
                    "path": map_path,
                    "contentDigest": map_digest,
                    "byteLength": map_length,
                    "historicalModel": "human-authored/deterministic-map-v1",
                    "acceptanceAuthority": False,
                }
                source_path = FIXTURES / workbook_path
                _assert_exact_sha(source_path.read_bytes(), workbook_digest)
                assert source_path.stat().st_size == workbook_length
                replay_path = FIXTURES / map_path
                _assert_exact_sha(replay_path.read_bytes(), map_digest)
                assert replay_path.stat().st_size == map_length
                assert json.loads(replay_path.read_text())["version"] == (
                    "semantic-table-map-v1"
                )
                key = str(source_path)
                if key not in workbook_cache:
                    workbook_cache[key] = load_workbook(
                        source_path, read_only=False, data_only=False
                    )
                    data_cache[key] = load_workbook(
                        source_path, read_only=False, data_only=True
                    )
                worksheet = workbook_cache[key][sheet]
                data_sheet = data_cache[key][sheet]
                assert (worksheet.max_row, worksheet.max_column) == (
                    max_row,
                    max_column,
                )
                physical_formulas = tuple(
                    (cell.coordinate, cell.value)
                    for source_row in worksheet.iter_rows()
                    for cell in source_row
                    if cell.data_type == "f"
                )
                assert physical_formulas == FDV_BREACH_PHYSICAL_FORMULAS[(year, sheet)]
                physical_formula_count += len(physical_formulas)
                year_rows = rows_by_year[year]
                assert len(year_rows) == expected_rows
                assert Counter(row["measure_id"] for row in year_rows) == {
                    "defendant-count": count_rows,
                    "mean-age": mean_rows,
                    "median-age": median_rows,
                }
                expected_statuses = {
                    key: value
                    for key, value in {
                        "observed": observed_rows,
                        "not_available": na_rows,
                        "not_applicable": not_applicable_rows,
                    }.items()
                    if value
                }
                assert Counter(row["value_status"] for row in year_rows) == (
                    expected_statuses
                )
                assert (
                    sum(
                        row["raw_value"] == 0 and not isinstance(row["raw_value"], bool)
                        for row in year_rows
                    )
                    == zeros
                )
                selected = {row["source_cell"] for row in year_rows}
                candidates = {
                    f"R{cell.row}C{cell.column}"
                    for source_row in worksheet.iter_rows()
                    for cell in source_row
                    if (
                        isinstance(cell.value, int | float)
                        and not isinstance(cell.value, bool)
                    )
                    or (
                        isinstance(cell.value, str)
                        and cell.value.strip().lower() in markers
                    )
                }
                assert selected == candidates
                formula_sources = {
                    f"R{cell.row}C{cell.column}"
                    for source_row in worksheet.iter_rows()
                    for cell in source_row
                    if cell.data_type == "f"
                }
                assert selected.isdisjoint(formula_sources)
                source_strings = {
                    cell.value
                    for source in (worksheet, data_sheet)
                    for source_row in source.iter_rows()
                    for cell in source_row
                    if isinstance(cell.value, str)
                }
                for row in year_rows:
                    source_row, source_column = _r1c1_parts(str(row["source_cell"]))
                    cell = worksheet.cell(source_row, source_column)
                    assert cell.value == row["raw_value"]
                    selected_formula_count += cell.data_type == "f"
                    assert cell.data_type != "f"
                    if row["value_status"] == "observed":
                        assert cell.data_type == "n"
                        assert type(cell.value) is type(row["raw_value"])
                        assert row["value"] == row["raw_value"]
                    else:
                        assert cell.data_type == "s"
                        assert (
                            markers[str(cell.value).strip().lower()]
                            == (row["value_status"])
                        )
                        assert row["value"] is None
                    assert (
                        worksheet.cell(source_row, 1).value
                        == row["raw_characteristic_category"]
                    )
                    assert (
                        row["raw_statistic_basis"] == row["raw_characteristic_category"]
                    )
                    assert all(
                        row[field] in source_strings
                        for field in FDV_BREACH_RAW_FIELDS.values()
                    )
                rows_by_release[year] += len(year_rows)
            all_rows.extend(rows)
    finally:
        for workbook in (
            list(workbook_cache.values())
            + list(data_cache.values())
            + list(membership_workbook_cache.values())
        ):
            workbook.close()

    assert aliases_declared == 1882
    assert physical_formula_count == 45
    assert selected_formula_count == 0
    assert len(all_rows) == 6463
    assert rows_by_release == {2021: 1054, 2022: 1465, 2023: 1826, 2024: 2118}
    assert Counter(row["measure_id"] for row in all_rows) == {
        "defendant-count": 6097,
        "mean-age": 183,
        "median-age": 183,
    }
    assert Counter(row["value_status"] for row in all_rows) == {
        "observed": 6357,
        "not_available": 94,
        "not_applicable": 12,
    }
    assert (
        sum(
            row["raw_value"] == 0 and not isinstance(row["raw_value"], bool)
            for row in all_rows
        )
        == 227
    )
    assert (
        len(
            {
                (row["source_workbook_digest"], row["source_sheet"], row["source_cell"])
                for row in all_rows
            }
        )
        == 6463
    )
    assert {
        row["classification_context_id"]
        for row in all_rows
        if row["publication_vintage_date"] < "2025-01-01"
    } == {"FDV_BREACH_OF_VIOLENCE_ORDERS_EXPERIMENTAL"}
    historical_rows = [
        row for row in all_rows if row["publication_vintage_date"] < "2025-01-01"
    ]
    rows_2024 = [
        row for row in all_rows if row["publication_vintage_date"] == "2025-06-30"
    ]
    assert {row["classification_context_id"] for row in rows_2024} == {
        "FDV_BREACH_OF_RESTRAINING_ORDER_VIOLENCE_EXPERIMENTAL"
    }
    for era, era_rows in (("historical", historical_rows), ("2024", rows_2024)):
        assert {str(row["characteristic_group_id"]) for row in era_rows} == set(
            FDV_BREACH_TAXONOMY_DIGESTS[era]
        )
        for group_id, expected in FDV_BREACH_TAXONOMY_DIGESTS[era].items():
            assert _fdv_taxonomy_digest(era_rows, group_id) == expected
    assert {
        row["characteristic_category_id"]
        for row in historical_rows
        if row["characteristic_group_id"] == "GROUP_AGE"
    } == {
        "CHAR_10_19_YEARS",
        "CHAR_20_29_YEARS",
        "CHAR_30_39_YEARS",
        "CHAR_40_49_YEARS",
        "CHAR_50_59_YEARS",
        "CHAR_60_YEARS_AND_OVER",
        "CHAR_MEAN_YEARS",
        "CHAR_MEDIAN_YEARS",
        "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
    }
    assert {
        row["characteristic_category_id"]
        for row in rows_2024
        if row["characteristic_group_id"] == "GROUP_AGE"
    } == {
        "CHAR_10_19_YEARS",
        "CHAR_20_24_YEARS",
        "CHAR_25_29_YEARS",
        "CHAR_30_34_YEARS",
        "CHAR_35_39_YEARS",
        "CHAR_40_44_YEARS",
        "CHAR_45_49_YEARS",
        "CHAR_50_54_YEARS",
        "CHAR_55_YEARS_AND_OVER",
        "CHAR_MEAN_YEARS",
        "CHAR_MEDIAN_YEARS",
        "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
    }
    assert {
        row["characteristic_category_id"]
        for row in historical_rows
        if row["characteristic_group_id"] == "GROUP_INDIGENOUS_STATUS"
    } == {
        "CHAR_ABORIGINAL_AND_TORRES_STRAIT_ISLANDER",
        "CHAR_NON_INDIGENOUS",
        "CHAR_NON_INDIGENOUS_AND_NOT_STATED",
        "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
    }
    assert {
        row["characteristic_category_id"]
        for row in rows_2024
        if row["characteristic_group_id"] == "GROUP_INDIGENOUS_STATUS"
    } == {
        "CHAR_ABORIGINAL_AND_TORRES_STRAIT_ISLANDER",
        "CHAR_NON_INDIGENOUS",
        "CHAR_NOT_STATED",
        "CHAR_TOTAL_FINALISED_EXCLUDING_TRANSFER_TO_OTHER_COURT_LEVELS",
    }
    act_2023_rows = [
        row
        for row in all_rows
        if row["source_sheet"] == "FDV Table 22"
        and row["publication_vintage_date"] == "2024-06-30"
        and row["jurisdiction_id"] == "ACT"
    ]
    assert act_2023_rows
    assert {row["raw_jurisdiction"] for row in act_2023_rows} == {
        FDV_BREACH_ACT_2023_TITLE
    }
    assert {row["raw_classification_context"] for row in act_2023_rows} == {
        FDV_BREACH_ACT_2023_TITLE
    }
    assert csv_whitespace_difference_count == 3087
    assert len(csv_whitespace_differences) == 446
    assert {item[2] for item in csv_whitespace_differences} == {
        "raw_characteristic_category",
        "raw_characteristic_group",
        "raw_statistic_basis",
        "raw_jurisdiction",
        "raw_classification_context",
    }
    whitespace_payload = json.dumps(
        sorted(csv_whitespace_differences),
        separators=(",", ":"),
        ensure_ascii=False,
    ).encode()
    assert sha256_digest(whitespace_payload) == (
        "sha256:6df3b55535510df2046e654b8b9063e495409693ba5f169ce7990ffbbf518aeb"
    )
    assert {
        item
        for item in csv_whitespace_differences
        if item[2] in {"raw_jurisdiction", "raw_classification_context"}
    } == {
        (
            FDV_BREACH_ACT_2023_FAMILY,
            "2024-06-30",
            "raw_jurisdiction",
            FDV_BREACH_ACT_2023_TITLE,
        ),
        (
            FDV_BREACH_ACT_2023_FAMILY,
            "2024-06-30",
            "raw_classification_context",
            FDV_BREACH_ACT_2023_TITLE,
        ),
    }


def test_fdv_breach_source_truth_rejects_coordinated_and_digest_mutations() -> None:
    family_id = next(
        item for item in FDV_BREACH_FAMILIES if item.endswith("88e189a36b")
    )
    contract_path = FIXTURES / "acceptance" / f"{family_id}-v1.json"
    cohort_path = FIXTURES / f"{family_id}.json"
    evidence = FIXTURES / f"{family_id}-evidence"
    contract = _load(contract_path)
    cohort = _load(cohort_path)
    rows = json.loads((evidence / "canonical-observations.json").read_text())
    title = str(rows[0]["raw_jurisdiction"])
    mutation_cases = (
        ("characteristic_category", "20\N{EN DASH}29 years", "CHAR_30_39_YEARS"),
        ("characteristic_group", "Age", "GROUP_SEX"),
        ("observation_period", "2018\N{EN DASH}19", "2020-06-30"),
        ("statistic_basis", "Mean (years)", "MEDIAN_AGE"),
        ("jurisdiction", title, "AUS"),
        (
            "classification_context",
            title,
            "FDV_BREACH_OF_RESTRAINING_ORDER_VIOLENCE_EXPERIMENTAL",
        ),
    )
    for dimension, normalized_raw, wrong_target in mutation_cases:
        mutated_contract = copy.deepcopy(contract)
        mutated_rows = copy.deepcopy(rows)
        alias_raw = next(
            raw
            for raw in mutated_contract["aliases"][dimension]
            if _fdv_normalize_label(raw) == _fdv_normalize_label(normalized_raw)
        )
        mutated_contract["aliases"][dimension][alias_raw] = wrong_target
        raw_field = FDV_BREACH_RAW_FIELDS[dimension]
        id_field = f"{dimension}_id"
        changed = 0
        for row in mutated_rows:
            if _fdv_normalize_label(str(row[raw_field])) == _fdv_normalize_label(
                normalized_raw
            ):
                row[id_field] = wrong_target
                changed += 1
        assert changed > 0
        with pytest.raises(AssertionError):
            _assert_fdv_aliases_against_source_truth(mutated_contract, mutated_rows)

    colliding_contract = copy.deepcopy(contract)
    colliding_contract["aliases"]["characteristic_category"]["  Males  "] = "CHAR_MALES"
    with pytest.raises(AssertionError):
        _assert_fdv_aliases_against_source_truth(colliding_contract, rows)

    matrix = _fdv_breach_matrix_by_family()[family_id]
    replay_path = FIXTURES / str(matrix[0][16])
    identity_pins = FDV_BREACH_IDENTITY_PINS[family_id[-10:]]
    for source, expected in (
        (contract_path, identity_pins[0]),
        (cohort_path, identity_pins[1]),
        (replay_path, str(matrix[0][17])),
    ):
        tampered = source.read_bytes() + b"tampered"
        with pytest.raises(AssertionError):
            _assert_exact_sha(tampered, expected)
    assert cohort["workbooks"][0]["replayResponse"]["acceptanceAuthority"] is False
