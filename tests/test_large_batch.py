from __future__ import annotations

import copy
import csv
import hashlib
import io
import json
import re
import runpy
import shutil
import subprocess
import zipfile
from collections import Counter
from collections.abc import Callable
from pathlib import Path
from xml.etree import ElementTree

import jsonschema
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import tidy_orchestrator.large_batch as large_batch_module
import tidy_orchestrator.product_prototype as product_prototype_module
from tidy_orchestrator.artifacts import (
    DecisionRecord,
    canonical_json_bytes,
    domain_digest,
    sha256_digest,
)
from tidy_orchestrator.large_batch import (
    LargeBatchError,
    load_large_batch_registry,
    verify_batch_normalization,
    verify_large_batch_complete_reproduction,
    verify_large_batch_evidence,
    verify_large_batch_reproduction,
)
from tidy_orchestrator.large_batch_cli import run_batch
from tidy_orchestrator.product_prototype import (
    RUN_SCHEMA,
    V2_REFERENCE_DATE_DECISION_IDENTITY,
    ProductPrototypeError,
    _acceptance_decision_payload,
    _validate_cohort,
    _validate_contract,
    _validate_expected_coverage,
)

PROJECT = Path(__file__).parents[1]


@pytest.fixture(scope="module", autouse=True)
def ensure_domain_worker_is_built() -> None:
    worker = PROJECT / "dist/tidy-domain-worker.cjs"
    if not worker.is_file():
        subprocess.run(
            ["npm", "run", "build"],
            cwd=PROJECT,
            check=True,
            capture_output=True,
            text=True,
            timeout=120,
        )


def test_large_batch_registry_and_all_evidence_close() -> None:
    registry = load_large_batch_registry(PROJECT)
    assert registry.batch_id == "justice-five-hundred-fifty-nine-worksheets-v1"
    assert registry.worksheet_count == 559
    assert registry.provider_calls == 0
    assert len(registry.entries) == 220
    normalization = verify_batch_normalization(PROJECT, registry)
    assert len(normalization["entries"]) == 64
    assert "normalization" not in normalization
    assert Counter(entry["normalization"] for entry in normalization["entries"]) == {
        "trim-pathological-styled-blank-cells-v1": 62,
        "trim-pathological-full-width-formatting-merge-v1": 2,
    }
    assert normalization["inRangeValuesChanged"] is True
    removed_cells = [
        cell
        for entry in normalization["entries"]
        if entry["correction"] is not None
        for cell in entry["correction"].get("removedCells", [])
    ]
    replaced_cells = [
        cell
        for entry in normalization["entries"]
        if entry["correction"] is not None
        for cell in entry["correction"].get("replacedCells", [])
    ]
    assert sum(cell["insideRetainedRange"] for cell in removed_cells) == 2
    assert {
        (entry["sourcePath"], cell["sheet"], cell["cell"])
        for entry in normalization["entries"]
        if entry["correction"] is not None
        for cell in entry["correction"].get("removedCells", [])
        if not cell["insideRetainedRange"]
    } == {
        (
            "fixtures/product-prototype/workbooks/"
            "recorded-crime-offenders-2024-25-source.xlsx",
            "Table 4",
            "XFC50",
        ),
        (
            "fixtures/product-prototype/workbooks/"
            "criminal-courts-2023-24-cube-17-source.xlsx",
            "FDV Table 16",
            "XEX59",
        ),
    }
    assert sum(cell["insideRetainedRange"] for cell in replaced_cells) == 1
    manifests = [
        verify_large_batch_evidence(PROJECT, spec) for spec in registry.entries
    ]
    assert sum(item["acceptedWorkbookCount"] for item in manifests) == 559
    assert sum(item["exceptionWorkbookCount"] for item in manifests) == 0
    assert sum(item["canonicalObservationCount"] for item in manifests) == 478120
    assert sum(item["providerCalls"] for item in manifests) == 0
    criminal_specs = [
        spec
        for spec in registry.entries
        if spec.family_id.startswith("criminal-courts-")
    ]
    assert len(criminal_specs) == 193
    assert sum(len(spec.expected_years) for spec in criminal_specs) == 430
    assert sum(spec.expected_canonical_count for spec in criminal_specs) == 422_103
    assert Counter(spec.acceptance_policy_version for spec in registry.entries) == {
        "tidy.table-family-acceptance/v1": 125,
        "tidy.table-family-acceptance/v2": 95,
    }
    nt_manifests = [
        item
        for item in manifests
        if item["familyId"].endswith(
            (
                "9e1eae4d24",
                "2dc791882d",
                "dff73e6680",
                "d4b9910476",
                "fb8ea665a2",
                "7b3957ee89",
                "2e3b1c96f5",
                "27c9d31040",
                "cd3b98cdfb",
            )
        )
    ]
    assert len(nt_manifests) == 9
    assert sum(item["acceptedWorkbookCount"] for item in nt_manifests) == 22
    assert sum(item["canonicalObservationCount"] for item in nt_manifests) == 16931
    assert sum(item["providerCalls"] for item in nt_manifests) == 0
    act_suffixes = (
        "4200e414a2",
        "79856bc0b9",
        "0b4eef6926",
        "7c61d6c40e",
        "6c72c5eba2",
        "139aaf92e0",
        "5705af16d6",
        "1f84e9447d",
        "b377949ac0",
    )
    act_manifests = [
        item for item in manifests if item["familyId"].endswith(act_suffixes)
    ]
    assert len(act_manifests) == 9
    assert sum(item["acceptedWorkbookCount"] for item in act_manifests) == 22
    assert sum(item["canonicalObservationCount"] for item in act_manifests) == 16057
    assert sum(item["providerCalls"] for item in act_manifests) == 0
    tas_suffixes = (
        "1e1718730a",
        "f2593de546",
        "cdc489d600",
        "34dbc091f5",
        "08dd480268",
        "0a5e590f31",
        "0cd32616d1",
        "1d97a0925b",
        "4a82019ceb",
        "d897254a26",
    )
    tas_manifests = [
        item for item in manifests if item["familyId"].endswith(tas_suffixes)
    ]
    assert len(tas_manifests) == 10
    assert sum(item["acceptedWorkbookCount"] for item in tas_manifests) == 22
    assert sum(item["canonicalObservationCount"] for item in tas_manifests) == 16545
    assert sum(item["providerCalls"] for item in tas_manifests) == 0
    fdv_breach_suffixes = (
        "00438881d8",
        "0964025147",
        "0e2d3059c2",
        "4f8a6d549a",
        "838adf447e",
        "85ca3806a6",
        "8d69cd5de3",
        "9bdd94cf9f",
        "ae9d5bfaeb",
        "27caf21795",
        "37a9caf571",
        "59939c8e3d",
        "6e7986a831",
        "7ced84ab33",
        "827a330079",
        "829327a2a7",
        "88e189a36b",
        "8c70b71645",
    )
    fdv_breach_manifests = [
        item for item in manifests if item["familyId"].endswith(fdv_breach_suffixes)
    ]
    assert len(fdv_breach_manifests) == 18
    assert sum(item["acceptedWorkbookCount"] for item in fdv_breach_manifests) == 36
    assert (
        sum(item["canonicalObservationCount"] for item in fdv_breach_manifests) == 6463
    )
    assert sum(item["providerCalls"] for item in fdv_breach_manifests) == 0
    fdv_offence_manifests = [
        item
        for item in manifests
        if item["familyId"].startswith("criminal-courts-family-domestic-violence-")
        and "breach-of-" not in item["familyId"]
    ]
    assert len(fdv_offence_manifests) == 31
    assert sum(item["acceptedWorkbookCount"] for item in fdv_offence_manifests) == 56
    assert (
        sum(item["canonicalObservationCount"] for item in fdv_offence_manifests)
        == 76_189
    )
    assert sum(item["providerCalls"] for item in fdv_offence_manifests) == 0
    offender_manifests = [
        item for item in manifests if item["familyId"].startswith("offenders-table-")
    ]
    assert len(offender_manifests) == 5
    assert all(
        item["manualReplayYears"] == [2021, 2022, 2023, 2024]
        and item["publicationVintagePreserved"] is True
        for item in offender_manifests
    )


def test_large_batch_cohorts_and_contracts_validate() -> None:
    registry = load_large_batch_registry(PROJECT)
    cohort_schema = json.loads(
        (PROJECT / "contracts/product-prototype/v1/cohort.schema.json").read_text()
    )
    validator = jsonschema.Draft202012Validator(
        cohort_schema,
        format_checker=jsonschema.FormatChecker(),
    )
    for spec in registry.entries:
        cohort_path = PROJECT / spec.cohort_path
        cohort = json.loads(cohort_path.read_text())
        contract = json.loads(
            (cohort_path.parent / cohort["acceptanceContract"]).read_text()
        )
        validator.validate(cohort)
        _validate_cohort(cohort)
        _validate_contract(contract, cohort)
        assert [item["year"] for item in cohort["workbooks"]] == list(
            spec.expected_years
        )
        assert [item["sheet"] for item in cohort["workbooks"]]
        assert all(
            item["replayResponse"]["acceptanceAuthority"] is False
            for item in cohort["workbooks"]
        )


def _write_registry(tmp_path: Path, value: dict[str, object]) -> None:
    registry_path = tmp_path / "fixtures/product-prototype/large-batch-assets-v1.json"
    registry_path.parent.mkdir(parents=True)
    registry_path.write_text(json.dumps(value))


def _registry_value() -> dict[str, object]:
    return json.loads(
        (PROJECT / "fixtures/product-prototype/large-batch-assets-v1.json").read_text()
    )


def test_registry_pins_acceptance_policy_and_v2_replay_timestamp() -> None:
    registry = load_large_batch_registry(PROJECT)
    v1 = [
        spec
        for spec in registry.entries
        if spec.acceptance_policy_version == "tidy.table-family-acceptance/v1"
    ]
    v2 = [
        spec
        for spec in registry.entries
        if spec.acceptance_policy_version == "tidy.table-family-acceptance/v2"
    ]
    assert len(v1) == 125
    assert len(v2) == 95
    assert all(spec.replay_recorded_at is None for spec in v1)
    assert Counter(spec.replay_recorded_at for spec in v2) == {
        "2026-08-15T09:00:00+00:00": 9,
        "2026-08-21T09:00:00+00:00": 9,
        "2026-08-22T09:00:00+00:00": 10,
        "2026-08-23T09:00:00+00:00": 18,
        "2026-08-24T09:00:00+00:00": 18,
        "2026-08-25T09:00:00+00:00": 31,
    }


@pytest.mark.parametrize("mutation", ["missing-policy", "v2-missing-timestamp"])
def test_registry_rejects_unpinned_policy_or_v2_timestamp(
    tmp_path: Path, mutation: str
) -> None:
    value = _registry_value()
    entry = value["entries"][-1]
    if mutation == "missing-policy":
        entry.pop("acceptancePolicyVersion")
    else:
        assert entry["acceptancePolicyVersion"].endswith("/v2")
        entry.pop("replayRecordedAt")
    _write_registry(tmp_path, value)
    with pytest.raises(LargeBatchError, match="entry shape is invalid"):
        load_large_batch_registry(tmp_path)


def test_registry_pins_exact_exclusions_for_every_year() -> None:
    registry = load_large_batch_registry(PROJECT)
    for spec in registry.entries:
        assert set(spec.expected_excluded_observation_counts_by_year) == set(
            spec.expected_years
        )
        expected = (
            {2021: 0, 2022: 36, 2023: 37, 2024: 37, 2025: 38}
            if spec.family_id == "national-selected-characteristics-by-offence-charge"
            else {year: 0 for year in spec.expected_years}
        )
        assert spec.expected_excluded_observation_counts_by_year == expected
        assert sum(expected.values()) == spec.expected_excluded_observation_count


def test_registry_rejects_output_path_escape(tmp_path: Path) -> None:
    value = _registry_value()
    value["entries"][0]["outputDirectory"] = "../escape"
    _write_registry(tmp_path, value)
    with pytest.raises(LargeBatchError, match="entry is invalid"):
        load_large_batch_registry(tmp_path)


def test_registry_rejects_uncustodied_family_separator(tmp_path: Path) -> None:
    value = _registry_value()
    value["entries"][0]["familyId"] = "invalid---family"
    _write_registry(tmp_path, value)
    with pytest.raises(LargeBatchError, match="entry is invalid"):
        load_large_batch_registry(tmp_path)


@pytest.mark.parametrize("invalid_count", [-1, "0", True])
def test_registry_rejects_invalid_per_year_exclusion_types(
    tmp_path: Path, invalid_count: object
) -> None:
    value = _registry_value()
    value["entries"][0]["expectedExcludedObservationCountsByYear"]["2021"] = (
        invalid_count
    )
    _write_registry(tmp_path, value)
    with pytest.raises(LargeBatchError, match="entry is invalid"):
        load_large_batch_registry(tmp_path)


@pytest.mark.parametrize("mutation", ["missing", "extra"])
def test_registry_rejects_missing_or_extra_exclusion_year(
    tmp_path: Path, mutation: str
) -> None:
    value = _registry_value()
    counts = value["entries"][0]["expectedExcludedObservationCountsByYear"]
    if mutation == "missing":
        counts.pop("2021")
    else:
        counts["2099"] = 0
    _write_registry(tmp_path, value)
    with pytest.raises(LargeBatchError, match="entry is invalid"):
        load_large_batch_registry(tmp_path)


def _normalization_manifest() -> dict[str, object]:
    return json.loads(
        (
            PROJECT / "fixtures/product-prototype/batch-workbook-normalization-v1.json"
        ).read_text()
    )


def _bind_normalization_manifest(manifest: dict[str, object]) -> None:
    semantic = {
        key: value for key, value in manifest.items() if key != "manifestDigest"
    }
    manifest["manifestDigest"] = domain_digest(
        large_batch_module.NORMALIZATION_SCHEMA, semantic
    )


def test_normalization_manifest_rejects_missing_entry_identity(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    manifest = _normalization_manifest()
    manifest["entries"][0].pop("normalization")
    _bind_normalization_manifest(manifest)
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if label == "normalization manifest":
            return manifest
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="manifest entry is invalid"):
        verify_batch_normalization(PROJECT, registry)


def test_normalization_manifest_rejects_relabelled_output(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    manifest = _normalization_manifest()
    manifest["entries"][0]["normalization"] = (
        "trim-pathological-full-width-formatting-merge-v1"
    )
    _bind_normalization_manifest(manifest)
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if label == "normalization manifest":
            return manifest
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="does not match normalization manifest"):
        verify_batch_normalization(PROJECT, registry)


def test_normalization_manifest_rejects_wrong_cohort_identity(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = registry.entries[0]
    cohort_path = (PROJECT / spec.cohort_path).resolve()
    cohort = json.loads(cohort_path.read_text())
    normalized = next(
        workbook for workbook in cohort["workbooks"] if "normalization" in workbook
    )
    normalized["normalization"] = "wrong-normalization-v1"
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if label == "large-batch cohort" and path == cohort_path:
            return cohort
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="does not match normalization manifest"):
        verify_batch_normalization(PROJECT, registry)


def _bind_run_mutation(run: dict[str, object], manifest: dict[str, object]) -> None:
    semantic = dict(run)
    semantic.pop("runDigest")
    run["runDigest"] = domain_digest(RUN_SCHEMA, semantic)
    manifest["runDigest"] = run["runDigest"]


def _v1_spec():
    registry = load_large_batch_registry(PROJECT)
    return next(
        item
        for item in registry.entries
        if item.acceptance_policy_version.endswith("/v1")
    )


def _v2_spec():
    registry = load_large_batch_registry(PROJECT)
    return next(
        item
        for item in registry.entries
        if item.family_id.endswith("all-courts-north-9e1eae4d24")
    )


def _copy_evidence_closure(tmp_path: Path, spec) -> tuple[Path, Path, Path]:
    cohort_relative = Path(spec.cohort_path)
    cohort = json.loads((PROJECT / cohort_relative).read_text())
    paths = [
        cohort_relative,
        cohort_relative.parent / cohort["acceptanceContract"],
    ]
    for workbook in cohort["workbooks"]:
        paths.extend(
            [
                cohort_relative.parent / workbook["path"],
                cohort_relative.parent / workbook["replayResponse"]["path"],
            ]
        )
    for relative in paths:
        destination = tmp_path / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(PROJECT / relative, destination)
    evidence_root = (tmp_path / spec.evidence_manifest_path).parent
    shutil.copytree((PROJECT / spec.evidence_manifest_path).parent, evidence_root)
    return (
        tmp_path / cohort_relative,
        tmp_path / paths[1],
        evidence_root,
    )


def _rebind_copied_evidence(
    cohort_path: Path,
    contract_path: Path,
    evidence_root: Path,
    rows: list[dict[str, object]],
    run: dict[str, object],
    manifest: dict[str, object],
) -> None:
    contract = json.loads(contract_path.read_text())
    canonical_bytes = canonical_json_bytes(rows) + b"\n"
    csv_bytes = product_prototype_module._canonical_csv(rows, contract)
    (evidence_root / "canonical-observations.json").write_bytes(canonical_bytes)
    (evidence_root / "canonical-observations.csv").write_bytes(csv_bytes)
    run["cohortDigest"] = sha256_digest(cohort_path.read_bytes())
    run["acceptanceContractDigest"] = sha256_digest(contract_path.read_bytes())
    run["canonicalJsonDigest"] = sha256_digest(canonical_bytes)
    run["canonicalCsvDigest"] = sha256_digest(csv_bytes)
    semantic_run = dict(run)
    semantic_run.pop("runDigest", None)
    run["runDigest"] = domain_digest(RUN_SCHEMA, semantic_run)
    run_bytes = canonical_json_bytes(run) + b"\n"
    (evidence_root / "run.json").write_bytes(run_bytes)
    manifest["cohortDigest"] = run["cohortDigest"]
    manifest["acceptanceContractDigest"] = run["acceptanceContractDigest"]
    manifest["runDigest"] = run["runDigest"]
    for declaration in manifest["files"]:
        data = (evidence_root / declaration["path"]).read_bytes()
        declaration["contentDigest"] = sha256_digest(data)
        declaration["byteLength"] = len(data)
    (evidence_root / "manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n"
    )


def _decision_id(
    workbook: dict[str, object],
    *,
    subject_id: str,
    policy_version: str,
    policy_digest: str,
    recorded_at: str,
    decision_identity_version: str | None = None,
) -> str:
    contract = (
        {"decisionIdentityVersion": decision_identity_version}
        if decision_identity_version is not None
        else {}
    )
    payload = _acceptance_decision_payload(
        contract=contract,
        acceptance_policy_version=policy_version,
        acceptance_policy_digest=policy_digest,
        year=workbook["year"],
        workbook_digest=workbook["workbookDigest"],
        sheet=workbook["sheet"],
        reference_date=workbook["referenceDate"],
        checks=workbook["checks"],
        issues=workbook["issues"],
    )
    return DecisionRecord.create(
        subject_id=subject_id,
        decision_type="prototype_auto_accepted",
        payload=payload,
        actor="tidy.product-prototype-policy/v1",
        recorded_at=recorded_at,
    ).decision_id


def test_evidence_rejects_redistributed_per_year_exclusions(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = next(
        item
        for item in registry.entries
        if item.family_id == "national-selected-characteristics-by-offence-charge"
    )
    manifest_path = (PROJECT / spec.evidence_manifest_path).resolve()
    evidence_root = manifest_path.parent
    manifest = json.loads(manifest_path.read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    run["workbooks"][0]["excludedObservationCount"] = 1
    run["workbooks"][1]["excludedObservationCount"] = 35
    _bind_run_mutation(run, manifest)
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if path == manifest_path:
            return manifest
        if path == evidence_root / "run.json":
            return run
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="Run evidence is invalid"):
        verify_large_batch_evidence(PROJECT, spec)


@pytest.mark.parametrize(
    "flag",
    ["historicalReplayIsAcceptanceAuthority", "trainingEligibility"],
)
def test_evidence_rejects_unsafe_run_authority_claims(
    monkeypatch: pytest.MonkeyPatch,
    flag: str,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = registry.entries[0]
    manifest_path = (PROJECT / spec.evidence_manifest_path).resolve()
    evidence_root = manifest_path.parent
    manifest = json.loads(manifest_path.read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    run[flag] = True
    _bind_run_mutation(run, manifest)
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if path == manifest_path:
            return manifest
        if path == evidence_root / "run.json":
            return run
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="Run evidence is invalid"):
        verify_large_batch_evidence(PROJECT, spec)


def test_evidence_rejects_fully_rebound_authoritative_replay_input(
    tmp_path: Path,
) -> None:
    spec = _v1_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    cohort = json.loads(cohort_path.read_text())
    cohort["workbooks"][0]["replayResponse"]["acceptanceAuthority"] = True
    cohort_path.write_text(json.dumps(cohort, indent=2, ensure_ascii=False) + "\n")
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )

    with pytest.raises(LargeBatchError, match="Evidence cohort is invalid"):
        verify_large_batch_evidence(tmp_path, spec)


def test_evidence_rejects_fully_rebound_training_eligible_contract(
    tmp_path: Path,
) -> None:
    spec = _v1_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    contract = json.loads(contract_path.read_text())
    contract["trainingEligibility"] = True
    contract_path.write_text(json.dumps(contract, indent=2, ensure_ascii=False) + "\n")
    policy_digest = sha256_digest(canonical_json_bytes(contract))
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    for row in rows:
        row["acceptance_policy_digest"] = policy_digest
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )

    with pytest.raises(
        LargeBatchError,
        match="Evidence acceptance contract is invalid",
    ):
        verify_large_batch_evidence(tmp_path, spec)


@pytest.mark.parametrize(
    "flag",
    ["historicalReplayIsAcceptanceAuthority", "trainingEligibility"],
)
def test_reproduction_rejects_unsafe_generated_run_authority_claims(
    tmp_path: Path,
    flag: str,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = registry.entries[0]
    evidence_root = (PROJECT / spec.evidence_manifest_path).parent
    output_root = tmp_path / "reproduction"
    output_root.mkdir()
    for filename in (
        "canonical-observations.csv",
        "canonical-observations.json",
        "collation-report.json",
        "exceptions.json",
    ):
        (output_root / filename).write_bytes((evidence_root / filename).read_bytes())
    run = json.loads((evidence_root / "run.json").read_text())
    run[flag] = True
    (output_root / "run.json").write_text(json.dumps(run))

    with pytest.raises(LargeBatchError, match="run authority claims are invalid"):
        verify_large_batch_reproduction(PROJECT, spec, output_root)


def _csv_fixture_rows() -> tuple[dict[str, object], list[dict[str, object]]]:
    spec = _v1_spec()
    cohort_path = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (cohort_path.parent / cohort["acceptanceContract"]).read_text()
    )
    evidence_root = (PROJECT / spec.evidence_manifest_path).parent
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())[:2]
    raw_field = next(field for field in rows[0] if field.startswith("raw_"))
    rows[0][raw_field] = str(rows[0][raw_field]) + "  "
    rows[0]["source_sheet"] = str(rows[0]["source_sheet"]) + " "
    rows[0]["value"] = None
    rows[1]["value"] = 12.5
    return contract, rows


def test_canonical_csv_projection_preserves_policy_semantics() -> None:
    contract, rows = _csv_fixture_rows()
    data = product_prototype_module._canonical_csv(rows, contract)
    large_batch_module._verify_canonical_csv_projection(data, rows, contract)
    parsed = list(csv.DictReader(io.StringIO(data.decode(), newline="")))
    raw_field = next(field for field in rows[0] if field.startswith("raw_"))
    assert parsed[0][raw_field] == str(rows[0][raw_field]).rstrip()
    assert parsed[0]["source_sheet"] == rows[0]["source_sheet"]
    assert parsed[0]["value"] == ""
    assert parsed[1]["value"] == "12.5"


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("field-order", "field order"),
        ("extra-field", "field order"),
        ("missing-row", "rows do not match"),
        ("row-order", "rows do not match"),
        ("null-drift", "rows do not match"),
        ("numeric-drift", "rows do not match"),
        ("source-sheet-rstrip", "rows do not match"),
        ("raw-whitespace-untrimmed", "rows do not match"),
    ],
)
def test_canonical_csv_projection_rejects_semantic_drift(
    mutation: str,
    message: str,
) -> None:
    contract, rows = _csv_fixture_rows()
    table = list(
        csv.reader(
            io.StringIO(
                product_prototype_module._canonical_csv(rows, contract).decode(),
                newline="",
            )
        )
    )
    header = table[0]
    if mutation == "field-order":
        header[0], header[1] = header[1], header[0]
    elif mutation == "extra-field":
        header.append("extra")
        for row in table[1:]:
            row.append("")
    elif mutation == "missing-row":
        table.pop()
    elif mutation == "row-order":
        table[1], table[2] = table[2], table[1]
    elif mutation == "null-drift":
        table[1][header.index("value")] = "0"
    elif mutation == "numeric-drift":
        table[2][header.index("value")] = "12.50"
    elif mutation == "source-sheet-rstrip":
        table[1][header.index("source_sheet")] = str(rows[0]["source_sheet"]).rstrip()
    else:
        raw_field = next(field for field in rows[0] if field.startswith("raw_"))
        table[1][header.index(raw_field)] = str(rows[0][raw_field])
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerows(table)
    with pytest.raises(LargeBatchError, match=message):
        large_batch_module._verify_canonical_csv_projection(
            output.getvalue().encode(), rows, contract
        )


@pytest.mark.parametrize(
    "filename",
    [
        "README.md",
        "canonical-observations.csv",
        "canonical-observations.json",
        "collation-report.json",
        "exceptions.json",
        "manifest.json",
        "run.json",
    ],
)
def test_complete_reproduction_requires_all_seven_exact_files(
    tmp_path: Path,
    filename: str,
) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    generated = tmp_path / "generated"
    shutil.copytree(source, generated)
    verify_large_batch_complete_reproduction(PROJECT, spec, generated)
    (generated / filename).write_bytes((generated / filename).read_bytes() + b" ")
    with pytest.raises(LargeBatchError, match="differs from checked evidence"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_complete_reproduction_rejects_missing_and_extra_files(tmp_path: Path) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    generated = tmp_path / "generated"
    shutil.copytree(source, generated)
    (generated / "README.md").unlink()
    with pytest.raises(LargeBatchError, match="missing or extra files"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)
    shutil.copy2(source / "README.md", generated / "README.md")
    (generated / "extra.txt").write_text("extra")
    with pytest.raises(LargeBatchError, match="missing or extra files"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_complete_reproduction_rejects_root_symlink_to_checked_evidence(
    tmp_path: Path,
) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    generated = tmp_path / "generated"
    generated.symlink_to(source, target_is_directory=True)
    with pytest.raises(LargeBatchError, match="root must be a real directory"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_complete_reproduction_rejects_root_symlink_to_separate_copy(
    tmp_path: Path,
) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    copied = tmp_path / "copied"
    shutil.copytree(source, copied)
    generated = tmp_path / "generated"
    generated.symlink_to(copied, target_is_directory=True)
    with pytest.raises(LargeBatchError, match="root must be a real directory"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_complete_reproduction_rejects_expected_file_symlink(tmp_path: Path) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    generated = tmp_path / "generated"
    shutil.copytree(source, generated)
    readme = generated / "README.md"
    readme.unlink()
    readme.symlink_to(source / "README.md")
    with pytest.raises(LargeBatchError, match="reproduction is missing: README.md"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_complete_reproduction_rejects_non_directory_root(tmp_path: Path) -> None:
    spec = _v1_spec()
    generated = tmp_path / "generated"
    generated.write_text("not a directory")
    with pytest.raises(LargeBatchError, match="root must be a real directory"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


@pytest.mark.parametrize("relationship", ["identity", "parent"])
def test_complete_reproduction_rejects_checked_evidence_overlap(
    relationship: str,
) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    generated = source if relationship == "identity" else source.parent
    with pytest.raises(LargeBatchError, match="root overlaps checked evidence"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_complete_reproduction_rejects_generated_child_of_checked_evidence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    spec = _v1_spec()
    source = (PROJECT / spec.evidence_manifest_path).parent
    manifest = verify_large_batch_evidence(PROJECT, spec)
    evidence = tmp_path / "evidence"
    shutil.copytree(source, evidence)
    generated = evidence / "generated"
    shutil.copytree(source, generated)
    monkeypatch.setattr(
        large_batch_module,
        "verify_large_batch_evidence",
        lambda _project_root, _spec: manifest,
    )
    monkeypatch.setattr(
        large_batch_module,
        "_safe_path",
        lambda _root, _relative, _label: evidence / "manifest.json",
    )
    with pytest.raises(LargeBatchError, match="root overlaps checked evidence"):
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)


def test_evidence_rejects_warning_count_drift(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = next(
        item
        for item in registry.entries
        if item.family_id.endswith("new-south-wales-and-99870bae7c")
    )
    manifest_path = (PROJECT / spec.evidence_manifest_path).resolve()
    manifest = json.loads(manifest_path.read_text())
    manifest["warningCountsByYear"]["2024"] += 1
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if path == manifest_path:
            return manifest
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="warning counts do not match"):
        verify_large_batch_evidence(PROJECT, spec)


def test_evidence_rejects_self_consistently_rebound_acceptance_mismatch(
    tmp_path: Path,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = next(
        item
        for item in registry.entries
        if item.family_id.endswith("western-australia-and-be8fa3884d")
    )
    cohort_source = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_source.read_text())
    paths = [
        Path(spec.cohort_path),
        Path(spec.cohort_path).parent / cohort["acceptanceContract"],
    ]
    for workbook in cohort["workbooks"]:
        paths.extend(
            [
                Path(spec.cohort_path).parent / workbook["path"],
                Path(spec.cohort_path).parent / workbook["replayResponse"]["path"],
            ]
        )
    for relative in paths:
        destination = tmp_path / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(PROJECT / relative, destination)

    evidence_source = (PROJECT / spec.evidence_manifest_path).parent
    evidence_root = (tmp_path / spec.evidence_manifest_path).parent
    shutil.copytree(evidence_source, evidence_root)

    canonical_path = evidence_root / "canonical-observations.json"
    rows = json.loads(canonical_path.read_text())
    assert (
        rows[0]["publication_vintage_date"] == cohort["workbooks"][0]["referenceDate"]
    )
    rows[0]["acceptance_decision_digest"] = "sha256:" + "0" * 64
    canonical_bytes = (
        json.dumps(rows, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
        + "\n"
    ).encode()
    canonical_path.write_bytes(canonical_bytes)
    contract = json.loads((tmp_path / paths[1]).read_text())
    csv_bytes = product_prototype_module._canonical_csv(rows, contract)
    (evidence_root / "canonical-observations.csv").write_bytes(csv_bytes)

    run_path = evidence_root / "run.json"
    run = json.loads(run_path.read_text())
    run["canonicalCsvDigest"] = sha256_digest(csv_bytes)
    run["canonicalJsonDigest"] = "sha256:" + hashlib.sha256(canonical_bytes).hexdigest()
    semantic_run = dict(run)
    semantic_run.pop("runDigest")
    run["runDigest"] = domain_digest(RUN_SCHEMA, semantic_run)
    run_bytes = (
        json.dumps(run, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
        + "\n"
    ).encode()
    run_path.write_bytes(run_bytes)

    manifest_path = evidence_root / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["runDigest"] = run["runDigest"]
    rebound = {
        "canonical-observations.csv": csv_bytes,
        "canonical-observations.json": canonical_bytes,
        "run.json": run_bytes,
    }
    for declaration in manifest["files"]:
        if declaration["path"] in rebound:
            content = rebound[declaration["path"]]
            declaration["contentDigest"] = (
                "sha256:" + hashlib.sha256(content).hexdigest()
            )
            declaration["byteLength"] = len(content)
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n")

    with pytest.raises(
        LargeBatchError,
        match="Canonical acceptance decision does not match run evidence",
    ):
        verify_large_batch_evidence(tmp_path, spec)


@pytest.mark.parametrize(
    ("schema_version", "identity_version"),
    [
        ("tidy.table-family-acceptance/v1", V2_REFERENCE_DATE_DECISION_IDENTITY),
        ("tidy.table-family-acceptance/v2", "v2-reference-date-v2"),
        ("tidy.table-family-acceptance/v2", ""),
        ("tidy.table-family-acceptance/v2", None),
    ],
)
def test_decision_identity_version_is_exact_and_v2_only(
    schema_version: str,
    identity_version: str | None,
) -> None:
    spec = _v2_spec()
    cohort_path = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (cohort_path.parent / cohort["acceptanceContract"]).read_text()
    )
    contract["schemaVersion"] = schema_version
    contract["decisionIdentityVersion"] = identity_version
    if schema_version.endswith("/v1"):
        contract.pop("expectedRecipeDigestsByYear")
    with pytest.raises(ProductPrototypeError, match="contract is incompatible"):
        _validate_contract(contract, cohort)


def test_pending_v2_contracts_can_opt_into_reference_date_identity() -> None:
    fixture_root = PROJECT / "fixtures/product-prototype"
    membership = json.loads(
        (fixture_root / "criminal-courts-release-family-membership-v1.json").read_text()
    )
    family_ids = sorted(
        family["familyId"]
        for family in membership["families"]
        if family["members"]
        and all(
            member["cubeId"].startswith("family-and-domestic-violence-offences")
            for member in family["members"]
        )
    )
    assert len(family_ids) == 31
    for family_id in family_ids:
        cohort_path = fixture_root / f"{family_id}.json"
        contract_path = fixture_root / "acceptance" / f"{family_id}-v1.json"
        cohort = json.loads(cohort_path.read_text())
        contract = json.loads(contract_path.read_text())
        assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
        contract["decisionIdentityVersion"] = V2_REFERENCE_DATE_DECISION_IDENTITY
        _validate_contract(contract, cohort)


def test_v2_reference_date_opt_in_changes_decision_identity_only_with_date() -> None:
    spec = _v2_spec()
    cohort_path = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_path.read_text())
    contract_path = cohort_path.parent / cohort["acceptanceContract"]
    contract = json.loads(contract_path.read_text())
    workbook = json.loads(
        ((PROJECT / spec.evidence_manifest_path).parent / "run.json").read_text()
    )["workbooks"][0]
    policy_digest = sha256_digest(contract_path.read_bytes())
    pinned_recipe = contract["expectedRecipeDigestsByYear"][str(workbook["year"])]
    legacy_id = _decision_id(
        workbook,
        subject_id=pinned_recipe,
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=spec.replay_recorded_at,
    )
    assert legacy_id == workbook["decisionId"]
    bound_id = _decision_id(
        workbook,
        subject_id=pinned_recipe,
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=spec.replay_recorded_at,
        decision_identity_version=V2_REFERENCE_DATE_DECISION_IDENTITY,
    )
    changed = copy.deepcopy(workbook)
    changed["referenceDate"] = "2099-12-31"
    changed_id = _decision_id(
        changed,
        subject_id=pinned_recipe,
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=spec.replay_recorded_at,
        decision_identity_version=V2_REFERENCE_DATE_DECISION_IDENTITY,
    )
    assert bound_id != legacy_id
    assert changed_id != bound_id


def test_replay_map_digest_is_custodied_but_not_decision_authority(
    tmp_path: Path,
) -> None:
    spec = _v2_spec()
    cohort_path, _contract_path, _evidence_root = _copy_evidence_closure(tmp_path, spec)
    cohort = json.loads(cohort_path.read_text())
    workbook = cohort["workbooks"][0]
    response_path = cohort_path.parent / workbook["replayResponse"]["path"]
    response_path.write_bytes(response_path.read_bytes() + b"\n")
    with pytest.raises(LargeBatchError, match="Cohort input digest mismatch"):
        verify_large_batch_evidence(tmp_path, spec)

    # Map bytes are checked above as cohort/reproduction custody. They are omitted
    # from the acceptance payload deliberately; the accepted recipe is its subject.
    run_workbook = json.loads(
        ((PROJECT / spec.evidence_manifest_path).parent / "run.json").read_text()
    )["workbooks"][0]
    contract = {}
    first = _acceptance_decision_payload(
        contract=contract,
        acceptance_policy_version=spec.acceptance_policy_version,
        acceptance_policy_digest="sha256:" + "1" * 64,
        year=run_workbook["year"],
        workbook_digest=run_workbook["workbookDigest"],
        sheet=run_workbook["sheet"],
        reference_date=run_workbook["referenceDate"],
        checks=run_workbook["checks"],
        issues=run_workbook["issues"],
    )
    workbook["replayResponse"]["contentDigest"] = "sha256:" + "0" * 64
    second = _acceptance_decision_payload(
        contract=contract,
        acceptance_policy_version=spec.acceptance_policy_version,
        acceptance_policy_digest="sha256:" + "1" * 64,
        year=run_workbook["year"],
        workbook_digest=run_workbook["workbookDigest"],
        sheet=run_workbook["sheet"],
        reference_date=run_workbook["referenceDate"],
        checks=run_workbook["checks"],
        issues=run_workbook["issues"],
    )
    assert first == second
    assert "replayResponseDigest" not in first
    assert workbook["replayResponse"]["contentDigest"] not in first.values()


def test_v2_reference_date_opt_in_is_enforced_by_evidence_verifier(
    tmp_path: Path,
) -> None:
    spec = _v2_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    contract = json.loads(contract_path.read_text())
    contract["decisionIdentityVersion"] = V2_REFERENCE_DATE_DECISION_IDENTITY
    contract_path.write_text(json.dumps(contract, indent=2, ensure_ascii=False) + "\n")
    policy_digest = sha256_digest(contract_path.read_bytes())
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    for workbook in run["workbooks"]:
        pinned_recipe = contract["expectedRecipeDigestsByYear"][str(workbook["year"])]
        decision_id = _decision_id(
            workbook,
            subject_id=pinned_recipe,
            policy_version=contract["schemaVersion"],
            policy_digest=policy_digest,
            recorded_at=manifest["recordedAt"],
            decision_identity_version=V2_REFERENCE_DATE_DECISION_IDENTITY,
        )
        workbook["decisionId"] = decision_id
        for row in rows:
            if (
                row["source_workbook_digest"] == workbook["workbookDigest"]
                and row["source_sheet"] == workbook["sheet"]
            ):
                row["acceptance_policy_digest"] = policy_digest
                row["acceptance_decision_digest"] = decision_id
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )
    verify_large_batch_evidence(tmp_path, spec)

    cohort = json.loads(cohort_path.read_text())
    old_date = cohort["workbooks"][0]["referenceDate"]
    new_date = "2099-12-31"
    cohort["workbooks"][0]["referenceDate"] = new_date
    cohort_path.write_text(json.dumps(cohort, indent=2, ensure_ascii=False) + "\n")
    run["workbooks"][0]["referenceDate"] = new_date
    for row in rows:
        if row.get("publication_vintage_date") == old_date:
            row["publication_vintage_date"] = new_date
        elif row["reference_date"] == old_date:
            row["reference_date"] = new_date
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )
    with pytest.raises(
        LargeBatchError,
        match="V2 acceptance decision does not bind acceptance policy",
    ):
        verify_large_batch_evidence(tmp_path, spec)


def test_v2_contract_requires_exact_recipe_digest_year_map() -> None:
    spec = _v2_spec()
    cohort_path = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (cohort_path.parent / cohort["acceptanceContract"]).read_text()
    )
    contract.pop("expectedRecipeDigestsByYear")
    with pytest.raises(ProductPrototypeError, match="recipe digest pins are invalid"):
        _validate_contract(contract, cohort)


def test_v1_evidence_retains_canonical_contract_digest_compatibility() -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = registry.entries[0]
    cohort_path = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_path.read_text())
    contract_path = cohort_path.parent / cohort["acceptanceContract"]
    contract = json.loads(contract_path.read_text())
    assert contract["schemaVersion"] == "tidy.table-family-acceptance/v1"
    expected_policy_digest = sha256_digest(canonical_json_bytes(contract))
    manifest = verify_large_batch_evidence(PROJECT, spec)
    assert expected_policy_digest != manifest["acceptanceContractDigest"]
    rows = json.loads(
        (
            (PROJECT / spec.evidence_manifest_path).parent
            / "canonical-observations.json"
        ).read_text()
    )
    assert {row["acceptance_policy_version"] for row in rows} == {
        contract["schemaVersion"]
    }
    assert {row["acceptance_policy_digest"] for row in rows} == {expected_policy_digest}


def test_v2_evidence_rejects_fully_rebound_policy_downgrade(
    tmp_path: Path,
) -> None:
    spec = _v2_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    contract = json.loads(contract_path.read_text())
    contract["schemaVersion"] = "tidy.table-family-acceptance/v1"
    contract.pop("expectedRecipeDigestsByYear")
    contract_path.write_text(json.dumps(contract, indent=2, ensure_ascii=False) + "\n")
    policy_digest = sha256_digest(canonical_json_bytes(contract))
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    workbook = run["workbooks"][0]
    decision_id = _decision_id(
        workbook,
        subject_id=rows[0]["recipe_digest"],
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=manifest["recordedAt"],
    )
    workbook["decisionId"] = decision_id
    for row in rows:
        row["acceptance_policy_version"] = contract["schemaVersion"]
        row["acceptance_policy_digest"] = policy_digest
        row["acceptance_decision_digest"] = decision_id
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )

    with pytest.raises(LargeBatchError, match="schema does not match registry pin"):
        verify_large_batch_evidence(tmp_path, spec)


def test_v2_evidence_rejects_fully_rebound_arbitrary_recipe_digest(
    tmp_path: Path,
) -> None:
    spec = _v2_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    contract = json.loads(contract_path.read_text())
    policy_digest = sha256_digest(contract_path.read_bytes())
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    arbitrary_digest = "sha256:" + "0" * 64
    workbook = run["workbooks"][0]
    decision_id = _decision_id(
        workbook,
        subject_id=arbitrary_digest,
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=manifest["recordedAt"],
    )
    workbook["decisionId"] = decision_id
    for row in rows:
        row["recipe_digest"] = arbitrary_digest
        row["acceptance_decision_digest"] = decision_id
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )

    with pytest.raises(
        LargeBatchError, match="recipe identity does not match contract"
    ):
        verify_large_batch_evidence(tmp_path, spec)


def test_v2_evidence_rejects_fully_rebound_substituted_workbook_identity(
    tmp_path: Path,
) -> None:
    spec = _v2_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    contract = json.loads(contract_path.read_text())
    policy_digest = sha256_digest(contract_path.read_bytes())
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    substituted_digest = "sha256:" + "1" * 64
    workbook = run["workbooks"][0]
    workbook["workbookDigest"] = substituted_digest
    pinned_recipe = contract["expectedRecipeDigestsByYear"][str(workbook["year"])]
    decision_id = _decision_id(
        workbook,
        subject_id=pinned_recipe,
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=manifest["recordedAt"],
    )
    workbook["decisionId"] = decision_id
    for row in rows:
        row["source_workbook_digest"] = substituted_digest
        row["acceptance_decision_digest"] = decision_id
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )

    with pytest.raises(LargeBatchError, match="Run evidence is invalid"):
        verify_large_batch_evidence(tmp_path, spec)


@pytest.mark.parametrize("mutation", ["empty", "missing", "extra", "truthy"])
def test_v2_evidence_rejects_nonexact_checks(
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
) -> None:
    spec = _v2_spec()
    manifest_path = (PROJECT / spec.evidence_manifest_path).resolve()
    evidence_root = manifest_path.parent
    manifest = json.loads(manifest_path.read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    checks = run["workbooks"][0]["checks"]
    if mutation == "empty":
        run["workbooks"][0]["checks"] = {}
    elif mutation == "missing":
        checks.pop("coverage")
    elif mutation == "extra":
        checks["unrootedExtraCheck"] = True
    else:
        checks["coverage"] = 1
    _bind_run_mutation(run, manifest)
    original_load = large_batch_module._load_object

    def load(path: Path, label: str) -> dict[str, object]:
        if path == manifest_path:
            return manifest
        if path == evidence_root / "run.json":
            return run
        return original_load(path, label)

    monkeypatch.setattr(large_batch_module, "_load_object", load)
    with pytest.raises(LargeBatchError, match="Run evidence is invalid"):
        verify_large_batch_evidence(PROJECT, spec)


def test_v2_evidence_rejects_fully_rebound_manifest_timestamp(
    tmp_path: Path,
) -> None:
    spec = _v2_spec()
    cohort_path, contract_path, evidence_root = _copy_evidence_closure(tmp_path, spec)
    contract = json.loads(contract_path.read_text())
    policy_digest = sha256_digest(contract_path.read_bytes())
    rows = json.loads((evidence_root / "canonical-observations.json").read_text())
    run = json.loads((evidence_root / "run.json").read_text())
    manifest = json.loads((evidence_root / "manifest.json").read_text())
    manifest["recordedAt"] = "2026-08-16T09:00:00+00:00"
    workbook = run["workbooks"][0]
    pinned_recipe = contract["expectedRecipeDigestsByYear"][str(workbook["year"])]
    decision_id = _decision_id(
        workbook,
        subject_id=pinned_recipe,
        policy_version=contract["schemaVersion"],
        policy_digest=policy_digest,
        recorded_at=manifest["recordedAt"],
    )
    workbook["decisionId"] = decision_id
    for row in rows:
        row["acceptance_decision_digest"] = decision_id
    _rebind_copied_evidence(
        cohort_path, contract_path, evidence_root, rows, run, manifest
    )

    with pytest.raises(LargeBatchError, match="timestamp does not match registry pin"):
        verify_large_batch_evidence(tmp_path, spec)


def test_v2_evidence_rejects_stale_decisions_after_exact_policy_rebind(
    tmp_path: Path,
) -> None:
    registry = load_large_batch_registry(PROJECT)
    spec = next(
        item
        for item in registry.entries
        if item.family_id.endswith("all-courts-north-9e1eae4d24")
    )
    cohort_source = PROJECT / spec.cohort_path
    cohort = json.loads(cohort_source.read_text())
    assert len(cohort["workbooks"]) == 1
    paths = [
        Path(spec.cohort_path),
        Path(spec.cohort_path).parent / cohort["acceptanceContract"],
    ]
    workbook = cohort["workbooks"][0]
    paths.extend(
        [
            Path(spec.cohort_path).parent / workbook["path"],
            Path(spec.cohort_path).parent / workbook["replayResponse"]["path"],
        ]
    )
    for relative in paths:
        destination = tmp_path / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(PROJECT / relative, destination)

    evidence_source = (PROJECT / spec.evidence_manifest_path).parent
    evidence_root = (tmp_path / spec.evidence_manifest_path).parent
    shutil.copytree(evidence_source, evidence_root)

    contract_path = tmp_path / paths[1]
    contract = json.loads(contract_path.read_text())
    assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
    aliases = next(iter(contract["aliases"].values()))
    aliases["Unused syntactically valid policy alias"] = next(iter(aliases.values()))
    contract_bytes = (
        json.dumps(contract, indent=2, ensure_ascii=False) + "\n"
    ).encode()
    contract_path.write_bytes(contract_bytes)
    policy_digest = sha256_digest(contract_bytes)

    canonical_path = evidence_root / "canonical-observations.json"
    rows = json.loads(canonical_path.read_text())
    stale_decisions = {row["acceptance_decision_digest"] for row in rows}
    for row in rows:
        row["acceptance_policy_digest"] = policy_digest
    canonical_bytes = canonical_json_bytes(rows) + b"\n"
    canonical_path.write_bytes(canonical_bytes)
    csv_bytes = product_prototype_module._canonical_csv(rows, contract)
    (evidence_root / "canonical-observations.csv").write_bytes(csv_bytes)

    run_path = evidence_root / "run.json"
    run = json.loads(run_path.read_text())
    assert {item["decisionId"] for item in run["workbooks"]} == stale_decisions
    run["acceptanceContractDigest"] = policy_digest
    run["canonicalCsvDigest"] = sha256_digest(csv_bytes)
    run["canonicalJsonDigest"] = sha256_digest(canonical_bytes)
    semantic_run = dict(run)
    semantic_run.pop("runDigest")
    run["runDigest"] = domain_digest(RUN_SCHEMA, semantic_run)
    run_bytes = canonical_json_bytes(run) + b"\n"
    run_path.write_bytes(run_bytes)

    manifest_path = evidence_root / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["acceptanceContractDigest"] = policy_digest
    manifest["runDigest"] = run["runDigest"]
    rebound = {
        "canonical-observations.csv": csv_bytes,
        "canonical-observations.json": canonical_bytes,
        "run.json": run_bytes,
    }
    for declaration in manifest["files"]:
        content = rebound.get(declaration["path"])
        if content is not None:
            declaration["contentDigest"] = sha256_digest(content)
            declaration["byteLength"] = len(content)
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n")

    with pytest.raises(
        LargeBatchError,
        match="V2 acceptance decision does not bind acceptance policy",
    ):
        verify_large_batch_evidence(tmp_path, spec)


def test_multi_condition_measure_selection_rejects_overlap() -> None:
    cohort_path = (
        PROJECT
        / "fixtures/product-prototype/prisoners-selected-characteristics-2021-2025.json"
    )
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (
            PROJECT
            / "fixtures"
            / "product-prototype"
            / "acceptance"
            / "prisoners-selected-characteristics-v1.json"
        ).read_text()
    )
    _validate_contract(contract, cohort)
    contract["measures"][2]["selection"]["conditions"] = {
        "statistic_basis": ["NUMBER"],
        "characteristic_group": ["SEX"],
    }
    with pytest.raises(ProductPrototypeError, match="selection overlaps"):
        _validate_contract(contract, cohort)


def test_empty_total_equations_require_explicit_not_applicable_policy() -> None:
    cohort_path = (
        PROJECT
        / "fixtures/product-prototype/prisoners-selected-characteristics-2021-2025.json"
    )
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (
            PROJECT
            / "fixtures"
            / "product-prototype"
            / "acceptance"
            / "prisoners-selected-characteristics-v1.json"
        ).read_text()
    )
    _validate_contract(contract, cohort)
    contract.pop("totalValidation")
    with pytest.raises(ProductPrototypeError, match="total equations"):
        _validate_contract(contract, cohort)


def test_reviewed_combination_digest_rejects_marginal_preserving_swap() -> None:
    cohort_path = (
        PROJECT / "fixtures/product-prototype/prisoners-table-34-2021-2025.json"
    )
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (
            PROJECT / "fixtures/product-prototype/acceptance/prisoners-table-34-v1.json"
        ).read_text()
    )
    rows = json.loads(
        (
            PROJECT
            / "fixtures"
            / "product-prototype"
            / "table-34-five-year-evidence"
            / "canonical-observations.json"
        ).read_text()
    )
    digest_2025 = next(
        item["contentDigest"] for item in cohort["workbooks"] if item["year"] == 2025
    )
    year_rows = [
        dict(row) for row in rows if row["source_workbook_digest"] == digest_2025
    ]
    assert _validate_expected_coverage(year_rows, contract, {"year": 2025}) == []
    candidates = [row for row in year_rows if row["measure_id"] == "prisoner-count"]
    left = candidates[0]
    right = next(
        row
        for row in candidates[1:]
        if row["jurisdiction_id"] != left["jurisdiction_id"]
        and row["prison_location_id"] != left["prison_location_id"]
    )
    left["prison_location_id"], right["prison_location_id"] = (
        right["prison_location_id"],
        left["prison_location_id"],
    )
    issues = _validate_expected_coverage(year_rows, contract, {"year": 2025})
    assert {item["code"] for item in issues} == {"EXPECTED_COMBINATION_SET_MISMATCH"}


def test_vintage_contract_requires_date_codes_and_two_time_axes() -> None:
    cohort_path = (
        PROJECT / "fixtures/product-prototype/prisoners-table-27-2021-2025.json"
    )
    cohort = json.loads(cohort_path.read_text())
    contract = json.loads(
        (
            PROJECT / "fixtures/product-prototype/acceptance/prisoners-table-27-v1.json"
        ).read_text()
    )
    _validate_contract(contract, cohort)
    assert contract["referenceDateDimension"] == "observation_period"
    assert contract["preservePublicationVintage"] is True
    assert contract["uniqueKey"][:2] == [
        "publication_vintage_date",
        "reference_date",
    ]
    rows = json.loads(
        (
            PROJECT
            / "fixtures"
            / "product-prototype"
            / "table-27-five-year-evidence"
            / "canonical-observations.json"
        ).read_text()
    )
    assert all(row["reference_date"] == row["observation_period_id"] for row in rows)
    assert any(row["publication_vintage_date"] != row["reference_date"] for row in rows)
    invalid = json.loads(json.dumps(contract))
    invalid["aliases"]["observation_period"]["2011"] = "YEAR_2011"
    with pytest.raises(ProductPrototypeError, match="Reference-date dimension"):
        _validate_contract(invalid, cohort)


def test_south_australia_criminal_courts_cluster_is_source_bound_and_closed() -> None:
    base = PROJECT / "fixtures/product-prototype"
    membership = json.loads(
        (base / "criminal-courts-release-family-membership-v1.json").read_text()
    )
    families = []
    for family in membership["families"]:
        members = [
            member
            for member in family["members"]
            if member["cubeId"] == "defendants-finalised-south-australia"
        ]
        if members:
            families.append((family["familyId"], members))

    expected_ids = {
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-south-b87fdba651",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-0c1d73ebf3",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--4c3a7e7ff1",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--53f02c72ed",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-72e91423b2",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-a5e4379445",
        "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-south-australia-1e29fe0da9",
        "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-south-australia-and-f3ba572c03",
        "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-south-australia-d12b5a0a38",
    }
    assert {family_id for family_id, _ in families} == expected_ids
    assert Counter(
        member["releaseId"] for _, members in families for member in members
    ) == {"2021-22": 5, "2022-23": 5, "2023-24": 6, "2024-25": 6}
    assert {
        (member["releaseId"], member["physicalSheetName"])
        for _, members in families
        for member in members
    } == {
        *(("2021-22", f"Table {number}") for number in range(31, 36)),
        *(("2022-23", f"Table {number}") for number in range(31, 36)),
        *(("2023-24", f"Table {number}") for number in range(34, 40)),
        *(("2024-25", f"Table {number}") for number in range(39, 45)),
    }
    assert all(
        member["registered"] is True for _, members in families for member in members
    )
    expected_source_digests = {
        "2021-22": "sha256:"
        "b50c0470380ab2efde73f486aedbe25429fbbeded6e49ea987ff41dcfbd971e8",
        "2022-23": "sha256:"
        "396d779d3214894f5113a2c70d3788271d78d0c60ca4c3cf71d6aa2050f6a766",
        "2023-24": "sha256:"
        "3fce8cd072186b2c4110664e5d68e7281d7906daa0fbf9b0054c1fd838d57ce0",
        "2024-25": "sha256:"
        "6b0b24c022a6d84e248845cd6af1dac3d229b3ccadcdb305dccd3c931985dad0",
    }
    for _, members in families:
        for member in members:
            source = base / member["sourcePath"]
            assert (
                "sha256:" + hashlib.sha256(source.read_bytes()).hexdigest()
                == (expected_source_digests[member["releaseId"]])
            )
            assert (
                member["sourceDigest"] == expected_source_digests[member["releaseId"]]
            )

    rows = []
    family_rows = {}
    warning_count = 0
    for family_id, members in families:
        cohort = json.loads((base / f"{family_id}.json").read_text())
        contract = json.loads(
            (base / "acceptance" / f"{family_id}-v1.json").read_text()
        )
        run = json.loads((base / f"{family_id}-evidence" / "run.json").read_text())
        manifest = json.loads(
            (base / f"{family_id}-evidence" / "manifest.json").read_text()
        )
        canonical_rows = json.loads(
            (base / f"{family_id}-evidence" / "canonical-observations.json").read_text()
        )
        family_rows[family_id] = canonical_rows
        rows.extend(canonical_rows)
        warning_count += sum(item["executionWarningCount"] for item in run["workbooks"])
        assert run["providerCalls"] == manifest["providerCalls"] == 0
        assert run["historicalReplayIsAcceptanceAuthority"] is False
        assert run["trainingEligibility"] is False
        assert contract["trainingEligibility"] is False
        assert contract["totalEquations"] == []
        assert (
            contract["expectedWarningCountsByYear"] == manifest["warningCountsByYear"]
        )
        assert contract["expectedWarningCountsByYear"] == {
            str(item["year"]): item["executionWarningCount"]
            for item in run["workbooks"]
        }
        assert all(
            rule["code"] == "AMBIGUOUS_HEADER"
            and rule["dimension"] in {"court_level", "observation_period"}
            and rule["requireCanonicalOutputEquivalence"] is True
            and set(rule)
            == {
                "code",
                "dimension",
                "requireCanonicalOutputEquivalence",
                "expectedHeaderSourcesByYear",
            }
            and all(
                sources
                for by_output in rule["expectedHeaderSourcesByYear"].values()
                for sources in by_output.values()
            )
            for rule in contract["allowedExecutionWarnings"]
        )
        assert [item["sheet"] for item in cohort["workbooks"]] == [
            member["physicalSheetName"] for member in members
        ]
        assert all(
            item["replayResponse"]["acceptanceAuthority"] is False
            and item["replayResponse"]["historicalModel"]
            == "human-authored/deterministic-map-v1"
            for item in cohort["workbooks"]
        )

    assert len(rows) == 17495
    assert Counter(row["measure_id"] for row in rows) == {
        "defendant-count": 16631,
        "mean-defendant-age": 216,
        "median-defendant-age": 216,
        "mean-case-duration": 216,
        "median-case-duration": 216,
    }
    assert Counter(row["value_status"] for row in rows) == {
        "observed": 17290,
        "not_applicable": 57,
        "not_available": 148,
    }
    assert (
        sum(row["value"] == 0 and row["value_status"] == "observed" for row in rows)
        == 1522
    )
    assert Counter(
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    ) == {("..", "not_applicable"): 57, ("na", "not_available"): 148}
    assert warning_count == 10734
    assert {row["jurisdiction_id"] for row in rows} == {"SA"}
    assert {row["classification_context_id"] for row in rows} == {
        "ANZSOC_2011",
        "ANZSOC_2023",
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
    }
    assert all(row["source_sheet"].startswith("Table ") for row in rows)
    assert all(row["raw_value"] == 0 for row in rows if row["value"] == 0)

    assert all(
        row.get("characteristic_group_id") != "GROUP_GUILTY_EX_PARTE" for row in rows
    )
    method_categories = {
        "CHAR_GUILTY_EX_PARTE",
        "CHAR_TRANSFER_TO_OTHER_COURT_LEVELS",
        "CHAR_WITHDRAWN_BY_PROSECUTION",
        "CHAR_TOTAL_FINALISED",
    }
    higher_courts_family = (
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-"
        "summary-outcomes-by-selected-principal-offence-higher-cour-72e91423b2"
    )
    historical_summary_family = (
        "criminal-courts-main-defendants-finalised-summary-characteristics-by-"
        "court-level-south-australia-d12b5a0a38"
    )
    mixed_summary_family = (
        "criminal-courts-main-defendants-finalised-summary-characteristics-by-"
        "court-level-south-australia-and-f3ba572c03"
    )
    affected_method_families = {
        higher_courts_family: (
            {"2022-06-30", "2023-06-30", "2024-06-30", "2025-06-30"},
            {"HIGHER_COURTS"},
        ),
        historical_summary_family: (
            {"2022-06-30", "2023-06-30", "2024-06-30"},
            {
                "ALL_COURTS",
                "CHILDRENS_COURTS",
                "HIGHER_COURTS",
                "MAGISTRATES_COURTS",
            },
        ),
        mixed_summary_family: (
            {"2025-06-30"},
            {
                "ALL_COURTS",
                "CHILDRENS_COURTS",
                "HIGHER_COURTS",
                "MAGISTRATES_COURTS",
            },
        ),
    }
    corrected_rows = []
    for family_id, (
        expected_vintages,
        expected_court_levels,
    ) in affected_method_families.items():
        selected = [
            row
            for row in family_rows[family_id]
            if row["characteristic_category_id"] in method_categories
        ]
        corrected_rows.extend(
            row for row in selected if row["court_level_id"] == "HIGHER_COURTS"
        )
        assert {row["publication_vintage_date"] for row in selected} == (
            expected_vintages
        )
        assert {row["court_level_id"] for row in selected} == expected_court_levels
        assert {
            (row["characteristic_group_id"], row["raw_characteristic_group"])
            for row in selected
        } == {("GROUP_METHOD_OF_FINALISATION", "Method of finalisation")}
    assert len(corrected_rows) == 387


def test_western_australia_criminal_courts_cluster_is_source_bound_and_closed() -> None:
    base = PROJECT / "fixtures/product-prototype"
    membership = json.loads(
        (base / "criminal-courts-release-family-membership-v1.json").read_text()
    )
    families = []
    for family in membership["families"]:
        members = [
            member
            for member in family["members"]
            if member["cubeId"] == "defendants-finalised-western-australia"
        ]
        if members:
            families.append((family["familyId"], members))

    expected_ids = {
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-all-courts-weste-dbc6bd4c00",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-all-principal-offence-magistrates-cour-38195c8963",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-all-courts--ff52555b4d",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-children-s--5674ad07c6",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-higher-cour-f953216c7c",
        "criminal-courts-main-defendants-finalised-and-with-a-guilty-outcome-summary-outcomes-by-selected-principal-offence-magistrates-b52ac75a8f",
        "criminal-courts-main-defendants-finalised-principal-offence-by-method-of-finalisation-western-australia-3c7004c375",
        "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-western-australia-and-be8fa3884d",
        "criminal-courts-main-defendants-finalised-summary-characteristics-by-court-level-western-australia-df3797a707",
    }
    assert {family_id for family_id, _ in families} == expected_ids
    assert Counter(
        member["releaseId"] for _, members in families for member in members
    ) == {"2021-22": 5, "2022-23": 5, "2023-24": 6, "2024-25": 6}
    assert {
        (member["releaseId"], member["physicalSheetName"])
        for _, members in families
        for member in members
    } == {
        *(("2021-22", f"Table {number}") for number in range(36, 41)),
        *(("2022-23", f"Table {number}") for number in range(36, 41)),
        *(("2023-24", f"Table {number}") for number in range(40, 46)),
        *(("2024-25", f"Table {number}") for number in range(45, 51)),
    }
    assert all(
        member["registered"] is True for _, members in families for member in members
    )

    rows = []
    warning_count = 0
    for family_id, members in families:
        cohort = json.loads((base / f"{family_id}.json").read_text())
        contract = json.loads(
            (base / "acceptance" / f"{family_id}-v1.json").read_text()
        )
        run = json.loads((base / f"{family_id}-evidence/run.json").read_text())
        manifest = json.loads(
            (base / f"{family_id}-evidence/manifest.json").read_text()
        )
        canonical_rows = json.loads(
            (base / f"{family_id}-evidence/canonical-observations.json").read_text()
        )
        rows.extend(canonical_rows)
        warning_count += sum(item["executionWarningCount"] for item in run["workbooks"])
        assert run["providerCalls"] == manifest["providerCalls"] == 0
        assert run["historicalReplayIsAcceptanceAuthority"] is False
        assert run["trainingEligibility"] is contract["trainingEligibility"] is False
        assert contract["totalEquations"] == []
        assert (
            contract["expectedWarningCountsByYear"] == manifest["warningCountsByYear"]
        )
        assert all(
            rule["code"] == "AMBIGUOUS_HEADER"
            and rule["dimension"] in {"court_level", "observation_period"}
            and rule["requireCanonicalOutputEquivalence"] is True
            and all(
                sources
                for by_output in rule["expectedHeaderSourcesByYear"].values()
                for sources in by_output.values()
            )
            for rule in contract["allowedExecutionWarnings"]
        )
        assert [item["sheet"] for item in cohort["workbooks"]] == [
            member["physicalSheetName"] for member in members
        ]
        assert all(
            item["replayResponse"]["acceptanceAuthority"] is False
            and item["replayResponse"]["historicalModel"]
            == "human-authored/deterministic-map-v1"
            for item in cohort["workbooks"]
        )
        for member in members:
            source = base / member["sourcePath"]
            assert (
                "sha256:" + hashlib.sha256(source.read_bytes()).hexdigest()
                == member["sourceDigest"]
            )

    assert len(rows) == 17327
    assert Counter(row["measure_id"] for row in rows) == {
        "defendant-count": 16463,
        "mean-case-duration": 216,
        "mean-defendant-age": 216,
        "median-case-duration": 216,
        "median-defendant-age": 216,
    }
    assert Counter(row["value_status"] for row in rows) == {
        "observed": 16968,
        "suppressed": 198,
        "not_available": 100,
        "not_applicable": 61,
    }
    assert Counter(
        (row["raw_value"], row["value_status"])
        for row in rows
        if row["value_status"] != "observed"
    ) == {
        ("np", "suppressed"): 198,
        ("na", "not_available"): 100,
        ("..", "not_applicable"): 61,
    }
    assert warning_count == 10576
    assert (
        sum(row["value"] == 0 and row["value_status"] == "observed" for row in rows)
        == 1441
    )
    assert {row["jurisdiction_id"] for row in rows} == {"WA"}
    assert {row["classification_context_id"] for row in rows} == {
        "ANZSOC_2011",
        "ANZSOC_2023",
        "MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
    }
    assert all(row["raw_value"] == 0 for row in rows if row["value"] == 0)

    method_categories = {
        "CHAR_GUILTY_EX_PARTE",
        "CHAR_TRANSFER_TO_OTHER_COURT_LEVELS",
        "CHAR_WITHDRAWN_BY_PROSECUTION",
        "CHAR_TOTAL_FINALISED",
    }
    selected = [
        row
        for row in rows
        if row.get("characteristic_category_id") in method_categories
    ]
    assert len(selected) == 2401
    assert all(
        row["characteristic_group_id"] == "GROUP_METHOD_OF_FINALISATION"
        and row["raw_characteristic_group"] == "Method of finalisation"
        for row in selected
    )
    assert all(
        row.get("characteristic_group_id") != "GROUP_GUILTY_EX_PARTE" for row in rows
    )


@pytest.mark.timeout(900)
def test_all_large_batch_cohorts_replay_cleanly(tmp_path: Path) -> None:
    report = run_batch(PROJECT, tmp_path / "batch", concurrency=3)
    assert report["passed"] is True
    assert report["providerCalls"] == 0
    assert report["acceptedWorksheetCount"] == 559
    assert report["exceptionWorksheetCount"] == 0
    assert report["canonicalObservationCount"] == 478120
    assert {item["familyId"] for item in report["cohorts"]} == {
        item.family_id for item in load_large_batch_registry(PROJECT).entries
    }
    expected_workbooks = {
        item.family_id: len(item.expected_years)
        for item in load_large_batch_registry(PROJECT).entries
    }
    assert all(
        item["passed"] is True
        and item["acceptedWorkbookCount"] == expected_workbooks[item["familyId"]]
        and item["exceptionWorkbookCount"] == 0
        and item["providerCalls"] == 0
        and item["crossYearIssues"] == []
        for item in report["cohorts"]
    )


def test_workbook_format_trim_is_deterministic(tmp_path: Path) -> None:
    output = tmp_path / "normalized.xlsx"
    completed = subprocess.run(
        [
            str(PROJECT / "scripts/trim-prototype-workbook-formatting.py"),
            str(
                PROJECT
                / "fixtures/product-prototype/workbooks/prisoners-australia-2021.xlsx"
            ),
            str(output),
            "--sheet",
            "Table_34=A1:B190",
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=30,
    )
    assert completed.returncode == 0, completed.stderr
    assert (
        output.read_bytes()
        == (
            PROJECT
            / "fixtures"
            / "product-prototype"
            / "workbooks"
            / "prisoners-australia-2021-batch-normalized.xlsx"
        ).read_bytes()
    )


def test_digest_bound_correction_emits_manifest_matching_receipt(
    tmp_path: Path,
) -> None:
    manifest = json.loads(
        (
            PROJECT / "fixtures/product-prototype/batch-workbook-normalization-v1.json"
        ).read_text()
    )
    entry = next(
        item
        for item in manifest["entries"]
        if item["year"] == 2024 and item["correction"] is not None
    )
    output = tmp_path / "corrected.xlsx"
    receipt = tmp_path / "receipt.json"
    completed = subprocess.run(
        [
            str(PROJECT / manifest["correctionScriptPath"]),
            str(PROJECT / entry["sourcePath"]),
            str(output),
            "--receipt",
            str(receipt),
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=30,
    )
    assert completed.returncode == 0, completed.stderr
    assert json.loads(receipt.read_text()) == entry["correction"]


def test_act_period_header_correction_is_exact_and_fails_closed(
    tmp_path: Path,
) -> None:
    script = PROJECT / "scripts/correct-known-workbook-artifacts.py"
    namespace = runpy.run_path(str(script))
    correct = namespace["correct"]
    declarations = namespace["CORRECTIONS"]
    source = (
        PROJECT
        / "fixtures/product-prototype/workbooks"
        / "criminal-courts-2021-22-cube-11-source.xlsx"
    )
    source_digest = hashlib.sha256(source.read_bytes()).hexdigest()
    original = copy.deepcopy(declarations[source_digest])

    output = tmp_path / "corrected.xlsx"
    receipt = correct(source, output)
    assert receipt["id"] == "criminal-courts-act-2021-22-period-header-v1"
    assert (
        load_workbook(source, read_only=True)["Table 51"]["M5"].value
        == "2022\N{EN DASH}22"
    )
    assert (
        load_workbook(output, read_only=True)["Table 51"]["M5"].value
        == "2021\N{EN DASH}22"
    )
    normalized = (
        PROJECT
        / "fixtures/product-prototype/workbooks"
        / "criminal-courts-2021-22-cube-11-normalized.xlsx"
    )
    assert (
        load_workbook(normalized, read_only=True)["Table 51"]["M5"].value
        == "2021\N{EN DASH}22"
    )

    tampered_source = tmp_path / "tampered-source.xlsx"
    tampered_source.write_bytes(source.read_bytes() + b"tampered")
    with pytest.raises(ValueError, match="exact reviewed correction source"):
        correct(tampered_source, tmp_path / "wrong-digest.xlsx")

    try:
        wrong_cell = copy.deepcopy(original)
        wrong_cell["replacedCells"][0]["cell"] = "ZZ999"
        declarations[source_digest] = wrong_cell
        with pytest.raises(
            ValueError, match="Expected exactly one Table 51!ZZ999 cell"
        ):
            correct(source, tmp_path / "wrong-cell.xlsx")

        wrong_value = copy.deepcopy(original)
        wrong_value["replacedCells"][0]["expectedValue"] = "2020\N{EN DASH}21"
        declarations[source_digest] = wrong_value
        with pytest.raises(ValueError, match="no longer matches"):
            correct(source, tmp_path / "wrong-value.xlsx")
    finally:
        declarations[source_digest] = original


def test_fdv_duplicate_footnote_correction_and_normalization_are_exact(
    tmp_path: Path,
) -> None:
    manifest_path = (
        PROJECT / "fixtures/product-prototype/batch-workbook-normalization-v1.json"
    )
    manifest = json.loads(manifest_path.read_text())
    semantic = {
        key: value for key, value in manifest.items() if key != "manifestDigest"
    }
    assert manifest["recordedAt"] == "2026-08-25T09:00:00+00:00"
    assert (
        domain_digest(manifest["schemaVersion"], semantic) == manifest["manifestDigest"]
    )
    entry = next(
        item for item in manifest["entries"] if "cube-17" in item["sourcePath"]
    )
    footnote = (
        "(f) Includes defendants for whom method of finalisation could not be "
        "determined, defendants deceased or unfit to plead, transfers to non-court "
        "agencies and other non-adjudicated finalisations n.e.c. "
    )
    assert entry["trimmedSheets"] == [
        {"sheet": "FDV Table 15", "retainedRange": "A1:G69"},
        {"sheet": "FDV Table 16", "retainedRange": "A1:G63"},
        {"sheet": "FDV Table 17", "retainedRange": "A1:G66"},
        {"sheet": "FDV Table 18", "retainedRange": "A1:F71"},
        {"sheet": "FDV Table 19", "retainedRange": "A1:G64"},
    ]
    assert entry["correction"] == {
        "id": "criminal-courts-fdv-2023-24-duplicate-footnote-v1",
        "reason": (
            "Remove the duplicate far-right Table 16 footnote at XEX59 before "
            "format trimming; preserve the identical retained footnote at A58 "
            "and the exact source workbook separately."
        ),
        "removedCells": [
            {
                "sheet": "FDV Table 16",
                "cell": "XEX59",
                "expectedStyle": "67",
                "expectedValue": footnote,
                "insideRetainedRange": False,
            }
        ],
    }
    source = PROJECT / entry["sourcePath"]
    normalized = PROJECT / entry["outputPath"]
    source_bytes = source.read_bytes()
    assert len(source_bytes) == entry["sourceByteLength"] == 85_082
    assert sha256_digest(source_bytes) == entry["sourceDigest"]
    assert sha256_digest(normalized.read_bytes()) == entry["outputDigest"]

    script_namespace = runpy.run_path(
        str(PROJECT / "scripts/correct-known-workbook-artifacts.py")
    )
    correct = script_namespace["correct"]
    sheet_targets = script_namespace["sheet_targets"]
    corrected = tmp_path / "corrected.xlsx"
    receipt = correct(source, corrected)
    assert receipt == entry["correction"]
    assert source.read_bytes() == source_bytes

    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def sheet_root(archive: zipfile.ZipFile, name: str) -> ElementTree.Element:
        return ElementTree.fromstring(archive.read(sheet_targets(archive)[name]))

    def valued_cells(root: ElementTree.Element) -> set[str]:
        return {
            cell.attrib["r"]
            for cell in root.findall(f".//{{{spreadsheet}}}c")
            if cell.find(f"{{{spreadsheet}}}f") is not None
            or (
                (value := cell.find(f"{{{spreadsheet}}}v")) is not None
                and value.text not in {None, ""}
            )
            or (
                (text := cell.find(f".//{{{spreadsheet}}}t")) is not None
                and text.text not in {None, ""}
            )
        }

    with (
        zipfile.ZipFile(source) as source_archive,
        zipfile.ZipFile(corrected) as corrected_archive,
    ):
        source_root = sheet_root(source_archive, "FDV Table 16")
        corrected_root = sheet_root(corrected_archive, "FDV Table 16")
        assert valued_cells(source_root) - valued_cells(corrected_root) == {"XEX59"}
        assert valued_cells(corrected_root) - valued_cells(source_root) == set()
        source_cells = {
            cell.attrib["r"]: cell
            for cell in source_root.findall(f".//{{{spreadsheet}}}c")
        }
        corrected_cells = {
            cell.attrib["r"]: cell
            for cell in corrected_root.findall(f".//{{{spreadsheet}}}c")
        }
        assert source_cells["XEX59"].attrib == {"r": "XEX59", "s": "67", "t": "s"}
        assert source_cells["XEX59"].find(f"{{{spreadsheet}}}v").text == "89"
        assert source_cells["A58"].find(f"{{{spreadsheet}}}v").text == "89"
        assert corrected_cells["A58"].attrib == {"r": "A58", "s": "67", "t": "s"}
        assert corrected_cells["A58"].find(f"{{{spreadsheet}}}v").text == "89"
        for name in set(sheet_targets(source_archive)) - {"FDV Table 16"}:
            assert source_archive.read(sheet_targets(source_archive)[name]) == (
                corrected_archive.read(sheet_targets(corrected_archive)[name])
            )

    reproduced = tmp_path / "normalized.xlsx"
    completed = subprocess.run(
        [
            str(PROJECT / manifest["scriptPath"]),
            str(corrected),
            str(reproduced),
            *[
                argument
                for item in entry["trimmedSheets"]
                for argument in (
                    "--sheet",
                    f"{item['sheet']}={item['retainedRange']}",
                )
            ],
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=60,
    )
    assert completed.returncode == 0, completed.stderr
    assert reproduced.read_bytes() == normalized.read_bytes()

    limits = {
        "FDV Table 15": (69, 7),
        "FDV Table 16": (63, 7),
        "FDV Table 17": (66, 7),
        "FDV Table 18": (71, 6),
        "FDV Table 19": (64, 7),
    }
    source_workbook = load_workbook(source, read_only=True, data_only=False)
    normalized_workbook = load_workbook(normalized, read_only=True, data_only=False)
    try:
        assert source_workbook.sheetnames == normalized_workbook.sheetnames
        for sheet_name, (max_row, max_column) in limits.items():
            source_rows = source_workbook[sheet_name].iter_rows(
                max_row=max_row, max_col=max_column
            )
            normalized_rows = normalized_workbook[sheet_name].iter_rows(
                max_row=max_row, max_col=max_column
            )
            for source_row, normalized_row in zip(
                source_rows, normalized_rows, strict=True
            ):
                assert [
                    (
                        cell.value,
                        getattr(cell, "data_type", "n"),
                        getattr(cell, "_style_id", 0),
                        getattr(cell, "number_format", "General"),
                    )
                    for cell in source_row
                ] == [
                    (
                        cell.value,
                        getattr(cell, "data_type", "n"),
                        getattr(cell, "_style_id", 0),
                        getattr(cell, "number_format", "General"),
                    )
                    for cell in normalized_row
                ]
    finally:
        source_workbook.close()
        normalized_workbook.close()

    expected_removed_merges = {
        "FDV Table 15": {"A2:XFD2"},
        "FDV Table 16": {"A2:XFD2", "XEX59:XFD59"},
        "FDV Table 17": {"A2:XFD2"},
        "FDV Table 18": {"A2:XFD2"},
        "FDV Table 19": {"A2:XFD2"},
    }
    with (
        zipfile.ZipFile(corrected) as corrected_archive,
        zipfile.ZipFile(normalized) as normalized_archive,
    ):
        for sheet_name, (max_row, max_column) in limits.items():
            corrected_root = sheet_root(corrected_archive, sheet_name)
            normalized_root = sheet_root(normalized_archive, sheet_name)
            corrected_merges = {
                item.attrib["ref"]
                for item in corrected_root.findall(f".//{{{spreadsheet}}}mergeCell")
            }
            normalized_merges = {
                item.attrib["ref"]
                for item in normalized_root.findall(f".//{{{spreadsheet}}}mergeCell")
            }
            assert (
                corrected_merges - normalized_merges
                == expected_removed_merges[sheet_name]
            )
            assert normalized_merges == (
                corrected_merges - expected_removed_merges[sheet_name]
            )
            for cell in corrected_root.findall(f".//{{{spreadsheet}}}c"):
                address = cell.attrib["r"]
                match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", address)
                assert match is not None
                column = 0
                for character in match.group(1):
                    column = column * 26 + ord(character) - ord("A") + 1
                outside = int(match.group(2)) > max_row or column > max_column
                assert not (outside and address in valued_cells(corrected_root))
        for name in set(sheet_targets(corrected_archive)) - set(limits):
            assert corrected_archive.read(sheet_targets(corrected_archive)[name]) == (
                normalized_archive.read(sheet_targets(normalized_archive)[name])
            )


def test_fdv_duplicate_footnote_correction_fails_closed(tmp_path: Path) -> None:
    script = PROJECT / "scripts/correct-known-workbook-artifacts.py"
    namespace = runpy.run_path(str(script))
    correct = namespace["correct"]
    declarations = namespace["CORRECTIONS"]
    source = (
        PROJECT
        / "fixtures/product-prototype/workbooks"
        / "criminal-courts-2023-24-cube-17-source.xlsx"
    )
    source_digest = hashlib.sha256(source.read_bytes()).hexdigest()
    original = copy.deepcopy(declarations[source_digest])
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rebound_digests: list[str] = []

    def write_modified_workbook(
        name: str, archive_path: str, replacement: bytes
    ) -> Path:
        modified = tmp_path / name
        with (
            zipfile.ZipFile(source) as archive,
            zipfile.ZipFile(modified, "w") as destination,
        ):
            for info in archive.infolist():
                destination.writestr(
                    copy.copy(info),
                    replacement
                    if info.filename == archive_path
                    else archive.read(info),
                )
        digest = hashlib.sha256(modified.read_bytes()).hexdigest()
        declaration = copy.deepcopy(original)
        declaration["byteLength"] = len(modified.read_bytes())
        declarations[digest] = declaration
        rebound_digests.append(digest)
        return modified

    tampered_source = tmp_path / "wrong-digest.xlsx"
    tampered_source.write_bytes(source.read_bytes() + b"tampered")
    with pytest.raises(ValueError, match="exact reviewed correction source"):
        correct(tampered_source, tmp_path / "wrong-digest-output.xlsx")

    try:
        wrong_length = copy.deepcopy(original)
        wrong_length["byteLength"] += 1
        declarations[source_digest] = wrong_length
        with pytest.raises(ValueError, match="exact reviewed correction source"):
            correct(source, tmp_path / "wrong-length.xlsx")

        wrong_sheet = copy.deepcopy(original)
        wrong_sheet["cells"][0]["sheet"] = "FDV Table 15"
        declarations[source_digest] = wrong_sheet
        with pytest.raises(
            ValueError,
            match=r"Expected exactly one FDV Table 15!XEX59 cell",
        ):
            correct(source, tmp_path / "wrong-sheet.xlsx")

        mutations = [
            ("cell", "ZZ999", r"Expected exactly one FDV Table 16!ZZ999 cell"),
            ("expectedStyle", "999", "no longer matches"),
            ("expectedType", "str", "no longer matches"),
            ("expectedValue", "different footnote", "no longer matches"),
            ("expectedMerge", "XEX59:XFC59", "no longer matches"),
            ("retainedRange", "B1:G63", "no longer matches"),
            ("retainedRange", "A1B2:G63", "Invalid cell address"),
        ]
        for key, value, message in mutations:
            changed = copy.deepcopy(original)
            changed["cells"][0][key] = value
            declarations[source_digest] = changed
            with pytest.raises(ValueError, match=message):
                correct(source, tmp_path / f"wrong-{key}.xlsx")
        for key, value in [
            ("cell", "A59"),
            ("expectedStyle", "999"),
            ("expectedType", "str"),
            ("expectedValue", "different duplicate footnote"),
            ("expectedMerge", "A59:G59"),
        ]:
            changed = copy.deepcopy(original)
            changed["cells"][0]["duplicateOf"][key] = value
            declarations[source_digest] = changed
            with pytest.raises(ValueError, match="no longer matches"):
                correct(source, tmp_path / f"wrong-duplicate-{key}.xlsx")

        with zipfile.ZipFile(source) as archive:
            target = namespace["sheet_targets"](archive)["FDV Table 16"]
            source_sheet = archive.read(target)
            root = ElementTree.fromstring(source_sheet)
            cells = {
                item.attrib["r"]: item
                for item in root.findall(f".//{{{spreadsheet}}}c")
            }
            assert cells["XEX59"].find(f"{{{spreadsheet}}}f") is None
            ElementTree.SubElement(cells["XEX59"], f"{{{spreadsheet}}}f").text = "1+1"
            target_formula_source = write_modified_workbook(
                "target-formula-drift.xlsx",
                target,
                ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
            )
        with pytest.raises(ValueError, match="no longer matches"):
            correct(target_formula_source, tmp_path / "target-formula-output.xlsx")

        root = ElementTree.fromstring(source_sheet)
        duplicate_cell = next(
            item
            for item in root.findall(f".//{{{spreadsheet}}}c")
            if item.attrib["r"] == "A58"
        )
        assert duplicate_cell.find(f"{{{spreadsheet}}}f") is None
        ElementTree.SubElement(duplicate_cell, f"{{{spreadsheet}}}f").text = "1+1"
        duplicate_formula_source = write_modified_workbook(
            "duplicate-formula-drift.xlsx",
            target,
            ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
        )
        with pytest.raises(ValueError, match="no longer matches"):
            correct(
                duplicate_formula_source,
                tmp_path / "duplicate-formula-output.xlsx",
            )

        root = ElementTree.fromstring(source_sheet)
        target_cell = next(
            item
            for item in root.findall(f".//{{{spreadsheet}}}c")
            if item.attrib["r"] == "XEX59"
        )
        target_value = target_cell.find(f"{{{spreadsheet}}}v")
        assert target_value is not None and target_value.text == "89"
        target_value.text = "212"
        index_drift_source = write_modified_workbook(
            "shared-string-index-drift.xlsx",
            target,
            ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
        )
        with pytest.raises(ValueError, match="no longer matches"):
            correct(index_drift_source, tmp_path / "index-drift-output.xlsx")

        with zipfile.ZipFile(source) as archive:
            shared_strings_path = "xl/sharedStrings.xml"
            shared_root = ElementTree.fromstring(archive.read(shared_strings_path))
        shared_items = shared_root.findall(f"{{{spreadsheet}}}si")
        shared_text = next(shared_items[89].iter(f"{{{spreadsheet}}}t"))
        assert shared_text.text is not None and shared_text.text.startswith(
            "(f) Includes defendants"
        )
        shared_text.text += "drift"
        shared_value_drift_source = write_modified_workbook(
            "shared-string-value-drift.xlsx",
            shared_strings_path,
            ElementTree.tostring(shared_root, encoding="utf-8", xml_declaration=True),
        )
        with pytest.raises(ValueError, match="no longer matches"):
            correct(
                shared_value_drift_source,
                tmp_path / "shared-value-drift-output.xlsx",
            )
    finally:
        declarations[source_digest] = original
        for digest in rebound_digests:
            declarations.pop(digest, None)


def test_fdv_offence_geometry_normalization_is_exact(tmp_path: Path) -> None:
    source = (
        PROJECT
        / "fixtures/product-prototype/workbooks"
        / "criminal-courts-2023-24-cube-16-source.xlsx"
    )
    normalized = (
        PROJECT
        / "fixtures/product-prototype/workbooks"
        / "criminal-courts-2023-24-cube-16-normalized.xlsx"
    )
    source_bytes = source.read_bytes()
    assert len(source_bytes) == 413_928
    assert hashlib.sha256(source_bytes).hexdigest() == (
        "65e4e00dc4062415fb33e6716135b5f18751a1b6589ed01af1943542c5b59815"
    )
    normalized_bytes = normalized.read_bytes()
    assert len(normalized_bytes) == 252_687
    assert hashlib.sha256(normalized_bytes).hexdigest() == (
        "52118023aefddf18a8197ea44c2de441050383c083d1505c99a41e0fd2a146fc"
    )
    manifest = json.loads(
        (
            PROJECT / "fixtures/product-prototype/batch-workbook-normalization-v1.json"
        ).read_text()
    )
    manifest_entry = next(
        entry
        for entry in manifest["entries"]
        if "cube-16-source" in entry["sourcePath"]
    )
    assert manifest_entry == {
        "year": 2023,
        "sourcePath": (
            "fixtures/product-prototype/workbooks/"
            "criminal-courts-2023-24-cube-16-source.xlsx"
        ),
        "sourceDigest": (
            "sha256:65e4e00dc4062415fb33e6716135b5f18751a1b6589ed01af1943542c5b59815"
        ),
        "sourceByteLength": 413_928,
        "outputPath": (
            "fixtures/product-prototype/workbooks/"
            "criminal-courts-2023-24-cube-16-normalized.xlsx"
        ),
        "outputDigest": (
            "sha256:52118023aefddf18a8197ea44c2de441050383c083d1505c99a41e0fd2a146fc"
        ),
        "outputByteLength": 252_687,
        "trimmedSheets": [
            {"sheet": "FDV Table 8", "retainedRange": "A1:L256"},
            {"sheet": "FDV Table 13", "retainedRange": "A1:K261"},
        ],
        "correction": {
            "id": "criminal-courts-fdv-offence-2023-24-geometry-v1",
            "reason": (
                "Validate the exact empty far-right Table 8 merge geometry and the "
                "full-width Table 13 title merge before format trimming; preserve "
                "the legitimate Table 8 A2:L2 title merge and exact source workbook."
            ),
            "validatedTrim": [
                {
                    "sheet": "FDV Table 8",
                    "retainedRange": "A1:L256",
                    "removedMergeCount": 4_680,
                    "removedMergeDigest": (
                        "db11301c90f1ff90c226aaf0daa186375e0926908a3cb70c5939763cde410e26"
                    ),
                },
                {
                    "sheet": "FDV Table 13",
                    "retainedRange": "A1:K261",
                    "removedMergeCount": 1,
                    "removedMergeDigest": (
                        "a6b604395e7a58579c5654f026de5ed1507c32b78d5a8480ed738c5444e7fd58"
                    ),
                },
            ],
        },
        "normalization": "trim-pathological-full-width-formatting-merge-v1",
    }

    script = PROJECT / "scripts/correct-known-workbook-artifacts.py"
    namespace = runpy.run_path(str(script))
    corrected = tmp_path / "corrected.xlsx"
    receipt = namespace["correct"](source, corrected)
    assert receipt == {
        "id": "criminal-courts-fdv-offence-2023-24-geometry-v1",
        "reason": (
            "Validate the exact empty far-right Table 8 merge geometry and the "
            "full-width Table 13 title merge before format trimming; preserve "
            "the legitimate Table 8 A2:L2 title merge and exact source workbook."
        ),
        "validatedTrim": [
            {
                "sheet": "FDV Table 8",
                "retainedRange": "A1:L256",
                "removedMergeCount": 4_680,
                "removedMergeDigest": (
                    "db11301c90f1ff90c226aaf0daa186375e0926908a3cb70c5939763cde410e26"
                ),
            },
            {
                "sheet": "FDV Table 13",
                "retainedRange": "A1:K261",
                "removedMergeCount": 1,
                "removedMergeDigest": (
                    "a6b604395e7a58579c5654f026de5ed1507c32b78d5a8480ed738c5444e7fd58"
                ),
            },
        ],
    }
    assert source.read_bytes() == source_bytes
    assert corrected.read_bytes() == source_bytes

    reproduced = tmp_path / "normalized.xlsx"
    trim = subprocess.run(
        [
            str(PROJECT / "scripts/trim-prototype-workbook-formatting.py"),
            str(corrected),
            str(reproduced),
            "--sheet",
            "FDV Table 8=A1:L256",
            "--sheet",
            "FDV Table 13=A1:K261",
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=60,
    )
    assert trim.returncode == 0, trim.stderr
    assert reproduced.read_bytes() == normalized_bytes

    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    limits = {
        "FDV Table 8": (256, 12, "A1:XFD256", "A1:L256"),
        "FDV Table 13": (261, 11, "A1:AT261", "A1:K261"),
    }
    expected_removed = {
        "FDV Table 8": (
            4_680,
            "db11301c90f1ff90c226aaf0daa186375e0926908a3cb70c5939763cde410e26",
        ),
        "FDV Table 13": (
            1,
            "a6b604395e7a58579c5654f026de5ed1507c32b78d5a8480ed738c5444e7fd58",
        ),
    }

    def merge_digest(merges: set[str]) -> str:
        return hashlib.sha256(("\n".join(sorted(merges)) + "\n").encode()).hexdigest()

    def root_for(archive: zipfile.ZipFile, sheet: str) -> ElementTree.Element:
        target = namespace["sheet_targets"](archive)[sheet]
        return ElementTree.fromstring(archive.read(target))

    def cell_has_content(cell: ElementTree.Element) -> bool:
        return any(
            item.tag.rsplit("}", 1)[-1] == "f"
            or (
                item.tag.rsplit("}", 1)[-1] in {"v", "t"}
                and item.text not in {None, ""}
            )
            for item in cell.iter()
        )

    with (
        zipfile.ZipFile(source) as source_archive,
        zipfile.ZipFile(normalized) as normalized_archive,
    ):
        assert list(namespace["sheet_targets"](source_archive)) == list(
            namespace["sheet_targets"](normalized_archive)
        )
        source_targets = namespace["sheet_targets"](source_archive)
        normalized_targets = namespace["sheet_targets"](normalized_archive)
        assert (
            sum(
                len(root_for(source_archive, sheet).findall(f".//{{{spreadsheet}}}f"))
                for sheet in source_targets
            )
            == 0
        )
        assert (
            sum(
                len(
                    root_for(normalized_archive, sheet).findall(
                        f".//{{{spreadsheet}}}f"
                    )
                )
                for sheet in normalized_targets
            )
            == 0
        )
        for sheet, (_, _, source_dimension, output_dimension) in limits.items():
            source_root = root_for(source_archive, sheet)
            output_root = root_for(normalized_archive, sheet)
            assert source_root.find(f"{{{spreadsheet}}}dimension").attrib["ref"] == (
                source_dimension
            )
            assert output_root.find(f"{{{spreadsheet}}}dimension").attrib["ref"] == (
                output_dimension
            )
            source_merges = {
                item.attrib["ref"]
                for item in source_root.findall(f".//{{{spreadsheet}}}mergeCell")
            }
            output_merges = {
                item.attrib["ref"]
                for item in output_root.findall(f".//{{{spreadsheet}}}mergeCell")
            }
            removed = source_merges - output_merges
            expected_count, expected_digest = expected_removed[sheet]
            assert len(removed) == expected_count
            assert merge_digest(removed) == expected_digest
            assert output_merges == source_merges - removed
            output_addresses = {
                retained.attrib["r"]
                for retained in output_root.findall(f".//{{{spreadsheet}}}c")
            }
            assert all(
                not cell_has_content(cell)
                for cell in source_root.findall(f".//{{{spreadsheet}}}c")
                if cell.attrib["r"] not in output_addresses
            )
        assert "A2:L2" in {
            item.attrib["ref"]
            for item in root_for(normalized_archive, "FDV Table 8").findall(
                f".//{{{spreadsheet}}}mergeCell"
            )
        }
        assert "A2:XFD2" not in {
            item.attrib["ref"]
            for item in root_for(normalized_archive, "FDV Table 13").findall(
                f".//{{{spreadsheet}}}mergeCell"
            )
        }
        for sheet in set(source_targets) - set(limits):
            assert source_archive.read(
                source_targets[sheet]
            ) == normalized_archive.read(normalized_targets[sheet])

    source_workbook = load_workbook(source, read_only=True, data_only=False)
    normalized_workbook = load_workbook(normalized, read_only=True, data_only=False)
    try:
        assert source_workbook.sheetnames == normalized_workbook.sheetnames
        for sheet, (max_row, max_column, _, _) in limits.items():
            assert normalized_workbook[sheet].max_row == max_row
            assert normalized_workbook[sheet].max_column == max_column
            for source_row, output_row in zip(
                source_workbook[sheet].iter_rows(
                    max_row=max_row,
                    max_col=max_column,
                ),
                normalized_workbook[sheet].iter_rows(
                    max_row=max_row,
                    max_col=max_column,
                ),
                strict=True,
            ):
                assert [
                    (
                        cell.value,
                        getattr(cell, "data_type", "n"),
                        getattr(cell, "_style_id", 0),
                        getattr(cell, "number_format", "General"),
                    )
                    for cell in source_row
                ] == [
                    (
                        cell.value,
                        getattr(cell, "data_type", "n"),
                        getattr(cell, "_style_id", 0),
                        getattr(cell, "number_format", "General"),
                    )
                    for cell in output_row
                ]
    finally:
        source_workbook.close()
        normalized_workbook.close()


def test_fdv_offence_geometry_correction_fails_closed(tmp_path: Path) -> None:
    script = PROJECT / "scripts/correct-known-workbook-artifacts.py"
    namespace = runpy.run_path(str(script))
    correct = namespace["correct"]
    declarations = namespace["CORRECTIONS"]
    source = (
        PROJECT
        / "fixtures/product-prototype/workbooks"
        / "criminal-courts-2023-24-cube-16-source.xlsx"
    )
    source_digest = "65e4e00dc4062415fb33e6716135b5f18751a1b6589ed01af1943542c5b59815"
    original = copy.deepcopy(declarations[source_digest])
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rebound_digests: list[str] = []

    def write_modified_workbook(
        name: str,
        sheet: str,
        mutate: Callable[[ElementTree.Element], None],
        *,
        rebind_sheet_digest: bool = False,
    ) -> Path:
        modified = tmp_path / name
        with zipfile.ZipFile(source) as archive:
            target = namespace["sheet_targets"](archive)[sheet]
            root = ElementTree.fromstring(archive.read(target))
            mutate(root)
            replacement = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )
            with zipfile.ZipFile(modified, "w") as destination:
                for info in archive.infolist():
                    destination.writestr(
                        copy.copy(info),
                        replacement if info.filename == target else archive.read(info),
                    )
        modified_bytes = modified.read_bytes()
        digest = hashlib.sha256(modified_bytes).hexdigest()
        declaration = copy.deepcopy(original)
        declaration["byteLength"] = len(modified_bytes)
        if rebind_sheet_digest:
            guard = next(
                item for item in declaration["trimGuards"] if item["sheet"] == sheet
            )
            guard["expectedSheetDigest"] = hashlib.sha256(replacement).hexdigest()
        declarations[digest] = declaration
        rebound_digests.append(digest)
        return modified

    wrong_digest = tmp_path / "wrong-digest.xlsx"
    wrong_digest.write_bytes(source.read_bytes() + b"tampered")
    with pytest.raises(ValueError, match="exact reviewed correction source"):
        correct(wrong_digest, tmp_path / "wrong-digest-output.xlsx")

    try:
        wrong_length = copy.deepcopy(original)
        wrong_length["byteLength"] += 1
        declarations[source_digest] = wrong_length
        with pytest.raises(ValueError, match="exact reviewed correction source"):
            correct(source, tmp_path / "wrong-length.xlsx")

        mutations = [
            ("sheet", "FDV Table 7"),
            ("retainedRange", "A1:K256"),
            ("expectedSheetDigest", "0" * 64),
            ("expectedDimension", "A1:L256"),
            ("expectedCellCount", 35_551),
            ("expectedValuedCellCount", 2_039),
            ("expectedFormulaCount", 1),
            ("expectedOutOfRangeCellCount", 32_743),
            ("expectedOutOfRangeValuedCellCount", 1),
            ("expectedMergeCount", 4_696),
            ("expectedRetainedMergeCount", 16),
            ("expectedRetainedMergeDigest", "1" * 64),
            ("expectedRemovedMergeCount", 4_679),
            ("expectedRemovedMergeDigest", "2" * 64),
        ]
        for key, value in mutations:
            changed = copy.deepcopy(original)
            changed["trimGuards"][0][key] = value
            declarations[source_digest] = changed
            with pytest.raises(ValueError, match="trim guard"):
                correct(source, tmp_path / f"wrong-{key}.xlsx")

        declarations[source_digest] = original

        def drift_style(root: ElementTree.Element) -> None:
            cell = next(
                item
                for item in root.findall(f".//{{{spreadsheet}}}c")
                if item.attrib["r"] == "A1"
            )
            cell.attrib["s"] = "999"

        def drift_type(root: ElementTree.Element) -> None:
            cell = next(
                item
                for item in root.findall(f".//{{{spreadsheet}}}c")
                if item.attrib["r"] == "A1"
            )
            cell.attrib["t"] = "str"

        def drift_value(root: ElementTree.Element) -> None:
            cell = next(
                item
                for item in root.findall(f".//{{{spreadsheet}}}c")
                if item.attrib["r"] == "A1"
            )
            value = cell.find(f"{{{spreadsheet}}}v")
            assert value is not None
            value.text = "0"

        for name, mutate in [
            ("style", drift_style),
            ("type", drift_type),
            ("value", drift_value),
        ]:
            modified = write_modified_workbook(
                f"{name}-drift.xlsx",
                "FDV Table 8",
                mutate,
            )
            with pytest.raises(ValueError, match="trim guard"):
                correct(modified, tmp_path / f"{name}-drift-output.xlsx")

        def add_outside_formula(root: ElementTree.Element) -> None:
            cell = next(
                item
                for item in root.findall(f".//{{{spreadsheet}}}c")
                if item.attrib["r"] == "M252"
            )
            ElementTree.SubElement(cell, f"{{{spreadsheet}}}f").text = "1+1"

        formula_drift = write_modified_workbook(
            "formula-drift.xlsx",
            "FDV Table 8",
            add_outside_formula,
            rebind_sheet_digest=True,
        )
        with pytest.raises(ValueError, match="trim guard"):
            correct(formula_drift, tmp_path / "formula-drift-output.xlsx")

        def drift_merge(root: ElementTree.Element) -> None:
            merge = next(
                item
                for item in root.findall(f".//{{{spreadsheet}}}mergeCell")
                if item.attrib["ref"] == "A2:L2"
            )
            merge.attrib["ref"] = "A2:K2"

        merge_drift = write_modified_workbook(
            "merge-drift.xlsx",
            "FDV Table 8",
            drift_merge,
            rebind_sheet_digest=True,
        )
        with pytest.raises(ValueError, match="trim guard"):
            correct(merge_drift, tmp_path / "merge-drift-output.xlsx")
    finally:
        declarations[source_digest] = original
        for digest in rebound_digests:
            declarations.pop(digest, None)


def test_prior_digest_bound_correction_outputs_remain_byte_identical(
    tmp_path: Path,
) -> None:
    manifest = json.loads(
        (
            PROJECT / "fixtures/product-prototype/batch-workbook-normalization-v1.json"
        ).read_text()
    )
    entries = [
        entry
        for entry in manifest["entries"]
        if entry["correction"] is not None
        and "cube-17" not in entry["sourcePath"]
        and "cube-16" not in entry["sourcePath"]
    ]
    assert len(entries) == 3
    for index, entry in enumerate(entries):
        corrected = tmp_path / f"corrected-{index}.xlsx"
        receipt = tmp_path / f"receipt-{index}.json"
        correction = subprocess.run(
            [
                str(PROJECT / manifest["correctionScriptPath"]),
                str(PROJECT / entry["sourcePath"]),
                str(corrected),
                "--receipt",
                str(receipt),
            ],
            cwd=PROJECT,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert correction.returncode == 0, correction.stderr
        assert json.loads(receipt.read_text()) == entry["correction"]
        reproduced = tmp_path / f"normalized-{index}.xlsx"
        trim = subprocess.run(
            [
                str(PROJECT / manifest["scriptPath"]),
                str(corrected),
                str(reproduced),
                *[
                    argument
                    for item in entry["trimmedSheets"]
                    for argument in (
                        "--sheet",
                        f"{item['sheet']}={item['retainedRange']}",
                    )
                ],
            ],
            cwd=PROJECT,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert trim.returncode == 0, trim.stderr
        assert reproduced.read_bytes() == (PROJECT / entry["outputPath"]).read_bytes()


def _source_cell(address: str) -> str:
    match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", address)
    assert match is not None
    column = 0
    for character in match.group(1):
        column = column * 26 + ord(character) - ord("A") + 1
    return f"R{match.group(2)}C{column}"


def test_prisoners_state_cluster_geometry_and_acceptance_are_independent() -> None:
    completed = subprocess.run(
        [
            "uv",
            "run",
            "python",
            str(PROJECT / "scripts/generate-prisoners-state-cluster.py"),
            "--check",
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=30,
    )
    assert completed.returncode == 0, completed.stderr
    geometry = json.loads(
        (
            PROJECT
            / "fixtures/product-prototype/prisoners-state-cluster-geometry-v1.json"
        ).read_text()
    )
    audit = json.loads(
        (
            PROJECT
            / "fixtures/product-prototype"
            / "prisoners-state-cluster-acceptance-audit-v1.json"
        ).read_text()
    )
    assert geometry["authority"] == "human-authored-reviewed-physical-geometry"
    assert audit["authority"] == "independent-of-replay-output"
    assert audit["handoffCorrection"] == {
        "year": 2025,
        "sheet": "Table 16",
        "publishedGrid": "B:K",
        "publishedColumns": 10,
        "expectedRows": 900,
        "familyRows": 4068,
        "clusterRows": 8406,
        "reason": audit["handoffCorrection"]["reason"],
    }
    assert sum(item["expectedCanonicalCount"] for item in audit["families"]) == 8406
    assert sum(len(item["members"]) for item in geometry["families"]) == 24

    geometry_by_family = {item["familyId"]: item for item in geometry["families"]}
    crosswalk = json.loads(
        (
            PROJECT
            / "fixtures/product-prototype/prisoners-release-family-crosswalk-v1.json"
        ).read_text()
    )
    crosswalk_by_family = {
        item["familyId"]: item["members"] for item in crosswalk["families"]
    }
    expected_workbook_digests = {
        2021: "sha256:9a5be165da58005a2a31634568491645192785a96f27d9eee3c17e45175d1710",
        2022: "sha256:a16a47e574d8da8d851f904f8ee60324cac870a89e76f4ea9114680bdde40a2b",
        2023: "sha256:61366170db7a4da717332a3ad1ca9e11f884d95905bb4556f5f44ce51d31c66f",
        2024: "sha256:609a96a96e2e359ae3e534252bcb1a6b6a329eb91fcf289834d1014dc61273d1",
        2025: "sha256:007a1c21fc2a2b256cbde672405a4710edcac85eff035949f403ca3fbed6ab6e",
    }
    for family in audit["families"]:
        evidence = json.loads(
            (
                PROJECT
                / "fixtures/product-prototype"
                / f"{family['familyId']}-five-year-evidence/canonical-observations.json"
            ).read_text()
        )
        assert len(evidence) == family["expectedCanonicalCount"]
        assert (
            dict(sorted(Counter(row["measure_id"] for row in evidence).items()))
            == family["measureCounts"]
        )
        cohort = json.loads(
            (
                PROJECT
                / "fixtures/product-prototype"
                / f"prisoners-{family['familyId']}.json"
            ).read_text()
        )
        reviewed_members = crosswalk_by_family[family["familyId"]]
        assert all(item["cube"] == 2 for item in reviewed_members)
        assert [(item["year"], item["sheet"]) for item in reviewed_members] == [
            (item["year"], item["sheet"]) for item in cohort["workbooks"]
        ]
        assert all(
            item["contentDigest"] == expected_workbook_digests[item["year"]]
            for item in cohort["workbooks"]
        )
        assert {
            item["year"] for item in cohort["workbooks"] if "normalization" in item
        } == ({2021, 2022, 2023, 2025} & set(family["years"]))
        rows_by_digest = {
            workbook["contentDigest"]: [
                row
                for row in evidence
                if row["source_workbook_digest"] == workbook["contentDigest"]
            ]
            for workbook in cohort["workbooks"]
        }
        assert [
            len(rows_by_digest[item["contentDigest"]]) for item in cohort["workbooks"]
        ] == family["expectedYearCounts"]
        for member, workbook in zip(
            geometry_by_family[family["familyId"]]["members"],
            cohort["workbooks"],
            strict=True,
        ):
            cells = {
                row["source_cell"] for row in rows_by_digest[workbook["contentDigest"]]
            }
            for band in member["valueBands"]:
                first, last = band.split(":")
                assert {_source_cell(first), _source_cell(last)} <= cells


def test_prisoners_state_cluster_retains_vintages_totals_and_null_markers() -> None:
    base = PROJECT / "fixtures/product-prototype"
    crude = json.loads(
        (base / "prisoners-state-crude-imprisonment-rate.json").read_text()
    )
    assert [item["year"] for item in crude["workbooks"]] == [2021, 2022, 2023, 2024]
    assert all(
        item["sheet"].replace("_", " ") == "Table 19" for item in crude["workbooks"]
    )

    selected = json.loads(
        (
            base
            / "state-selected-characteristics-time-series-five-year-evidence"
            / "canonical-observations.json"
        ).read_text()
    )
    assert all(
        row["reference_date"] == row["observation_period_id"] for row in selected
    )
    assert any(
        row["publication_vintage_date"] != row["reference_date"] for row in selected
    )
    vintages_by_period: dict[str, set[str]] = {}
    for row in selected:
        vintages_by_period.setdefault(row["reference_date"], set()).add(
            row["publication_vintage_date"]
        )
    assert any(len(vintages) > 1 for vintages in vintages_by_period.values())

    sex_status = json.loads(
        (
            base
            / "state-sex-by-indigenous-status-five-year-evidence"
            / "canonical-observations.json"
        ).read_text()
    )
    markers = Counter(
        (row["raw_value"], row["value_status"])
        for row in sex_status
        if row["value_status"] != "observed"
    )
    assert set(markers) == {
        ("n.p.", "suppressed"),
        ("np", "suppressed"),
        ("n.a.", "not_applicable"),
        ("n.a", "not_applicable"),
        ("na", "not_applicable"),
    }
    assert all(
        row["value"] is None for row in sex_status if row["value_status"] != "observed"
    )
    count_rows = [row for row in sex_status if row["measure_id"] == "prisoner-count"]
    assert len(count_rows) == 81
    assert {row["source_workbook_digest"] for row in count_rows} == {
        "sha256:007a1c21fc2a2b256cbde672405a4710edcac85eff035949f403ca3fbed6ab6e"
    }
    assert all(row["unit_id"] == "person" for row in count_rows)
    assert {"AUS"} <= {row["jurisdiction_id"] for row in sex_status}
    assert {"PERSONS"} <= {row["sex_id"] for row in sex_status}
    assert {"TOTAL"} <= {row["indigenous_status_id"] for row in sex_status}
    assert any(
        row["measure_id"] == "indigenous-to-non-indigenous-rate-ratio"
        and row["statistic_basis_id"] == "RATE_RATIO"
        and row["rate_basis_id"] in {"CRUDE", "AGE_STANDARDISED"}
        for row in sex_status
    )
    for family in (
        "state-selected-characteristics-time-series",
        "state-sex-by-indigenous-status",
        "state-age-standardised-rate-by-indigenous-status",
        "state-crude-imprisonment-rate",
        "state-crude-rate-by-indigenous-status",
    ):
        contract = json.loads(
            (base / f"acceptance/prisoners-{family}-v1.json").read_text()
        )
        assert contract["totalValidation"] == "not_applicable"
        assert contract["totalEquations"] == []


def _fdv_offence_families() -> list[dict[str, object]]:
    membership = json.loads(
        (
            PROJECT
            / "fixtures/product-prototype"
            / "criminal-courts-release-family-membership-v1.json"
        ).read_text()
    )
    return sorted(
        (
            family
            for family in membership["families"]
            if family["members"]
            and all(
                member["cubeId"].startswith("family-and-domestic-violence-offences")
                for member in family["members"]
            )
        ),
        key=lambda family: family["familyId"],
    )


def _worksheet_xml(path: Path, sheet: str) -> ElementTree.Element:
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package = "http://schemas.openxmlformats.org/package/2006/relationships"
    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relation_id = next(
            item.attrib[f"{{{relationships}}}id"]
            for item in workbook.findall(f".//{{{main}}}sheet")
            if item.attrib["name"] == sheet
        )
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = next(
            item.attrib["Target"]
            for item in rels.findall(f"{{{package}}}Relationship")
            if item.attrib["Id"] == relation_id
        )
        member = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
        return ElementTree.fromstring(archive.read(member))


def _fdv_sheet_cells(
    cache: dict[tuple[Path, str], tuple[dict[str, bytes], dict[str, str]]],
    path: Path,
    sheet: str,
) -> tuple[dict[str, bytes], dict[str, str]]:
    key = (path, sheet)
    if key not in cache:
        main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        root = _worksheet_xml(path, sheet)
        cells: dict[str, bytes] = {}
        formulas: dict[str, str] = {}
        for cell in root.findall(f".//{{{main}}}c"):
            address = cell.attrib["r"]
            cells[address] = ElementTree.tostring(cell, encoding="utf-8")
            formula = cell.find(f"{{{main}}}f")
            if formula is not None:
                formulas[address] = "=" + (formula.text or "")
        cache[key] = cells, formulas
    return cache[key]


def _a1_from_source_cell(address: str) -> str:
    match = re.fullmatch(r"R([1-9][0-9]*)C([1-9][0-9]*)", address)
    assert match is not None
    return f"{get_column_letter(int(match.group(2)))}{match.group(1)}"


def test_registered_fdv_formula_census_and_selected_source_cells_are_exact() -> None:
    fixture_root = PROJECT / "fixtures/product-prototype"
    families = _fdv_offence_families()
    assert len(families) == 31
    members = [
        (family["familyId"], member)
        for family in families
        for member in family["members"]
    ]
    assert len(members) == 56
    cell_cache: dict[tuple[Path, str], tuple[dict[str, bytes], dict[str, str]]] = {}
    formulas: list[tuple[str, str, str, str, str]] = []
    formula_counts: Counter[str] = Counter()
    historical_sheet_counts: Counter[tuple[str, str]] = Counter()
    source_member_by_release_sheet = {}
    for _family_id, member in members:
        release = member["releaseId"]
        sheet = member["physicalSheetName"]
        source_path = fixture_root / member["sourcePath"]
        source_member_by_release_sheet[(release, sheet)] = member
        _cells, sheet_formulas = _fdv_sheet_cells(cell_cache, source_path, sheet)
        formula_counts[release] += len(sheet_formulas)
        if release in {"2021-22", "2022-23"}:
            historical_sheet_counts[(release, sheet)] = len(sheet_formulas)
        for coordinate, text in sheet_formulas.items():
            formulas.append((release, member["sourceDigest"], sheet, coordinate, text))
    assert formula_counts == Counter(
        {"2021-22": 39, "2022-23": 39, "2023-24": 0, "2024-25": 0}
    )
    assert len(historical_sheet_counts) == 26
    assert set(historical_sheet_counts.values()) == {3}
    assert {item[4] for item in formulas} == {
        "=Contents!A2",
        "=Contents!A3",
        "=Contents!B34",
    }
    footer_rows = {
        "2021-22": {
            1: 74,
            2: 72,
            3: 74,
            4: 189,
            5: 138,
            6: 143,
            7: 131,
            8: 143,
            9: 98,
            10: 131,
            11: 143,
            12: 143,
            13: 135,
        },
        "2022-23": {
            1: 74,
            2: 78,
            3: 74,
            4: 189,
            5: 181,
            6: 200,
            7: 184,
            8: 200,
            9: 152,
            10: 184,
            11: 200,
            12: 200,
            13: 192,
        },
    }
    for release, rows_by_table in footer_rows.items():
        for table, footer_row in rows_by_table.items():
            observed = {
                (coordinate, text)
                for item_release, _digest, sheet, coordinate, text in formulas
                if item_release == release and sheet == f"FDV Table {table}"
            }
            assert observed == {
                ("A2", "=Contents!A2"),
                ("A3", "=Contents!A3"),
                (f"A{footer_row}", "=Contents!B34"),
            }
    formula_payload = "".join(
        "\t".join(item) + "\n" for item in sorted(formulas)
    ).encode()
    assert sha256_digest(formula_payload) == (
        "sha256:6b4a47b14bf27809a9d0c7633b1d89f943f7b6940bcb5538ec91a8768ef87de1"
    )

    selected: set[tuple[str, str, str]] = set()
    selected_by_workbook: dict[tuple[Path, str], set[str]] = {}
    selected_formula_count = 0
    normalized_parity_count = 0
    for family in families:
        family_id = family["familyId"]
        cohort = json.loads((fixture_root / f"{family_id}.json").read_text())
        rows = json.loads(
            (
                fixture_root / f"{family_id}-evidence" / "canonical-observations.json"
            ).read_text()
        )
        workbook_by_identity = {
            (item["contentDigest"], item["sheet"]): item for item in cohort["workbooks"]
        }
        for row in rows:
            identity = (
                row["source_workbook_digest"],
                row["source_sheet"],
                row["source_cell"],
            )
            assert identity not in selected
            selected.add(identity)
            workbook = workbook_by_identity[
                (row["source_workbook_digest"], row["source_sheet"])
            ]
            workbook_path = fixture_root / workbook["path"]
            a1 = _a1_from_source_cell(row["source_cell"])
            selected_by_workbook.setdefault(
                (workbook_path, row["source_sheet"]), set()
            ).add(a1)
    assert len(selected) == 76_189
    for (workbook_path, sheet), addresses in selected_by_workbook.items():
        _cells, workbook_formulas = _fdv_sheet_cells(cell_cache, workbook_path, sheet)
        selected_formula_count += len(addresses & set(workbook_formulas))
        if "2023-24-cube-16-normalized" not in workbook_path.name:
            continue
        source_member = source_member_by_release_sheet[("2023-24", sheet)]
        source_path = fixture_root / source_member["sourcePath"]
        source_cells, _source_formulas = _fdv_sheet_cells(
            cell_cache, source_path, sheet
        )
        normalized_cells, _normalized_formulas = _fdv_sheet_cells(
            cell_cache, workbook_path, sheet
        )
        for address in addresses:
            assert normalized_cells.get(address) == source_cells.get(address)
            normalized_parity_count += 1
    assert selected_formula_count == 0
    assert normalized_parity_count == 25_203


def test_registered_fdv_json_csv_projection_and_whitespace_delta_are_exact() -> None:
    fixture_root = PROJECT / "fixtures/product-prototype"
    difference_count = 0
    differences: set[tuple[str, str, str, str]] = set()
    row_count = 0
    for family in _fdv_offence_families():
        family_id = family["familyId"]
        contract = json.loads(
            (fixture_root / "acceptance" / f"{family_id}-v1.json").read_text()
        )
        evidence_root = fixture_root / f"{family_id}-evidence"
        rows = json.loads((evidence_root / "canonical-observations.json").read_text())
        csv_bytes = (evidence_root / "canonical-observations.csv").read_bytes()
        large_batch_module._verify_canonical_csv_projection(csv_bytes, rows, contract)
        csv_rows = list(
            csv.DictReader(io.StringIO(csv_bytes.decode("utf-8"), newline=""))
        )
        assert len(csv_rows) == len(rows)
        row_count += len(rows)
        for json_row, csv_row in zip(rows, csv_rows, strict=True):
            for field, value in json_row.items():
                if isinstance(value, str) and csv_row[field] != value:
                    assert field != "source_sheet"
                    assert csv_row[field] == value.rstrip()
                    difference_count += 1
                    differences.add(
                        (family_id, json_row["reference_date"], field, value)
                    )
    assert row_count == 76_189
    assert difference_count == 3_484
    assert len(differences) == 110
    assert {item[2] for item in differences} == {
        "raw_characteristic_category",
        "raw_indigenous_status",
        "raw_method_of_finalisation",
        "raw_sentence_statistic",
        "raw_statistic_basis",
    }
    payload = json.dumps(
        sorted(differences), separators=(",", ":"), ensure_ascii=False
    ).encode()
    assert sha256_digest(payload) == (
        "sha256:49f7bf9006282668731b9f87665fd3fd42e9714519cea7b3964d987d397a2f32"
    )


def test_registered_fdv_authority_and_aggregate_digests_are_literal() -> None:
    fixture_root = PROJECT / "fixtures/product-prototype"
    families = _fdv_offence_families()
    family_ids = [family["familyId"] for family in families]
    registry = load_large_batch_registry(PROJECT)
    specs = [spec for spec in registry.entries if spec.family_id in family_ids]
    assert len(specs) == 31
    assert sum(len(spec.expected_years) for spec in specs) == 56
    assert sum(spec.expected_canonical_count for spec in specs) == 76_189
    assert all(
        spec.acceptance_policy_version == "tidy.table-family-acceptance/v2"
        and spec.replay_recorded_at == "2026-08-25T09:00:00+00:00"
        and spec.expected_excluded_observation_count == 0
        for spec in specs
    )

    contracts = [
        fixture_root / "acceptance" / f"{family_id}-v1.json" for family_id in family_ids
    ]
    cohorts = [fixture_root / f"{family_id}.json" for family_id in family_ids]

    def path_manifest_digest(paths: list[Path]) -> str:
        payload = "".join(
            f"{path.relative_to(PROJECT).as_posix()} "
            f"{sha256_digest(path.read_bytes())}\n"
            for path in sorted(paths)
        ).encode()
        return sha256_digest(payload)

    assert path_manifest_digest(contracts) == (
        "sha256:c4ce08444432078619eea0e9bc958a1b172912695da9e83477d84691c62b352e"
    )
    assert path_manifest_digest(cohorts) == (
        "sha256:97787b28379ac7b4e1d452536ff05144fdbb6564694bdc45fe9ce5d0a5283050"
    )

    map_paths: list[Path] = []
    decision_ids: list[list[str]] = []
    evidence_paths: list[Path] = []
    rows: list[dict[str, object]] = []
    warning_count = 0
    recipe_pin_count = 0
    for family_id, contract_path, cohort_path in zip(
        family_ids, contracts, cohorts, strict=True
    ):
        contract = json.loads(contract_path.read_text())
        cohort = json.loads(cohort_path.read_text())
        assert (
            contract["decisionIdentityVersion"] == V2_REFERENCE_DATE_DECISION_IDENTITY
        )
        recipe_pin_count += len(contract["expectedRecipeDigestsByYear"])
        map_paths.extend(
            fixture_root / workbook["replayResponse"]["path"]
            for workbook in cohort["workbooks"]
        )
        evidence_root = fixture_root / f"{family_id}-evidence"
        evidence_paths.extend(sorted(evidence_root.iterdir()))
        run = json.loads((evidence_root / "run.json").read_text())
        assert [item["referenceDate"] for item in run["workbooks"]] == [
            item["referenceDate"] for item in cohort["workbooks"]
        ]
        decision_ids.append([item["decisionId"] for item in run["workbooks"]])
        manifest = verify_large_batch_evidence(
            PROJECT, next(spec for spec in specs if spec.family_id == family_id)
        )
        warning_count += sum(manifest["warningCountsByYear"].values())
        rows.extend(
            json.loads((evidence_root / "canonical-observations.json").read_text())
        )
    assert recipe_pin_count == 56
    assert len(map_paths) == 56
    assert path_manifest_digest(map_paths) == (
        "sha256:63618d1292eaa0fe686613c62093f1f85f61978c5fac36e300cee987ef1f3a04"
    )
    assert sha256_digest(b"".join(path.read_bytes() for path in sorted(map_paths))) == (
        "sha256:73fc3d77d30039e5cd9f1b40afa69dd75839518cecc4ca37ad94cd74abd4b29b"
    )
    assert (
        sha256_digest(
            json.dumps(
                decision_ids, separators=(",", ":"), ensure_ascii=False, sort_keys=True
            ).encode()
        )
        == "sha256:6a71922d3640465c562217ca6a7c863b3db151ba49181f75f541aa6081c82479"
    )
    assert len(evidence_paths) == 217
    assert path_manifest_digest(evidence_paths) == (
        "sha256:cfa8418988c1eef774b812a50679779c87926a6323ffcd6addd09338a6461105"
    )

    assert len(rows) == 76_189
    assert Counter(row["measure_id"] for row in rows) == {
        "defendant-count": 67_425,
        "mean-age": 2_331,
        "mean-duration": 2_051,
        "median-age": 2_331,
        "median-duration": 2_051,
    }
    assert Counter(row["value_status"] for row in rows) == {
        "not_applicable": 118,
        "not_available": 633,
        "observed": 75_320,
        "suppressed": 118,
    }
    assert Counter(
        row["raw_value"] for row in rows if row["value_status"] != "observed"
    ) == {"..": 118, "na": 633, "np": 118}
    assert (
        sum(
            row["value_status"] == "observed"
            and row["value"] == 0
            and not isinstance(row["value"], bool)
            for row in rows
        )
        == 10_152
    )
    assert warning_count == 53_808
    assert {row["classification_context_id"] for row in rows} == {
        "FDV_EXPERIMENTAL_ANZSOC_2011",
        "FDV_EXPERIMENTAL_ANZSOC_2023",
        "FDV_EXPERIMENTAL_MIXED_CONCORDED_ANZSOC_2011_AND_ANZSOC_2023",
        "FDV_EXPERIMENTAL_RELEASE_2021_22",
        "FDV_EXPERIMENTAL_RELEASE_2022_23",
        "FDV_EXPERIMENTAL_RELEASE_2023_24",
        "FDV_EXPERIMENTAL_RELEASE_2024_25",
    }
    raw_keys = sorted(
        {key for row in rows for key in row if key.startswith("raw_")} - {"raw_value"}
    )
    raw_tuples = [
        (
            row["source_workbook_digest"],
            row["source_sheet"],
            row["measure_id"],
            row["statistic_basis_id"],
            row["unit_id"],
            *(row.get(key) for key in raw_keys),
        )
        for row in rows
    ]
    assert len(raw_tuples) == len(set(raw_tuples)) == 76_189


@pytest.mark.timeout(300)
def test_registered_fdv_complete_seven_file_bundles_verify(tmp_path: Path) -> None:
    family_ids = {family["familyId"] for family in _fdv_offence_families()}
    specs = [
        spec
        for spec in load_large_batch_registry(PROJECT).entries
        if spec.family_id in family_ids
    ]
    assert len(specs) == 31
    compared_files = 0
    for spec in specs:
        checked = (PROJECT / spec.evidence_manifest_path).parent
        generated = tmp_path / spec.family_id
        shutil.copytree(checked, generated)
        verify_large_batch_complete_reproduction(PROJECT, spec, generated)
        compared_files += len(list(generated.iterdir()))
    assert compared_files == 217


@pytest.mark.timeout(180)
def test_large_batch_cli_verifies_committed_evidence() -> None:
    completed = subprocess.run(
        [str(PROJECT / "scripts/tidy-prototype-batch"), "verify"],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=120,
    )
    assert completed.returncode == 0, completed.stderr
    report = json.loads(completed.stdout)
    assert report == {
        "batchId": "justice-five-hundred-fifty-nine-worksheets-v1",
        "worksheetCount": 559,
        "cohortCount": 220,
        "canonicalObservationCount": 478120,
        "providerCalls": 0,
        "verified": True,
    }
