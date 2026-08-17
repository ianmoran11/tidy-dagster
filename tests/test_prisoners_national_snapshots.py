from __future__ import annotations

import json
import re
import subprocess
from collections import Counter
from pathlib import Path

import openpyxl

from tidy_orchestrator.prisoners_release import semantic_cells
from tidy_orchestrator.product_prototype import _map_alias

PROJECT = Path(__file__).parents[1]
FIX = PROJECT / "fixtures/product-prototype"
FAMILIES = {
    "national-selected-characteristics-by-offence-charge": [450, 450, 449, 449, 538],
    "national-age-by-sex": [117, 117, 117, 117, 117],
    "national-sex-offence-charge-by-indigenous-status": [306, 306, 306, 306, 306],
    "national-age-by-offence-charge": [221, 221, 221, 221, 221],
    "national-country-of-birth-by-offence-charge": [221, 238, 221, 221, 238],
}


def _load(path: Path) -> dict[str, object] | list[dict[str, object]]:
    return json.loads(path.read_text())


def test_reviewed_source_geometry_cardinality_and_markers() -> None:
    expected_markers = {
        2021: Counter({None: 36}),
        2022: Counter({"n.a.": 36}),
        2023: Counter({"n.a.": 36, "n.p.": 1}),
        2024: Counter({"na": 36, "np": 1}),
        2025: Counter({"na": 36, "np": 2}),
    }
    expected_numeric = {
        "age": [117] * 5,
        "status": [306] * 5,
        "age_offence": [221] * 5,
        "country": [221, 238, 221, 221, 238],
    }
    for year in range(2021, 2026):
        workbook = openpyxl.load_workbook(
            FIX / "workbooks" / f"prisoners-australia-{year}-national-source.xlsx",
            data_only=False,
            read_only=False,
        )
        table1 = workbook["Table_1" if year < 2024 else "Table 1"]
        table1_rows = (
            [
                7,
                *range(9, 11),
                *range(12, 15),
                *range(16, 19),
                *range(20, 22),
                *range(23, 28),
                *range(29, 34),
                *range(36, 38),
                *range(39, 42),
                *range(43, 46),
                *range(47, 50),
            ]
            if year == 2025
            else [
                7,
                *range(9, 11),
                *range(12, 15),
                *range(16, 19),
                *range(20, 22),
                *range(23, 28),
                *range(30, 32),
                *range(33, 36),
                *range(37, 40),
                *range(41, 44),
            ]
        )
        values = [
            table1.cell(row, column).value
            for row in table1_rows
            for column in range(2, 20)
        ]
        markers = Counter(
            value for value in values if not isinstance(value, int | float)
        )
        assert markers == expected_markers[year]
        assert (
            sum(isinstance(value, int | float) for value in values)
            == FAMILIES["national-selected-characteristics-by-offence-charge"][
                year - 2021
            ]
        )

        physical = {
            "age": (4, [(7, 19)], 10),
            "status": (5 if year < 2025 else 6, [(8, 24), (26, 42), (44, 60)], 7),
            "age_offence": (6 if year < 2025 else 7, [(6, 18)], 18),
            "country": (
                7 if year < 2025 else 8,
                [(6, 19 if year in {2022, 2025} else 18)],
                18,
            ),
        }
        for family, (table, bands, last_column) in physical.items():
            sheet = workbook[f"Table_{table}" if year < 2024 else f"Table {table}"]
            body = [
                sheet.cell(row, column).value
                for first, last in bands
                for row in range(first, last + 1)
                for column in range(2, last_column + 1)
            ]
            assert len(body) == expected_numeric[family][year - 2021]
            assert all(
                isinstance(value, int | float)
                or (isinstance(value, str) and value.startswith("="))
                for value in body
            )
        workbook.close()


def test_generator_geometry_and_acceptance_inputs_are_independent() -> None:
    completed = subprocess.run(
        [
            "uv",
            "run",
            "python",
            "scripts/generate-prisoners-national-snapshots.py",
            "--check",
        ],
        cwd=PROJECT,
        capture_output=True,
        text=True,
        timeout=30,
    )
    assert completed.returncode == 0, completed.stderr
    geometry = _load(FIX / "prisoners-national-snapshots-geometry-v1.json")
    audit = _load(FIX / "prisoners-national-snapshots-acceptance-audit-v1.json")
    assert geometry["authority"] == "human-authored-reviewed-physical-geometry"
    assert audit["authority"] == "independent-of-replay-output"
    assert audit["expectedCanonicalCount"] == 6695
    assert sum(item["expectedCanonicalCount"] for item in audit["families"]) == 6695
    assert sum(len(item["members"]) for item in geometry["families"]) == 25
    assert {
        item["familyId"]: item["expectedYearCounts"] for item in audit["families"]
    } == FAMILIES
    assert all(item["ANZSOCVersion"] == "2011" for item in audit["families"])


def test_all_five_evidence_closures_have_exact_reviewed_counts_and_source_cells() -> (
    None
):
    for family, expected in FAMILIES.items():
        cohort = _load(FIX / f"prisoners-{family}.json")
        run = _load(FIX / f"{family}-five-year-evidence/run.json")
        rows = _load(FIX / f"{family}-five-year-evidence/canonical-observations.json")
        assert run["providerCalls"] == 0
        assert run["acceptedWorkbookCount"] == 5
        assert run["exceptionWorkbookCount"] == 0
        assert run["crossYearIssues"] == []
        assert [item["observationCount"] for item in run["workbooks"]] == expected
        assert len(rows) == sum(expected)
        for workbook in cohort["workbooks"]:
            year_rows = [
                row
                for row in rows
                if row["source_workbook_digest"] == workbook["contentDigest"]
                and row["source_sheet"] == workbook["sheet"]
            ]
            assert len({row["source_cell"] for row in year_rows}) == len(year_rows)
            assert all(
                re.fullmatch(r"R[1-9][0-9]*C[1-9][0-9]*", row["source_cell"])
                for row in year_rows
            )


def test_table_1_measure_sections_and_exact_missing_markers() -> None:
    family = "national-selected-characteristics-by-offence-charge"
    rows = _load(FIX / f"{family}-five-year-evidence/canonical-observations.json")
    run = _load(FIX / f"{family}-five-year-evidence/run.json")
    contract = _load(FIX / f"acceptance/prisoners-{family}-v1.json")
    assert [item["rawObservationCount"] for item in run["workbooks"]] == [
        450,
        486,
        486,
        486,
        576,
    ]
    assert [item["excludedObservationCount"] for item in run["workbooks"]] == [
        0,
        36,
        37,
        37,
        38,
    ]
    assert Counter(row["measure_id"] for row in rows) == {
        "prisoner-count": 900,
        "prisoner-proportion": 900,
        "median-age": 447,
        "mean-age": 89,
    }
    assert all(
        row["value_status"] == "observed" and row["value"] is not None for row in rows
    )
    assert contract["strictAliasMatching"] is True
    assert contract["totalEquations"] == []
    assert contract["totalValidation"] == "not_applicable"
    for measure in contract["measures"]:
        assert measure["missingValues"] == {
            "n.a.": "not_applicable",
            "n.p.": "suppressed",
            "na": "not_applicable",
            "np": "suppressed",
        }
        assert measure["excludeMissingValues"] is True
    assert (
        _map_alias(contract, "characteristic_group", "Mean age (years)") == "MEAN_AGE"
    )
    assert _map_alias(contract, "characteristic_group", "Mean age (years) (z)") is None


def test_2022_formula_results_preserve_exact_source_coordinates() -> None:
    age_rows = _load(
        FIX / "national-age-by-sex-five-year-evidence/canonical-observations.json"
    )
    status_rows = _load(
        FIX
        / "national-sex-offence-charge-by-indigenous-status-five-year-evidence"
        / "canonical-observations.json"
    )
    age_digest = (
        "sha256:0b6f14ccf90fded702181e26e023262afbb01016986c630e8bf27648fd7828be"
    )
    formula_age = {"R19C3", "R19C6", "R19C9"}
    formula_status = {
        f"R{row}C{column}" for row in (24, 42, 60) for column in (3, 5, 7)
    }
    assert {
        row["source_cell"]
        for row in age_rows
        if row["source_workbook_digest"] == age_digest
        and row["source_cell"] in formula_age
        and row["value"] == 100
    } == formula_age
    assert {
        row["source_cell"]
        for row in status_rows
        if row["source_workbook_digest"] == age_digest
        and row["source_cell"] in formula_status
        and row["value"] == 100
    } == formula_status


def test_full_width_2025_tables_are_bounded_without_payload_change() -> None:
    workbooks = FIX / "workbooks"
    source = workbooks / "prisoners-australia-2025-national-source.xlsx"
    bounded = workbooks / "prisoners-australia-2025-national-snapshots-bounded.xlsx"
    assert semantic_cells(source) == semantic_cells(bounded)
    geometry = _load(FIX / "prisoners-national-snapshots-geometry-v1.json")
    assert geometry["boundedNormalization"]["sheets"] == {
        "Table 7": "A1:R25",
        "Table 8": "A1:R28",
    }
    target_members = [
        member
        for family in geometry["families"]
        for member in family["members"]
        if member["year"] == 2025 and member["sheet"] in {"Table 7", "Table 8"}
    ]
    assert {item["semanticMaximumColumn"] for item in target_members} == {"R"}


def test_replay_names_use_collision_safe_physical_table_identity() -> None:
    paths = {
        workbook["replayResponse"]["path"]
        for family in FAMILIES
        for workbook in _load(FIX / f"prisoners-{family}.json")["workbooks"]
    }
    assert len(paths) == 25
    assert all(
        re.fullmatch(
            r"replay/prisoners-australia-national-table-(?:1|4|5|6|7|8)-20(?:21|22|23|24|25)\.response\.txt",
            path,
        )
        for path in paths
    )
    assert not list((FIX / "replay").glob("prisoners-national-national-*.response.txt"))


def test_total_equations_are_measure_scoped_to_counts() -> None:
    for family in FAMILIES:
        contract = _load(FIX / f"acceptance/prisoners-{family}-v1.json")
        for equation in contract["totalEquations"]:
            assert equation["measureIds"] == ["prisoner-count"]
        assert not any(
            measure in equation.get("measureIds", [])
            for equation in contract["totalEquations"]
            for measure in (
                "prisoner-proportion",
                "imprisonment-rate",
                "mean-age",
                "median-age",
            )
        )
