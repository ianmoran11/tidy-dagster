from __future__ import annotations

import copy
import importlib.util
import json
from pathlib import Path
from types import ModuleType

import openpyxl
import pytest

from tidy_orchestrator.artifacts import sha256_digest
from tidy_orchestrator.prisoners_release import (
    PrisonersReleaseError,
    _registered_members,
    _validate_atomic_remaining_registration,
    semantic_cells,
)
from tidy_orchestrator.product_prototype import (
    ProductPrototypeError,
    _validate_cohort,
    _validate_contract,
)

PROJECT = Path(__file__).resolve().parents[1]
FIX = PROJECT / "fixtures" / "product-prototype"


def _load_registration_script() -> ModuleType:
    path = PROJECT / "scripts" / "register-prisoners-remaining.py"
    spec = importlib.util.spec_from_file_location("register_prisoners_remaining", path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _tree_snapshot(root: Path) -> dict[str, tuple[str, bytes]]:
    return {
        item.relative_to(root).as_posix(): (
            ("directory", b"") if item.is_dir() else ("file", item.read_bytes())
        )
        for item in sorted(root.rglob("*"))
    }


def _load(path: Path) -> dict:
    return json.loads(path.read_text())


def test_remaining_prisoners_campaign_is_exact_and_v2_bound() -> None:
    membership = _load(FIX / "prisoners-release-family-membership-v1.json")
    registered = _registered_members(PROJECT, membership)
    source_members = {
        (member["year"], member["downloadOrdinal"], member["sheet"]): (
            family["familyId"],
            member,
        )
        for family in membership["families"]
        for member in family["members"]
    }
    pending = {
        key: value for key, value in source_members.items() if key not in registered
    }
    assert len(registered) == 203
    assert pending == {}

    plan = _load(FIX / "prisoners-remaining-semantic-map-plan-v1.json")
    assert plan["acceptanceAuthority"] is False
    assert plan["trainingEligibility"] is False
    assert len(plan["families"]) == 21
    planned = {
        (
            member["year"],
            next(
                source["downloadOrdinal"]
                for source in membership["families"]
                if source["familyId"] == family["familyId"]
                for source in source["members"]
                if source["year"] == member["year"]
                and source["sheet"] == member["sheet"]
            ),
            member["sheet"],
        )
        for family in plan["families"]
        for member in family["members"]
    }
    assert len(planned) == 69
    assert planned == {
        key
        for key, (family_id, _member) in source_members.items()
        if family_id in {family["familyId"] for family in plan["families"]}
    }
    assert planned <= registered

    for family in plan["families"]:
        family_id = family["familyId"]
        cohort = _load(FIX / f"prisoners-{family_id}.json")
        contract_path = FIX / "acceptance" / f"prisoners-{family_id}-v1.json"
        contract = _load(contract_path)
        _validate_cohort(cohort)
        _validate_contract(contract, cohort)
        assert contract["schemaVersion"] == "tidy.table-family-acceptance/v2"
        assert contract["decisionIdentityVersion"] == "v2-reference-date-v1"
        assert contract["trainingEligibility"] is False
        assert contract["automaticAcceptance"] is True
        assert contract["strictAliasMatching"] is True
        assert contract["totalEquations"] == []
        assert contract["totalValidation"] == "not_applicable"
        years = {str(workbook["year"]) for workbook in cohort["workbooks"]}
        assert set(contract["expectedRecipeDigestsByYear"]) == years
        assert all(
            digest.startswith("sha256:") and len(digest) == 71
            for digest in contract["expectedRecipeDigestsByYear"].values()
        )
        assert contract["expectedWarningCountsByYear"] == {year: 0 for year in years}
        for workbook in cohort["workbooks"]:
            replay = workbook["replayResponse"]
            assert replay["acceptanceAuthority"] is False
            replay_bytes = (FIX / replay["path"]).read_bytes()
            assert replay["contentDigest"] == sha256_digest(replay_bytes)
            assert replay["byteLength"] == len(replay_bytes)


def test_fixed_scratch_reproduction_restores_original_tree_on_failure(
    tmp_path: Path,
) -> None:
    registration = _load_registration_script()
    scratch = tmp_path / "scratch"
    scratch.mkdir()
    (scratch / "family-a").mkdir()
    (scratch / "family-a" / "run.json").write_bytes(b"original-run\n")
    (scratch / "root.bin").write_bytes(b"original-root")
    before = _tree_snapshot(scratch)

    def fail_after_partial_reproduction(empty_scratch: Path) -> None:
        assert not empty_scratch.exists()
        (empty_scratch / "family-new").mkdir(parents=True)
        (empty_scratch / "family-new" / "partial.json").write_bytes(b"partial")
        raise OSError("injected reproduction failure")

    with pytest.raises(OSError, match="injected reproduction failure"):
        registration._with_fixed_scratch_reproduction(
            scratch, fail_after_partial_reproduction
        )

    assert _tree_snapshot(scratch) == before
    assert not list(tmp_path.glob(".scratch-reproduction-backup-*"))


def test_registration_title_preserves_anzsoc_capitalization() -> None:
    registration = _load_registration_script()
    assert (
        registration._title("preliminary-anzsoc-2023-table-1")
        == "Preliminary ANZSOC 2023 table 1"
    )


@pytest.mark.parametrize(
    ("failure_point", "failure_number"),
    [
        ("evidence-swap", 11),
        ("registry-write", 2),
        ("post-swap-validation", 1),
    ],
)
def test_registration_writer_rolls_back_every_output_on_injected_failure(
    tmp_path: Path, failure_point: str, failure_number: int
) -> None:
    registration = _load_registration_script()
    installed = tmp_path / "installed"
    staged = tmp_path / "staged"
    installed.mkdir()
    staged.mkdir()
    replacements = []

    for index in range(21):
        destination = installed / f"evidence-{index:02d}"
        if index != 20:
            destination.mkdir()
            (destination / "nested").mkdir()
            (destination / "nested" / "evidence.bin").write_bytes(
                f"original-evidence-{index}".encode()
            )
        source = staged / f"evidence-{index:02d}"
        source.mkdir()
        (source / "registered.bin").write_bytes(f"new-evidence-{index}".encode())
        replacements.append(("evidence-swap", source, destination))

    for index, name in enumerate(("normalization.json", "large.json", "status.json")):
        destination = installed / name
        if index != 2:
            destination.write_bytes(f"original-registry-{index}".encode())
        source = staged / name
        source.write_bytes(f"new-registry-{index}".encode())
        replacements.append(("registry-write", source, destination))

    before = _tree_snapshot(installed)
    calls = 0
    validated = False

    def fail_at_point(point: str, _destination: Path) -> None:
        nonlocal calls
        if point == failure_point:
            calls += 1
            if calls == failure_number:
                raise OSError(f"injected {point} failure")

    def validate() -> None:
        nonlocal validated
        validated = True

    with pytest.raises(OSError, match=f"injected {failure_point} failure"):
        registration._transactional_replace(
            replacements,
            staged / "backups",
            validate,
            fail_at_point,
        )

    assert _tree_snapshot(installed) == before
    assert not (staged / "backups").exists()
    assert validated is (failure_point == "post-swap-validation")


def test_remaining_prisoners_registration_and_evidence_are_atomic() -> None:
    membership = _load(FIX / "prisoners-release-family-membership-v1.json")
    registered = _registered_members(PROJECT, membership)
    plan = _load(FIX / "prisoners-remaining-semantic-map-plan-v1.json")
    family_ids = {family["familyId"] for family in plan["families"]}
    cohort_ids = {f"prisoners-australia-{family_id}" for family_id in family_ids}

    large = _load(FIX / "large-batch-assets-v1.json")
    status = _load(FIX / "data-asset-status-v1.json")
    assert {item["familyId"] for item in large["entries"]} >= family_ids
    assert {item["cohortId"] for item in status["cohorts"]} >= cohort_ids
    for family_id in family_ids:
        evidence = FIX / f"prisoners-{family_id}-evidence"
        assert {item.name for item in evidence.iterdir()} == {
            "README.md",
            "canonical-observations.csv",
            "canonical-observations.json",
            "collation-report.json",
            "exceptions.json",
            "manifest.json",
            "run.json",
        }

    _validate_atomic_remaining_registration(PROJECT, membership, registered)
    campaign_keys = {
        (
            member["year"],
            member["downloadOrdinal"],
            member["sheet"],
        )
        for family in membership["families"]
        if family["familyId"] in family_ids
        for member in family["members"]
    }
    partial = registered - {next(iter(campaign_keys))}
    with pytest.raises(PrisonersReleaseError, match="must be atomic"):
        _validate_atomic_remaining_registration(PROJECT, membership, partial)


def test_remaining_prisoners_special_semantics_and_worker_limits() -> None:
    table_five = _load(
        FIX / "acceptance" / "prisoners-preliminary-anzsoc-2023-table-5-v1.json"
    )
    assert table_five["requiredDimensions"] == [
        "principal_offence",
        "principal_offence_anzsoc_2011",
        "jurisdiction",
    ]
    negative = _load(
        FIX
        / "acceptance"
        / "prisoners-national-selected-characteristics-time-series-v1.json"
    )["measures"][0]
    assert negative["minimum"] == -10
    assert negative["allowNegative"] is True

    plan = _load(FIX / "prisoners-remaining-semantic-map-plan-v1.json")
    limited = {}
    for family in plan["families"]:
        cohort = _load(FIX / f"prisoners-{family['familyId']}.json")
        if "workerLimits" in cohort:
            limited[family["familyId"]] = cohort["workerLimits"]
    assert limited == {
        "preliminary-anzsoc-2023-table-5": {"maxWarnings": 100000},
        "state-indigenous-sex-age-standardised-rate-time-series": {
            "maxWarnings": 100000
        },
        "state-indigenous-sex-crude-rate-time-series": {"maxWarnings": 100000},
        "state-indigenous-sex-prisoner-count-time-series": {"maxWarnings": 100000},
    }


@pytest.mark.parametrize("allow_negative", [False, "true", 1, None])
def test_negative_measure_requires_literal_true_opt_in(allow_negative: object) -> None:
    family_id = "national-selected-characteristics-time-series"
    cohort = _load(FIX / f"prisoners-{family_id}.json")
    contract = _load(FIX / "acceptance" / f"prisoners-{family_id}-v1.json")
    invalid = copy.deepcopy(contract)
    invalid["measures"][0]["allowNegative"] = allow_negative

    with pytest.raises(ProductPrototypeError):
        _validate_contract(invalid, cohort)


def test_negative_measure_without_opt_in_fails_closed() -> None:
    family_id = "national-selected-characteristics-time-series"
    cohort = _load(FIX / f"prisoners-{family_id}.json")
    contract = _load(FIX / "acceptance" / f"prisoners-{family_id}-v1.json")
    invalid = copy.deepcopy(contract)
    invalid["measures"][0].pop("allowNegative")
    assert invalid["measures"][0]["minimum"] < 0

    with pytest.raises(ProductPrototypeError):
        _validate_contract(invalid, cohort)


def test_remaining_prisoners_bounded_workbooks_preserve_semantic_cells() -> None:
    pairs = [
        (
            "prisoners-australia-2025-national-source.xlsx",
            "prisoners-australia-2025-national-remaining-bounded.xlsx",
        ),
        (
            "prisoners-australia-2024-federal-source.xlsx",
            "prisoners-australia-2024-federal-remaining-bounded.xlsx",
        ),
        (
            "prisoners-australia-2025-federal-source.xlsx",
            "prisoners-australia-2025-federal-remaining-bounded.xlsx",
        ),
    ]
    for source, bounded in pairs:
        assert semantic_cells(FIX / "workbooks" / source) == semantic_cells(
            FIX / "workbooks" / bounded
        )


def _logical_workbook_differences(source: str, bounded: str) -> dict:
    source_book = openpyxl.load_workbook(FIX / "workbooks" / source)
    bounded_book = openpyxl.load_workbook(FIX / "workbooks" / bounded)
    try:
        assert source_book.sheetnames == bounded_book.sheetnames
        differences = {}
        for sheet_name in source_book.sheetnames:
            source_sheet = source_book[sheet_name]
            bounded_sheet = bounded_book[sheet_name]
            source_merges = {str(item) for item in source_sheet.merged_cells.ranges}
            bounded_merges = {str(item) for item in bounded_sheet.merged_cells.ranges}
            style_differences = {}
            for coordinate in set(source_sheet._cells) | set(bounded_sheet._cells):
                source_cell = source_sheet._cells.get(coordinate)
                bounded_cell = bounded_sheet._cells.get(coordinate)
                source_style = source_cell.style_id if source_cell is not None else None
                bounded_style = (
                    bounded_cell.style_id if bounded_cell is not None else None
                )
                if source_style != bounded_style:
                    style_differences[coordinate] = (source_style, bounded_style)
            source_geometry = (source_sheet.max_row, source_sheet.max_column)
            bounded_geometry = (bounded_sheet.max_row, bounded_sheet.max_column)
            if (
                source_geometry != bounded_geometry
                or source_merges != bounded_merges
                or style_differences
            ):
                differences[sheet_name] = {
                    "geometry": (source_geometry, bounded_geometry),
                    "removedMerges": source_merges - bounded_merges,
                    "addedMerges": bounded_merges - source_merges,
                    "styles": style_differences,
                }
        return differences
    finally:
        source_book.close()
        bounded_book.close()


def test_bounded_workbook_style_merge_and_geometry_diffs_are_exact() -> None:
    national = _logical_workbook_differences(
        "prisoners-australia-2025-national-source.xlsx",
        "prisoners-australia-2025-national-remaining-bounded.xlsx",
    )
    expected_national = {
        "Table 10": ((69, 16384), (69, 11), {66}, {"A66:XFD66"}),
        "Table 11": ((48, 16384), (48, 18), {43, 44}, {"A43:XFD43", "A44:XFD44"}),
        "Table 12": ((69, 16384), (69, 16), {62}, {"A62:XFD62"}),
        "Table 13": ((69, 16384), (69, 16), {62}, {"A62:XFD62"}),
        "Table 14": ((67, 16384), (67, 6), {63}, {"A63:XFD63"}),
    }
    assert set(national) == set(expected_national)
    for sheet_name, (
        source_geometry,
        bounded_geometry,
        rows,
        merges,
    ) in expected_national.items():
        difference = national[sheet_name]
        assert difference["geometry"] == (source_geometry, bounded_geometry)
        assert difference["removedMerges"] == merges
        assert difference["addedMerges"] == set()
        assert difference["styles"] == {
            (row, column): (0, None) for row in rows for column in range(2, 16385)
        }

    federal_2024 = _logical_workbook_differences(
        "prisoners-australia-2024-federal-source.xlsx",
        "prisoners-australia-2024-federal-remaining-bounded.xlsx",
    )
    assert set(federal_2024) == {"Table 38", "Table 39"}
    expected_2024_styles = {
        "Table 38": {7, 19, 31, 43, 55, 67, 79, 91, 103},
        "Table 39": {7, 18, 29, 40, 51, 62, 73},
    }
    expected_2024_geometry = {
        "Table 38": (114, 12),
        "Table 39": (83, 12),
    }
    for sheet_name, rows in expected_2024_styles.items():
        difference = federal_2024[sheet_name]
        geometry = expected_2024_geometry[sheet_name]
        assert difference["geometry"] == (geometry, geometry)
        assert difference["removedMerges"] == set()
        assert difference["addedMerges"] == set()
        assert difference["styles"] == {(row, 1): (36, 38) for row in rows}

    federal_2025 = _logical_workbook_differences(
        "prisoners-australia-2025-federal-source.xlsx",
        "prisoners-australia-2025-federal-remaining-bounded.xlsx",
    )
    assert set(federal_2025) == {"Table 39"}
    assert federal_2025["Table 39"] == {
        "geometry": ((87, 11), (87, 11)),
        "removedMerges": set(),
        "addedMerges": set(),
        "styles": {(row, 1): (46, 48) for row in {7, 18, 29, 40, 51, 62, 73}},
    }
